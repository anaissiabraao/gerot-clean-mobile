#!/usr/bin/env python3
from __future__ import annotations

from flask import (
    Flask,
    render_template,
    redirect,
    url_for,
    session,
    jsonify,
    request,
    flash,
    send_from_directory,
    g,
)
from flask_cors import CORS
from flask_compress import Compress
from flask_restful import Api, Resource
from dotenv import load_dotenv
import os
import time
import socket
import json
import gzip
from urllib.parse import urlparse

# Carregar variáveis de ambiente do arquivo .env
load_dotenv()

import secrets
import re
from pathlib import Path
import bcrypt
import psycopg2
import psycopg2.extras
import psycopg2.errors
from datetime import datetime, date, timedelta
from functools import wraps
from typing import Dict, List, Tuple
import mimetypes
from io import BytesIO
from werkzeug.utils import secure_filename
from psycopg2 import pool
import requests
from openpyxl import load_workbook

from utils.planner_client import PlannerClient, PlannerIntegrationError
from utils.library_reports import build_tabular_report, generate_report_pdf, render_report_html
from utils.indicadores_executivos import montar_indicadores_executivos

from automate.automation_catalog import get_automation, list_automations


app = Flask(__name__)
CORS(app)
Compress(app)  # Habilita compressão Gzip
api = Api(app)


@app.after_request
def add_header(response):
    if 'Cache-Control' not in response.headers:
        # Cache para arquivos estáticos (static/...)
        if request.path.startswith('/static/'):
            response.headers['Cache-Control'] = 'public, max-age=31536000'
        # Não cachear API ou HTML dinâmico para evitar dados velhos
        else:
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return response

# --------------------------------------------------------------------------- #
# Configuração base
# --------------------------------------------------------------------------- #
app.config["SECRET_KEY"] = os.getenv(
    "SECRET_KEY", "gerot-production-2025-super-secret"
)
app.config["DEBUG"] = os.getenv("FLASK_DEBUG", "false").lower() == "true"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=7)  # Sessões persistem por 7 dias

DATABASE_URL = (
    os.getenv("DATABASE_URL")
    or os.getenv("DIRECT_URL")
    or os.getenv("SUPABASE_DB_URL")
)

if not DATABASE_URL:
    raise RuntimeError(
        "DATABASE_URL não configurada. Defina a string de conexão do banco (PostgreSQL/Supabase) nas variáveis do serviço no Railway."
    )

app.config["DATABASE_URL"] = DATABASE_URL

RAG_API_URL = os.getenv("RAG_API_URL")
RAG_API_KEY = os.getenv("RAG_API_KEY", "local-dev")
if not RAG_API_URL:
    print("⚠️ AVISO: RAG_API_URL não configurada. Serviço RAG ficará indisponível.")
RAG_DEFAULT_TOP_K = int(os.getenv("RAG_TOP_K", "6"))
# Se true, usa o endpoint agentic do RAG (/v1/agent/qa) em vez do clássico (/v1/qa).
RAG_USE_AGENTIC = os.getenv("RAG_USE_AGENTIC", "true").lower() == "true"
# Em Railway, requests muito longos podem ser encerrados pelo proxy antes do Flask responder.
# Mantemos um default mais baixo; você pode aumentar via env se souber que o proxy aguenta.
RAG_TIMEOUT_SECONDS = int(os.getenv("RAG_TIMEOUT_SECONDS", "60"))
RAG_STRICT_MODE = os.getenv("RAG_STRICT_MODE", "true").lower() == "true"


def call_rag_service(question: str, *, top_k: int = None) -> dict:
    if not RAG_API_URL:
        raise RuntimeError("RAG_API_URL não configurada")

    payload = {
        "question": question,
        "top_k": top_k or RAG_DEFAULT_TOP_K,
    }
    headers = {"x-api-key": RAG_API_KEY}
    try:
        endpoint = "/v1/agent/qa" if RAG_USE_AGENTIC else "/v1/qa"
        response = requests.post(
            f"{RAG_API_URL.rstrip('/')}{endpoint}",
            json=payload,
            headers=headers,
            timeout=RAG_TIMEOUT_SECONDS,
        )
        response.raise_for_status()
        return response.json()
    except requests.HTTPError as http_err:
        app.logger.error("[RAG] Erro HTTP: %s - %s", http_err, http_err.response.text if http_err.response else "")
        raise
    except Exception as exc:
        app.logger.error("[RAG] Erro ao chamar serviço: %s", exc)
        raise


@app.route("/api/agent/rag/status", methods=["GET"])
def rag_status():
    """
    Diagnóstico rápido do RAG configurado no dashboard (Railway):
    - qual URL está ativa no processo atual
    - resolve DNS?
    - /health responde?
    """
    # Evita depender do decorator login_required aqui (ele é definido mais abaixo no arquivo).
    if not session.get("user_id"):
        return jsonify({"error": "Não autenticado"}), 401

    url = RAG_API_URL
    out = {
        "rag_api_url": url,
        "rag_timeout_seconds": RAG_TIMEOUT_SECONDS,
        "has_rag_api_key": bool(RAG_API_KEY),
        "time": datetime.utcnow().isoformat() + "Z",
    }
    if not url:
        return jsonify(out), 200

    try:
        parsed = urlparse(url)
        host = parsed.hostname
        out["host"] = host
        out["scheme"] = parsed.scheme
        out["port"] = parsed.port
        if host:
            try:
                addrs = socket.getaddrinfo(host, 443, proto=socket.IPPROTO_TCP)
                out["dns"] = sorted({a[4][0] for a in addrs})
            except Exception as e:
                out["dns_error"] = str(e)
    except Exception as e:
        out["parse_error"] = str(e)

    # Testar /health
    try:
        resp = requests.get(f"{url.rstrip('/')}/health", timeout=15)
        out["health_http_status"] = resp.status_code
        out["health_text_preview"] = (resp.text or "")[:300]
    except Exception as e:
        out["health_error"] = str(e)

    return jsonify(out), 200


@app.route("/api/agent/rag/test", methods=["GET"])
def rag_test():
    """
    Teste do lado do Railway: chama /v1/qa e devolve tempo + status.
    Útil para diagnosticar "Erro de conexão" no frontend (timeout/proxy vs RAG).
    """
    if not session.get("user_id"):
        return jsonify({"error": "Não autenticado"}), 401

    q = (request.args.get("q") or "qual a capital da alemanha?").strip()
    full = (request.args.get("full") or "").strip().lower() in {"1", "true", "yes", "y"}
    t0 = time.time()
    result = {"question": q, "rag_api_url": RAG_API_URL, "timeout_s": RAG_TIMEOUT_SECONDS, "full": full}
    if not RAG_API_URL:
        return jsonify({**result, "error": "RAG_API_URL não configurada"}), 400
    try:
        payload = {"question": q, "top_k": RAG_DEFAULT_TOP_K}
        headers = {"x-api-key": RAG_API_KEY}
        resp = requests.post(f"{RAG_API_URL.rstrip('/')}/v1/qa", json=payload, headers=headers, timeout=RAG_TIMEOUT_SECONDS)
        result["elapsed_ms"] = int((time.time() - t0) * 1000)
        result["http_status"] = resp.status_code
        # tenta json
        try:
            j = resp.json()
            ans = (j.get("answer") or "")
            result["answer_preview"] = ans[:300]
            result["answer_len"] = len(ans)
            if full:
                # evita payload gigante no endpoint de diagnóstico
                result["answer"] = ans[:4000]
            result["sources_count"] = len(j.get("sources") or [])
        except Exception:
            result["text_preview"] = (resp.text or "")[:300]
        return jsonify(result), 200 if resp.ok else 502
    except Exception as e:
        result["elapsed_ms"] = int((time.time() - t0) * 1000)
        result["error"] = str(e)
        return jsonify(result), 502


@app.route("/api/agent/rag/ingest-file", methods=["POST"])
def rag_ingest_file():
    """
    Proxy seguro para upload de arquivo (dashboard -> RAG via tunnel) sem expor o x-api-key no browser.

    Envia multipart/form-data para o RAG local: POST {RAG_API_URL}/v1/ingest-file
    """
    if not session.get("user_id"):
        return jsonify({"error": "Não autenticado"}), 401

    if not RAG_API_URL:
        return jsonify({"error": "RAG_API_URL não configurada"}), 400

    if "file" not in request.files:
        return jsonify({"error": "Arquivo não enviado (campo 'file')"}), 400

    up = request.files["file"]
    if not up or not getattr(up, "filename", ""):
        return jsonify({"error": "Arquivo inválido"}), 400

    source_name = (request.form.get("source_name") or "upload").strip()
    category = (request.form.get("category") or "Uploads").strip()

    files = {"file": (up.filename, up.stream, up.mimetype or "application/octet-stream")}
    data = {"source_name": source_name, "category": category}
    headers = {"x-api-key": RAG_API_KEY}

    try:
        resp = requests.post(
            f"{RAG_API_URL.rstrip('/')}/v1/ingest-file",
            files=files,
            data=data,
            headers=headers,
            timeout=min(180, max(30, int(RAG_TIMEOUT_SECONDS or 60))),
        )
        # retorna body do RAG (json ou texto)
        ct = (resp.headers.get("content-type") or "").lower()
        if "application/json" in ct:
            payload = resp.json()
        else:
            payload = {"text": (resp.text or "")[:2000]}
        return jsonify({"ok": resp.ok, "http_status": resp.status_code, "rag": payload}), 200 if resp.ok else 502
    except Exception as e:
        return jsonify({"error": str(e)}), 502


def format_rag_response(raw: dict) -> str:
    answer = raw.get("answer") or "Não foi possível obter resposta."
    sources = raw.get("sources") or []
    if not sources:
        return answer

    import html as _html
    import os as _os

    max_chars = int(_os.getenv("RAG_UI_SOURCE_SNIPPET_MAX_CHARS", "20000"))
    max_chars = max(500, min(max_chars, 20000))

    items_html: list[str] = []
    for idx, src in enumerate(sources, start=1):
        doc_id = src.get("document_id")
        chunk_idx = src.get("chunk_index")
        score = src.get("score", 0.0)
        snippet = (src.get("snippet") or "").strip()
        snippet = snippet.replace("\r\n", "\n").replace("\r", "\n")
        truncated = False
        if len(snippet) > max_chars:
            snippet = snippet[:max_chars]
            truncated = True

        snippet_escaped = _html.escape(snippet)
        meta_line = (
            f"<strong>Fonte {idx}</strong>: "
            f"doc:<code>{_html.escape(str(doc_id))}</code> "
            f"chunk:<code>{_html.escape(str(chunk_idx))}</code> "
            f"score:<code>{float(score):.3f}</code>"
        )
        trecho_label = "Trecho (completo)" if not truncated else f"Trecho (parcial, truncado em {max_chars} chars)"
        items_html.append(
            "<li>"
            + meta_line
            + f"<details style='margin-top:6px;'><summary>{trecho_label}</summary>"
            + "<pre style='white-space:pre-wrap;overflow:auto;padding:10px;border:1px solid #e5e7eb;border-radius:10px;background:#fafafa;'>"
            + snippet_escaped
            + "</pre></details></li>"
        )

    sources_html = (
        f"<details><summary><strong>Fontes ({len(sources)})</strong></summary>"
        "<div style='margin-top:8px;'>"
        "<ul style='margin-left:18px;list-style:disc;'>"
        + "".join(items_html)
        + "</ul></div></details>"
    )

    return answer + "\n\n" + sources_html

def _strip_pgbouncer_param(dsn: str) -> str:
    """
    Remove apenas o parâmetro `pgbouncer=` da query-string (quando existir).

    OBS: Mantemos `sslmode=require` e quaisquer outros parâmetros úteis.
    """
    if "?" not in dsn:
        return dsn
    base_url, query_params = dsn.split("?", 1)
    params = [p for p in query_params.split("&") if not p.startswith("pgbouncer=")]
    return f"{base_url}?{'&'.join(params)}" if params else base_url


# Configuração do Pool de Conexões
# - Preferimos DATABASE_URL (pode ser pooler).
# - Se falhar (ex.: SSL/pgbouncer), tentamos DIRECT_URL (host direto), se existir.
pool_dsn_candidates: list[str] = []
pool_dsn_candidates.append(_strip_pgbouncer_param(DATABASE_URL))
direct_url_env = os.getenv("DIRECT_URL")
if direct_url_env and direct_url_env.strip() and direct_url_env.strip() != DATABASE_URL:
    pool_dsn_candidates.append(_strip_pgbouncer_param(direct_url_env.strip()))

# Min: 1, Max: 60 conexões simultâneas (com retry para PostgreSQL que sobe após o app no Railway)
db_pool = None
last_pool_error: Exception | None = None
_pool_retries = int(os.getenv("DB_POOL_INIT_RETRIES", "5"))
_pool_delay = float(os.getenv("DB_POOL_INIT_DELAY_SECONDS", "2.0"))
for attempt in range(1, _pool_retries + 1):
    for pool_dsn in pool_dsn_candidates:
        try:
            db_pool = psycopg2.pool.ThreadedConnectionPool(
                minconn=1,
                maxconn=60,
                dsn=pool_dsn,
                cursor_factory=psycopg2.extras.RealDictCursor,
                keepalives=1,
                keepalives_idle=30,
                keepalives_interval=10,
                keepalives_count=5,
            )
            print("✅ Pool de conexões criado com sucesso (1-60 conexões)")
            break
        except Exception as e:
            last_pool_error = e
            db_pool = None
            continue
    if db_pool is not None:
        break
    if attempt < _pool_retries:
        print(f"⏳ Tentativa {attempt}/{_pool_retries} do pool falhou; aguardando {_pool_delay}s...")
        time.sleep(_pool_delay)

if db_pool is None and last_pool_error is not None:
    print(f"❌ Erro ao criar pool de conexões após {_pool_retries} tentativas: {last_pool_error}")
    print("   O app seguirá; get_db() tentará conexão direta a cada request (ou use DATABASE_URL correto).")


BASE_DIR = Path(__file__).resolve().parent
PLANILHA_USUARIOS = BASE_DIR / "dados.xlsx"
FRONTEND_DIST_DIR = BASE_DIR / "frontend" / "dist"
FRONTEND_APP_URL = (os.getenv("FRONTEND_APP_URL") or "").strip().rstrip("/")
ADMIN_CARGOS = {"CONSULTOR", "COORDENADOR", "DIRETOR"}
ALLOWED_AVATAR_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "gif"}
MAX_AVATAR_SIZE_BYTES = 5 * 1024 * 1024  # 5 MB

# Configuração para usar o novo tema Tailwind (True) ou o tema antigo (False)
USE_TAILWIND_THEME = os.getenv("USE_TAILWIND_THEME", "true").lower() == "true"

DEFAULT_DASHBOARDS = [
    {
        "slug": "comercial_sc",
        "title": "Comercial SC",
        "description": "Relatório Vendas PortoEx - unidade SC.",
        "category": "Comercial",
        "embed_url": "https://app.powerbi.com/view?r=eyJrIjoiNDAwZTA5YjgtZWVlMC00MzQ2LWJmYmQtYTZiZDVlMDhlZTEyIiwidCI6IjM4MjViNTlkLTY1ZGMtNDM1Zi04N2M4LTkyM2QzMzkxYzMyOCJ9",
        "display_order": 1,
    },
    {
        "slug": "comercial_sp",
        "title": "Comercial SP",
        "description": "Relatório Vendas PortoEx - filial São Paulo.",
        "category": "Comercial",
        "embed_url": "https://app.powerbi.com/view?r=eyJrIjoiYjMyZTc5MzktNGFhYi00ZjE1LWFjMDctYjY2ODM4NTlhMWRmIiwidCI6IjM4MjViNTlkLTY1ZGMtNDM1Zi04N2M4LTkyM2QzMzkxYzMyOCJ9",
        "display_order": 2,
    },
    {
        "slug": "controladoria_target",
        "title": "Controladoria e Target Operação",
        "description": "Visão consolidada da controladoria e operação.",
        "category": "Controladoria",
        "embed_url": "https://app.powerbi.com/view?r=eyJrIjoiMzhjODY4OGYtY2UxMy00ZjkyLTkzNDEtOTcxZWIzNDY2ZGJlIiwidCI6IjM4MjViNTlkLTY1ZGMtNDM1Zi04N2M4LTkyM2QzMzkxYzMyOCJ9",
        "display_order": 3,
    },
    {
        "slug": "cotacao_sc",
        "title": "Cotação SC",
        "description": "Painel de cotações da filial Santa Catarina.",
        "category": "Cotação",
        "embed_url": "https://app.powerbi.com/view?r=eyJrIjoiNDcwMmM0NGUtOGJkZC00NmIyLTk3M2QtOTNjOWEzMDY5MjkwIiwidCI6IjM4MjViNTlkLTY1ZGMtNDM1Zi04N2M4LTkyM2QzMzkxYzMyOCJ9",
        "display_order": 4,
    },
    {
        "slug": "cotacao_sp",
        "title": "Cotação SP",
        "description": "Painel de cotações da filial São Paulo.",
        "category": "Cotação",
        "embed_url": "https://app.powerbi.com/view?r=eyJrIjoiYmNkMWQxMjUtZjExNC00ZGE5LWIxYTEtYzlmNzI3M2I3Mjg1IiwidCI6IjM4MjViNTlkLTY1ZGMtNDM1Zi04N2M4LTkyM2QzMzkxYzMyOCJ9",
        "display_order": 5,
    },
    {
        "slug": "gr_leandro",
        "title": "Gestão Regional - Leandro",
        "description": "Indicadores táticos da regional do Leandro.",
        "category": "Operações",
        "embed_url": "https://app.powerbi.com/view?r=eyJrIjoiYjI4YTIzMDEtZmRmOC00N2Y3LTkzNmQtMGEwYzE1N2VhMDViIiwidCI6IjM4MjViNTlkLTY1ZGMtNDM1Zi04N2M4LTkyM2QzMzkxYzMyOCJ9",
        "display_order": 6,
    },
    {
        "slug": "financeiro_fluxo",
        "title": "Financeiro - Fluxo de Caixa",
        "description": "Inadimplência e fluxo de caixa diário.",
        "category": "Financeiro",
        "embed_url": "https://app.powerbi.com/view?r=eyJrIjoiNzFlZWVkZjUtNTdiYy00ZjJiLTk3OTEtNzhiYzFhMzk3MmY3IiwidCI6IjM4MjViNTlkLTY1ZGMtNDM1Zi04N2M4LTkyM2QzMzkxYzMyOCJ9",
        "display_order": 7,
    },
]

PLANNER_CONFIG = {
    "tenant_id": os.getenv("MS_TENANT_ID"),
    "client_id": os.getenv("MS_CLIENT_ID"),
    "client_secret": os.getenv("MS_CLIENT_SECRET"),
    "plan_id": os.getenv("MS_PLANNER_PLAN_ID"),
    "bucket_id": os.getenv("MS_PLANNER_BUCKET_ID"),
}

planner_client = PlannerClient(**PLANNER_CONFIG)


# --------------------------------------------------------------------------- #
# Decorators e utilidades
# --------------------------------------------------------------------------- #
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            flash("Por favor, faça login.", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "admin":
            flash("Acesso restrito aos administradores.", "error")
            return redirect(url_for("team_dashboard"))
        return f(*args, **kwargs)

    return decorated_function


def is_admin_session() -> bool:
    return session.get("role") == "admin"


def get_template(template_base_name: str) -> str:
    """Retorna o template correto baseado na configuração USE_TAILWIND_THEME."""
    if USE_TAILWIND_THEME:
        # Verifica se existe uma versão Tailwind do template
        tailwind_template = f"{template_base_name.replace('.html', '')}_tailwind.html"
        template_path = Path(BASE_DIR) / "templates" / tailwind_template
        if template_path.exists():
            return tailwind_template
    return template_base_name


def _as_bytes(value):
    if value is None:
        return None
    if isinstance(value, memoryview):
        return value.tobytes()
    if isinstance(value, str):
        return value.encode("utf-8")
    return value


def _sanitize_optional(value: str | None):
    if value is None:
        return None
    value = value.strip()
    return value or None


def is_allowed_avatar_file(filename: str) -> bool:
    return bool(filename and "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_AVATAR_EXTENSIONS)


def refresh_session_user_cache():
    """Recarrega dados básicos do usuário na sessão após atualização."""
    if "user_id" not in session:
        return None
    updated = get_user_by_id(session["user_id"])
    if updated:
        session["username"] = updated.get("username", session.get("username"))
        session["nome_completo"] = updated.get("nome_completo", session.get("nome_completo"))
        session["role"] = updated.get("role", session.get("role"))
        if updated.get("departamento") is not None:
            session["departamento"] = updated.get("departamento")
        if updated.get("avatar_url"):
            session["avatar_url"] = updated["avatar_url"]
    return updated


@app.before_request
def update_last_seen():
    if session.get('user_id'):
        # Ignora rotas estáticas
        if request.path.startswith('/static') or request.endpoint == 'static':
            return

        try:
            conn = get_db()
            cursor = conn.cursor()
            # Usar request.path para mostrar a URL real
            current_page = request.path
            
            cursor.execute("""
                UPDATE users_new 
                SET last_seen_at = CURRENT_TIMESTAMP, current_page = %s 
                WHERE id = %s
            """, (current_page, session['user_id']))
            conn.commit()
            conn.close()
        except Exception as e:
            # Logar erro para debug
            app.logger.error(f"Erro ao atualizar last_seen: {e}")


@app.route("/admin/live-users")
@login_required
@admin_required
def admin_live_users():
    conn = get_db()
    cursor = conn.cursor()
    # Buscar usuários ativos nos últimos 5 minutos
    cursor.execute("""
        SELECT id, nome_completo, username, role, last_seen_at, current_page,
        EXTRACT(EPOCH FROM (NOW() - last_seen_at)) as seconds_ago
        FROM users_new
        WHERE last_seen_at > NOW() - INTERVAL '5 minutes'
        ORDER BY last_seen_at DESC
    """)
    active_users = cursor.fetchall()
    
    conn.close()
    return render_template(get_template("admin_live_users.html"), users=active_users)


@app.route("/admin/debug-time")
@login_required
@admin_required
def debug_time():
    """Rota temporária para diagnosticar problemas de time/fuso"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Verifica hora do DB
    cursor.execute("SELECT NOW() as db_time, CURRENT_TIMESTAMP as db_timestamp, NOW() - INTERVAL '5 minutes' as cutoff")
    times = cursor.fetchone()
    
    # Verifica últimos updates reais
    cursor.execute("""
        SELECT username, last_seen_at, current_page,
        NOW() - last_seen_at as time_diff
        FROM users_new
        WHERE last_seen_at IS NOT NULL
        ORDER BY last_seen_at DESC
        LIMIT 5
    """)
    recent_users = cursor.fetchall()
    conn.close()
    
    return jsonify({
        "server_time_utc": str(datetime.utcnow()),
        "db_time": str(times['db_time']),
        "cutoff_5min": str(times['cutoff']),
        "recent_users": [
            {
                "username": r['username'],
                "last_seen": str(r['last_seen_at']),
                "page": r['current_page'],
                "diff": str(r['time_diff'])
            } for r in recent_users
        ]
    })


# Classe Wrapper para interceptar o close() sem modificar o objeto C do psycopg2
class ConnectionWrapper:
    def __init__(self, conn, from_pool=True):
        self._conn = conn
        self._from_pool = from_pool
    
    def close(self):
        # Ignora chamadas explícitas de close() no código legado
        pass
    
    def real_close(self):
        # Fecha de verdade (usado para conexões fora do pool)
        self._conn.close()
        
    def __getattr__(self, name):
        return getattr(self._conn, name)


def get_db():
    """Obtém conexão do pool ou cria nova se necessário."""
    if 'db_wrapper' not in g:
        # Tentar pegar do Pool
        if db_pool:
            try:
                conn = db_pool.getconn()
                g.db_wrapper = ConnectionWrapper(conn, from_pool=True)
                return g.db_wrapper
            except Exception as e:
                app.logger.error(f"[DB] Erro ao pegar do pool: {e}")
        
        # Fallback: conexão direta (sem pool)
        # - Preferir DIRECT_URL (evita problemas com pgbouncer/pooler).
        # - Caso não exista, usar DATABASE_URL.
        database_url = (os.getenv("DIRECT_URL") or app.config["DATABASE_URL"]).strip()
        database_url = _strip_pgbouncer_param(database_url)
        
        max_retries = 3
        last_error = None
        for attempt in range(max_retries):
            try:
                conn = psycopg2.connect(
                    database_url,
                    cursor_factory=psycopg2.extras.RealDictCursor,
                    keepalives=1, keepalives_idle=30, keepalives_interval=10, keepalives_count=5
                )
                g.db_wrapper = ConnectionWrapper(conn, from_pool=False)
                return g.db_wrapper
            except psycopg2.OperationalError as e:
                last_error = e
                time.sleep(1 * (attempt + 1))
        raise last_error

    return g.db_wrapper


@app.teardown_appcontext
def close_db(error):
    """Devolve a conexão ao pool ou fecha ao final da requisição."""
    wrapper = g.pop('db_wrapper', None)
    if wrapper is not None:
        conn = wrapper._conn
        if wrapper._from_pool and db_pool:
            try:
                db_pool.putconn(conn)
            except Exception as e:
                app.logger.error(f"[DB] Erro ao devolver ao pool: {e}")
        else:
            # Se não veio do pool, fecha de verdade
            try:
                conn.close()
            except:
                pass


def seed_asset_library() -> None:
    """Cria ativos básicos e sincroniza gráficos internos do legado."""
    conn = None
    try:
        conn = get_db()
        cursor = conn.cursor()

        core_assets = [
            {
                "nome": "Auditoria Fiscal (Legado)",
                "tipo": "interno",
                "categoria": "Fiscal",
                "descricao": "Painel consolidado da Auditoria Fiscal já validado pelos usuários.",
                "status": "ativo",
                "ordem_padrao": 10,
                "resource_url": "/agent/legacy?tab=auditoria",
                "config": {"available_chart_types": ["bar", "line", "pie", "table"]},
            },
            {
                "nome": "Relatório de Resultados (Legado)",
                "tipo": "interno",
                "categoria": "Controladoria",
                "descricao": "Relatório de Resultados com gráficos e indicadores financeiros.",
                "status": "ativo",
                "ordem_padrao": 20,
                "resource_url": "/agent/legacy?tab=relatorio",
                "config": {
                    "embed_mode": True,
                    "legacy_view": "relatorio_resultados",
                    "available_components": [
                        {"key": "card_operacoes", "label": "Card: Operações"},
                        {"key": "card_frete", "label": "Card: Frete"},
                        {"key": "indicador_frete", "label": "Indicador: Frete (Meta)"},
                        {"key": "card_valor_nf", "label": "Card: Valor NF"},
                        {"key": "card_resultado", "label": "Card: Resultado Líquido"},
                        {"key": "card_ticket_medio", "label": "Card: Ticket Médio"},
                        {"key": "card_impostos", "label": "Card: Total Impostos"},
                        {"key": "card_prejuizos", "label": "Card: Prejuízos"},
                        {"key": "card_margem", "label": "Card: Margem %"},
                        {"key": "insight_melhor_vendedor", "label": "Insight: Melhor Vendedor"},
                        {"key": "insight_pior_vendedor", "label": "Insight: Pior Vendedor"},
                        {"key": "insight_resultado_medio", "label": "Insight: Resultado Médio"},
                        {"key": "analise_financeira_card_receita", "label": "Card: Análise Financeira - Receita"},
                        {"key": "analise_financeira_card_custos", "label": "Card: Análise Financeira - Custos"},
                        {"key": "analise_financeira_card_ebtida", "label": "Card: Análise Financeira - EBTIDA"},
                        {"key": "analise_financeira_card_impostos", "label": "Card: Análise Financeira - Impostos"},
                        {"key": "analise_financeira_grafico_composicao", "label": "Gráfico: Análise Financeira - Composição"},
                        {"key": "analise_financeira_grafico_ebtida", "label": "Gráfico: Análise Financeira - EBTIDA por Vendedor"},
                        {"key": "analise_financeira_tabela", "label": "Tabela: Análise Financeira (Custos e Margens)"},
                        {"key": "grafico_resultado_vendedor", "label": "Gráfico: Resultado por Vendedor"},
                        {"key": "grafico_operacoes", "label": "Gráfico: Distribuição de Operações"},
                        {"key": "grafico_custos", "label": "Gráfico: Custos vs Receita"},
                        {"key": "grafico_margem", "label": "Gráfico: Margem de Lucro (%)"},
                        {"key": "grafico_velocimetro_margem", "label": "Velocímetro: Margem (%)"},
                        {"key": "grafico_velocimetro_ebtida", "label": "Velocímetro: EBTIDA (%)"},
                        {"key": "grafico_velocimetro_custos", "label": "Velocímetro: Custos (%)"},
                        {"key": "grafico_velocimetro_impostos", "label": "Velocímetro: Impostos (%)"},
                        {"key": "grafico_velocimetro_margem_global", "label": "Velocímetro: Margem (%) (Geral)"},
                        {"key": "ranking_vendedores", "label": "Ranking de Vendedores"},
                        {"key": "analise_agente_filtros", "label": "Análise por Agente: Filtros"},
                        {"key": "analise_agente_detalhe", "label": "Análise por Agente: Detalhe do Agente"},
                        {"key": "analise_agente_grafico_impacto", "label": "Análise por Agente: Gráfico de Impacto"},
                        {"key": "analise_agente_ranking", "label": "Análise por Agente: Ranking"},
                        {"key": "analise_agente_tabela", "label": "Análise por Agente: Tabela Completa"},
                        {"key": "dre_card_rol", "label": "Card DRE: ROL"},
                        {"key": "dre_card_impostos", "label": "Card DRE: Impostos"},
                        {"key": "dre_card_lucro", "label": "Card DRE: Lucro Líquido"},
                        {"key": "dre_card_taxa", "label": "Card DRE: Taxa de Impostos"},
                        {"key": "dre_breakdown_cf", "label": "DRE: Contribuição Social (C.F)"},
                        {"key": "dre_breakdown_if", "label": "DRE: Imposto de Renda (I.F)"},
                        {"key": "dre_breakdown_outros", "label": "DRE: Outros Impostos"},
                        {"key": "dre_grafico_composicao", "label": "Gráfico DRE: Composição"},
                        {"key": "dre_grafico_impostos", "label": "Gráfico DRE: Breakdown de Impostos"},
                        {"key": "dre_grafico_evolucao", "label": "Gráfico DRE: Evolução"},
                        {"key": "ranking_clientes", "label": "Ranking de Clientes"},
                        {"key": "top_clientes", "label": "Top Clientes (Farol Verde)"},
                        {"key": "clientes_risco", "label": "Clientes em Risco (Farol Vermelho)"},
                        {"key": "analise_churn", "label": "Análise de Churn - Morte Silenciosa"},
                        {"key": "block_grafico_entregas", "label": "Gráfico: Relatório de Entregas por Status (Script 376)"},
                    ],
                },
            },
            {
                "nome": "RPA Selenium • Relatório Completo",
                "tipo": "rpa",
                "categoria": "Automação",
                "descricao": "Execução do RPA Selenium (relatório completo Brudam).",
                "status": "ativo",
                "ordem_padrao": 30,
                "resource_url": "/agent/legacy?tab=rpa",
            },
        ]

        for asset in core_assets:
            cursor.execute(
                """
                SELECT id FROM assets
                WHERE nome = %s AND tipo = %s
                  AND COALESCE(resource_url, '') = COALESCE(%s, '')
                  AND COALESCE(embed_url, '') = COALESCE(%s, '')
                """,
                (
                    asset["nome"],
                    asset["tipo"],
                    asset.get("resource_url"),
                    asset.get("embed_url"),
                ),
            )
            existing = cursor.fetchone()
            if existing:
                if asset.get("config"):
                    cursor.execute(
                        """
                        UPDATE assets
                        SET config = %s, updated_at = CURRENT_TIMESTAMP
                        WHERE id = %s
                        """,
                        (psycopg2.extras.Json(asset.get("config")), existing["id"]),
                    )
                continue
            cursor.execute(
                """
                INSERT INTO assets (
                    nome, tipo, categoria, descricao, status, ordem_padrao, resource_url, embed_url, config
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    asset["nome"],
                    asset["tipo"],
                    asset.get("categoria"),
                    asset.get("descricao"),
                    asset.get("status", "ativo"),
                    asset.get("ordem_padrao", 0),
                    asset.get("resource_url"),
                    asset.get("embed_url"),
                    psycopg2.extras.Json(asset.get("config")) if asset.get("config") else None,
                ),
            )

        cursor.execute(
            """
            SELECT title, description, category, embed_url, display_order
            FROM dashboards
            WHERE is_active = true
            """
        )
        for row in cursor.fetchall():
            embed_url = row.get("embed_url")
            if not embed_url:
                continue
            cursor.execute(
                "SELECT id FROM assets WHERE tipo = 'PBI' AND embed_url = %s",
                (embed_url,),
            )
            if cursor.fetchone():
                continue
            cursor.execute(
                """
                INSERT INTO assets (
                    nome, tipo, categoria, descricao, status, ordem_padrao, embed_url
                ) VALUES (%s, %s, %s, %s, 'ativo', %s, %s)
                """,
                (
                    row.get("title"),
                    "PBI",
                    row.get("category"),
                    row.get("description"),
                    row.get("display_order", 0),
                    embed_url,
                ),
            )

        try:
            cursor.execute(
                """
                SELECT id, title, description, category
                FROM agent_dashboard_templates
                WHERE is_published = true
                """
            )
            for row in cursor.fetchall():
                resource_url = f"/agent/dashboard/{row['id']}"
                cursor.execute(
                    "SELECT id FROM assets WHERE resource_url = %s",
                    (resource_url,),
                )
                if cursor.fetchone():
                    continue
                cursor.execute(
                    """
                    INSERT INTO assets (
                        nome, tipo, categoria, descricao, status, ordem_padrao, resource_url
                    ) VALUES (%s, %s, %s, %s, 'ativo', %s, %s)
                    """,
                    (
                        row.get("title"),
                        "interno",
                        row.get("category"),
                        row.get("description"),
                        40,
                        resource_url,
                    ),
                )
        except Exception as exc:
            app.logger.info(f"[ASSET] Templates internos não disponíveis: {exc}")

        conn.commit()
    except Exception as exc:
        if conn:
            conn.rollback()
        app.logger.error(f"[ASSET] Erro ao popular biblioteca de ativos: {exc}", exc_info=True)
    finally:
        if conn:
            conn.close()


def ensure_schema() -> None:
    conn = get_db()
    cursor = conn.cursor()
    
    # Adquirir lock exclusivo para garantir que apenas um worker execute a migração
    cursor.execute("SELECT pg_advisory_xact_lock(12345)")

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS users_new (
            id BIGSERIAL PRIMARY KEY,
            username TEXT NOT NULL UNIQUE,
            password BYTEA NOT NULL,
            nome_completo TEXT NOT NULL,
            cargo_original TEXT,
            departamento TEXT,
            role TEXT NOT NULL DEFAULT 'usuario',
            email TEXT,
            nome_usuario TEXT,
            unidade TEXT,
            is_active BOOLEAN NOT NULL DEFAULT TRUE,
            first_login BOOLEAN NOT NULL DEFAULT TRUE,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            updated_at TIMESTAMPTZ,
            last_login TIMESTAMPTZ
        );
        """
    )
    
    # Adicionar coluna nome_usuario se não existir (migration)
    cursor.execute(
        """
        DO $$ 
        BEGIN
            IF NOT EXISTS (
                SELECT 1 FROM information_schema.columns 
                WHERE table_name = 'users_new' AND column_name = 'nome_usuario'
            ) THEN
                ALTER TABLE users_new ADD COLUMN nome_usuario TEXT;
                CREATE UNIQUE INDEX IF NOT EXISTS users_new_nome_usuario_unique
                    ON users_new (LOWER(nome_usuario)) WHERE nome_usuario IS NOT NULL;
            END IF;
        END $$;
        """
    )

    cursor.execute(
        """
        DO $$ 
        BEGIN
            IF NOT EXISTS (
                SELECT 1 FROM information_schema.columns 
                WHERE table_name = 'users_new' AND column_name = 'avatar_url'
            ) THEN
                ALTER TABLE users_new ADD COLUMN avatar_url TEXT;
            END IF;
        END $$;
        """
    )

    cursor.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS users_new_email_unique
            ON users_new (LOWER(email));
        """
    )
    
    # Adicionar colunas para rastreamento de atividade em tempo real
    cursor.execute(
        """
        DO $$ 
        BEGIN
            IF NOT EXISTS (SELECT 1 FROM information_schema.columns WHERE table_name = 'users_new' AND column_name = 'last_seen_at') THEN
                ALTER TABLE users_new ADD COLUMN last_seen_at TIMESTAMPTZ;
            END IF;
            IF NOT EXISTS (SELECT 1 FROM information_schema.columns WHERE table_name = 'users_new' AND column_name = 'current_page') THEN
                ALTER TABLE users_new ADD COLUMN current_page TEXT;
            END IF;
        END $$;
        """
    )

    cursor.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS users_new_username_lower_unique
            ON users_new (LOWER(username));
        """
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS dashboards (
            id BIGSERIAL PRIMARY KEY,
            slug TEXT UNIQUE NOT NULL,
            title TEXT NOT NULL,
            description TEXT,
            category TEXT,
            embed_url TEXT NOT NULL,
            display_order INTEGER NOT NULL DEFAULT 0,
            is_active BOOLEAN NOT NULL DEFAULT TRUE,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        );
        """
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS user_dashboards (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL REFERENCES users_new(id) ON DELETE CASCADE,
            dashboard_id BIGINT NOT NULL REFERENCES dashboards(id) ON DELETE CASCADE,
            created_by BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            UNIQUE (user_id, dashboard_id)
        );
        """
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS planner_sync_logs (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
            user_name TEXT,
            dashboard_count INTEGER NOT NULL DEFAULT 0,
            status TEXT NOT NULL,
            message TEXT,
            task_id TEXT,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        );
        """
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS assets (
            id BIGSERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            tipo TEXT NOT NULL,
            categoria TEXT,
            descricao TEXT,
            status TEXT NOT NULL DEFAULT 'ativo',
            ordem_padrao INTEGER NOT NULL DEFAULT 0,
            embed_url TEXT,
            resource_url TEXT,
            config JSONB,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        );
        """
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS asset_assignments (
            id BIGSERIAL PRIMARY KEY,
            asset_id BIGINT NOT NULL REFERENCES assets(id) ON DELETE CASCADE,
            user_id BIGINT REFERENCES users_new(id) ON DELETE CASCADE,
            group_name TEXT,
            ordem INTEGER NOT NULL DEFAULT 0,
            visivel BOOLEAN NOT NULL DEFAULT TRUE,
            config JSONB,
            created_by BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            CHECK (
                (user_id IS NOT NULL AND group_name IS NULL)
                OR (user_id IS NULL AND group_name IS NOT NULL)
            )
        );
        """
    )

    cursor.execute(
        """
        DO $$ 
        BEGIN
            IF NOT EXISTS (
                SELECT 1 FROM information_schema.columns
                WHERE table_name = 'asset_assignments' AND column_name = 'config'
            ) THEN
                ALTER TABLE asset_assignments ADD COLUMN config JSONB;
            END IF;
        END $$;
        """
    )

    cursor.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS asset_assignments_user_unique
            ON asset_assignments (user_id, asset_id)
            WHERE user_id IS NOT NULL;
        """
    )

    cursor.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS asset_assignments_group_unique
            ON asset_assignments (group_name, asset_id)
            WHERE group_name IS NOT NULL;
        """
    )

    cursor.execute(
        """
        CREATE INDEX IF NOT EXISTS asset_assignments_asset_idx
            ON asset_assignments (asset_id);
        """
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS asset_logs (
            id BIGSERIAL PRIMARY KEY,
            action_type TEXT NOT NULL,
            asset_id BIGINT REFERENCES assets(id) ON DELETE SET NULL,
            actor_id BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
            target_type TEXT,
            target_value TEXT,
            details JSONB,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        );
        """
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS relatorio_meta_settings (
            id BIGINT PRIMARY KEY,
            meta_valor NUMERIC(16, 2) NOT NULL DEFAULT 0,
            meta_percentual NUMERIC(6, 2) NOT NULL DEFAULT 0,
            updated_by BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        );
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS relatorio_layouts (
            user_id BIGINT PRIMARY KEY REFERENCES users_new(id) ON DELETE CASCADE,
            layout JSONB,
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        );
        """
    )
    cursor.execute(
        """
        INSERT INTO relatorio_meta_settings (id, meta_valor, meta_percentual)
        VALUES (1, 0, 0)
        ON CONFLICT (id) DO NOTHING;
        """
    )

    conn.commit()
    conn.close()
    
    # Criar usuário admin anaissiabraao se não existir
    create_admin_user()
    # Popular biblioteca de ativos a partir do legado e templates internos
    seed_asset_library()


def fetch_relatorio_meta_settings() -> Dict[str, float]:
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            SELECT meta_valor, meta_percentual
            FROM relatorio_meta_settings
            WHERE id = 1
            """
        )
        row = cursor.fetchone()
        if not row:
            return {"meta_valor": 0.0, "meta_percentual": 0.0}
        return {
            "meta_valor": float(row.get("meta_valor") or 0),
            "meta_percentual": float(row.get("meta_percentual") or 0),
        }
    finally:
        conn.close()


def save_relatorio_meta_settings(meta_valor: float, meta_percentual: float, user_id: int) -> None:
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO relatorio_meta_settings (id, meta_valor, meta_percentual, updated_by, updated_at)
            VALUES (1, %s, %s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (id) DO UPDATE
            SET meta_valor = EXCLUDED.meta_valor,
                meta_percentual = EXCLUDED.meta_percentual,
                updated_by = EXCLUDED.updated_by,
                updated_at = CURRENT_TIMESTAMP
            """,
            (meta_valor, meta_percentual, user_id),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def fetch_relatorio_layout(user_id: int) -> Dict:
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT layout FROM relatorio_layouts WHERE user_id = %s",
            (user_id,),
        )
        row = cursor.fetchone()
        return row["layout"] if row and row.get("layout") else {}
    finally:
        conn.close()


def save_relatorio_layout(user_id: int, layout: Dict) -> None:
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO relatorio_layouts (user_id, layout, updated_at)
            VALUES (%s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (user_id) DO UPDATE
            SET layout = EXCLUDED.layout,
                updated_at = CURRENT_TIMESTAMP
            """,
            (user_id, psycopg2.extras.Json(layout)),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def create_admin_user() -> None:
    """Cria o usuário admin anaissiabraao com todos os privilégios"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Verifica se o usuário já existe
        cursor.execute(
            """
            SELECT id FROM users_new 
            WHERE LOWER(username) = LOWER('anaissiabraao@gmail.com')
               OR LOWER(nome_usuario) = LOWER('anaissiabraao')
            """
        )
        existing = cursor.fetchone()
        
        if not existing:
            # Cria senha padrão
            temp_password = "admin123"  # Senha temporária, deve ser alterada no primeiro acesso
            password_hash = bcrypt.hashpw(
                temp_password.encode("utf-8"), bcrypt.gensalt()
            )
            
            cursor.execute(
                """
                INSERT INTO users_new (
                    username, password, nome_completo, cargo_original,
                    departamento, role, email, nome_usuario, is_active, first_login
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (username) DO UPDATE SET
                    nome_usuario = EXCLUDED.nome_usuario,
                    role = EXCLUDED.role,
                    is_active = EXCLUDED.is_active
                """,
                (
                    "anaissiabraao@gmail.com",
                    psycopg2.Binary(password_hash),
                    "ABRAAO DE OLIVEIRA COSTA ANAISSI",
                    "DIRETOR",
                    "ADMINISTRATIVO",
                    "admin",
                    "anaissiabraao@portoex.com.br",
                    "anaissiabraao",
                    True,
                    True,  # Primeiro acesso
                ),
            )
            conn.commit()
            app.logger.info("[ADMIN] Usuário admin anaissiabraao criado com sucesso")
        else:
            # Atualiza o usuário existente para garantir que é admin (sem resetar senha)
            cursor.execute(
                """
                UPDATE users_new
                SET nome_usuario = 'anaissiabraao',
                    role = 'admin',
                    is_active = true
                WHERE id = %s
                """,
                (existing["id"],),
            )
            conn.commit()
            app.logger.info("[ADMIN] Usuário admin anaissiabraao atualizado (sem resetar senha)")
        
        conn.close()
    except Exception as exc:
        app.logger.error(f"[ADMIN] Erro ao criar usuário admin: {exc}")


def ensure_agent_tables():
    """Garante que as tabelas do agente existam via SQL direto."""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Verificar se tabelas já existem para evitar processamento desnecessário
        cursor.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_name = 'agent_rpa_types'
            )
        """)
        if cursor.fetchone()['exists']:
            conn.close()
            return

        app.logger.info("[AGENT] Criando tabelas do Agente IA...")
        
        # Criação das tabelas
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS agent_rpa_types (
                id BIGSERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                description TEXT,
                icon TEXT DEFAULT 'fa-cogs',
                is_active BOOLEAN NOT NULL DEFAULT TRUE,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS agent_rpas (
                id BIGSERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                description TEXT,
                rpa_type_id BIGINT REFERENCES agent_rpa_types(id) ON DELETE SET NULL,
                priority TEXT NOT NULL DEFAULT 'medium',
                frequency TEXT DEFAULT 'once',
                parameters JSONB,
                status TEXT NOT NULL DEFAULT 'pending',
                result JSONB,
                error_message TEXT,
                created_by BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
                executed_at TIMESTAMPTZ,
                completed_at TIMESTAMPTZ,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
            
            CREATE INDEX IF NOT EXISTS idx_agent_rpas_status ON agent_rpas(status);
            CREATE INDEX IF NOT EXISTS idx_agent_rpas_created_by ON agent_rpas(created_by);

            CREATE TABLE IF NOT EXISTS agent_data_sources (
                id BIGSERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                description TEXT,
                source_type TEXT NOT NULL DEFAULT 'database',
                connection_config JSONB,
                is_active BOOLEAN NOT NULL DEFAULT TRUE,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS agent_dashboard_requests (
                id BIGSERIAL PRIMARY KEY,
                title TEXT NOT NULL,
                description TEXT,
                category TEXT NOT NULL DEFAULT 'Outros',
                data_source_id BIGINT REFERENCES agent_data_sources(id) ON DELETE SET NULL,
                chart_types TEXT[],
                filters JSONB,
                status TEXT NOT NULL DEFAULT 'pending',
                leased_by TEXT,
                leased_until TIMESTAMPTZ,
                result_url TEXT,
                result_data JSONB,
                error_message TEXT,
                created_by BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
                processed_at TIMESTAMPTZ,
                completed_at TIMESTAMPTZ,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
            
            CREATE INDEX IF NOT EXISTS idx_agent_dashboard_requests_status ON agent_dashboard_requests(status);
            CREATE INDEX IF NOT EXISTS idx_agent_dashboard_requests_created_by ON agent_dashboard_requests(created_by);
            CREATE INDEX IF NOT EXISTS idx_agent_dashboard_requests_lease ON agent_dashboard_requests(leased_until);

            CREATE TABLE IF NOT EXISTS agent_settings (
                id BIGSERIAL PRIMARY KEY,
                setting_key TEXT NOT NULL UNIQUE,
                setting_value JSONB,
                description TEXT,
                updated_by BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS agent_logs (
                id BIGSERIAL PRIMARY KEY,
                action_type TEXT NOT NULL,
                entity_type TEXT,
                entity_id BIGINT,
                user_id BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
                details JSONB,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
            
            CREATE INDEX IF NOT EXISTS idx_agent_logs_action_type ON agent_logs(action_type);
            
            CREATE TABLE IF NOT EXISTS agent_dashboard_templates (
                id BIGSERIAL PRIMARY KEY,
                title TEXT NOT NULL,
                description TEXT,
                category TEXT NOT NULL DEFAULT 'Outros',
                data_source_id BIGINT REFERENCES agent_data_sources(id) ON DELETE SET NULL,
                query_config JSONB,
                layout_config JSONB,
                charts_config JSONB,
                filters_config JSONB,
                theme_config JSONB,
                is_published BOOLEAN DEFAULT false,
                is_public BOOLEAN DEFAULT false,
                thumbnail_url TEXT,
                linked_dashboard_id BIGINT,
                created_by BIGINT REFERENCES users_new(id) ON DELETE SET NULL,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
            
            CREATE INDEX IF NOT EXISTS idx_agent_dashboard_templates_created_by ON agent_dashboard_templates(created_by);
            CREATE INDEX IF NOT EXISTS idx_agent_dashboard_templates_published ON agent_dashboard_templates(is_published);
        """)

        # Inserção de dados iniciais
        cursor.execute("""
            INSERT INTO agent_rpa_types (name, description, icon) VALUES
            ('Extração de Dados', 'Extrai dados de sistemas externos (ERP, planilhas, APIs)', 'fa-download'),
            ('Processamento de Arquivos', 'Processa e transforma arquivos (PDF, Excel, CSV)', 'fa-file-alt'),
            ('Integração de Sistemas', 'Sincroniza dados entre sistemas diferentes', 'fa-sync'),
            ('Envio de Relatórios', 'Gera e envia relatórios automaticamente', 'fa-paper-plane'),
            ('Monitoramento', 'Monitora sistemas e envia alertas', 'fa-bell'),
            ('Backup de Dados', 'Realiza backup automático de dados', 'fa-database'),
            ('Web Scraping', 'Coleta dados de websites', 'fa-globe'),
            ('Automação de E-mail', 'Processa e responde e-mails automaticamente', 'fa-envelope')
            ON CONFLICT (name) DO UPDATE SET description = EXCLUDED.description;

            INSERT INTO agent_data_sources (name, description, source_type) VALUES
            ('Banco de Dados GeRot', 'Dados internos do sistema GeRot', 'database'),
            ('Power BI', 'Dados dos dashboards Power BI', 'api'),
            ('Planilhas Excel', 'Dados de planilhas compartilhadas', 'file'),
            ('ERP PortoEx', 'Sistema ERP da empresa', 'api'),
            ('API Externa', 'Dados de APIs de terceiros', 'api')
            ON CONFLICT (name) DO UPDATE SET description = EXCLUDED.description;
            
            INSERT INTO agent_settings (setting_key, setting_value, description) VALUES
            ('rpa_enabled', '{"enabled": true}', 'Habilita/desabilita funcionalidades de RPA'),
            ('dashboard_gen_enabled', '{"enabled": true}', 'Habilita/desabilita geração de dashboards'),
            ('max_concurrent_rpas', '{"value": 5}', 'Número máximo de RPAs executando simultaneamente')
            ON CONFLICT (setting_key) DO NOTHING;
        """)
        
        conn.commit()
        conn.close()
        app.logger.info("[AGENT] Tabelas e dados iniciais criados com sucesso.")
        
    except Exception as e:
        app.logger.error(f"[AGENT] Erro ao criar tabelas: {e}")


def seed_dashboards() -> None:
    conn = get_db()
    cursor = conn.cursor()

    for dash in DEFAULT_DASHBOARDS:
        cursor.execute(
            """
            INSERT INTO dashboards (slug, title, description, category, embed_url, display_order, is_active)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT(slug) DO UPDATE SET
                title=excluded.title,
                description=excluded.description,
                category=excluded.category,
                embed_url=excluded.embed_url,
                display_order=excluded.display_order,
                updated_at=CURRENT_TIMESTAMP
            """,
            (
                dash["slug"],
                dash["title"],
                dash["description"],
                dash["category"],
                dash["embed_url"],
                dash["display_order"],
                True,  # is_active como boolean
            ),
        )

    conn.commit()
    conn.close()


def normalize_roles() -> None:
    """Normaliza roles dos usuários com retry para evitar deadlocks"""
    max_retries = 3
    conn = None
    for attempt in range(max_retries):
        try:
            conn = get_db()
            cursor = conn.cursor()
            
            # Otimização: Só atualiza se o valor for diferente
            # Isso evita locks desnecessários em linhas que já estão corretas
            
            # 1. Normalizar admin_master -> admin
            cursor.execute(
                """
                UPDATE users_new
                SET role = 'admin'
                WHERE role = 'admin_master'
                """
            )
            
            # 2. Definir 'usuario' para quem não é admin e não é usuario (ex: null ou outros)
            cursor.execute(
                """
                UPDATE users_new
                SET role = 'usuario'
                WHERE role IS NULL OR (role != 'admin' AND role != 'usuario')
                """
            )
            
            conn.commit()
            conn.close()
            return
        except psycopg2.errors.DeadlockDetected:
            if attempt < max_retries - 1:
                import time
                time.sleep(0.1 * (attempt + 1))  # Backoff exponencial
                if conn:
                    try:
                        conn.rollback()
                    except:
                        pass
                    conn.close()
                continue
            else:
                app.logger.warning("[normalize_roles] Aviso: deadlock detected após múltiplas tentativas", exc_info=True)
                if conn:
                    try:
                        conn.rollback()
                    except:
                        pass
                    conn.close()
                return
        except Exception as exc:
            app.logger.warning(f"[normalize_roles] Aviso: {exc}")
            if conn:
                try:
                    conn.rollback()
                except:
                    pass
                conn.close()
            return
        

def _determine_role_from_cargo(cargo: str | None) -> str:
    cargo_normalizado = (cargo or "").strip().upper()
    return "admin" if cargo_normalizado in ADMIN_CARGOS else "usuario"


def import_users_from_excel() -> None:
    if not PLANILHA_USUARIOS.exists():
        app.logger.warning(
            "[IMPORTACAO] Arquivo dados.xlsx não encontrado em %s",
            PLANILHA_USUARIOS,
        )
        return

    try:
        workbook = load_workbook(
            filename=str(PLANILHA_USUARIOS), read_only=True, data_only=True
        )
    except Exception as exc:
        app.logger.error(
            "[IMPORTACAO] Falha ao abrir planilha de usuários: %s", exc
        )
        return

    conn = None
    inserted = updated = skipped = 0

    try:
        conn = get_db()
        cursor = conn.cursor()
        for row in workbook.active.iter_rows(min_row=2, values_only=True):
            if not row:
                skipped += 1
                continue

            values = list(row)
            while len(values) < 5:
                values.append(None)

            nome, email, cargo, unidade, departamento = values[:5]
            email = (email or "").strip()
            if not email:
                skipped += 1
                continue

            nome = (nome or "").strip() or email
            cargo = (cargo or "").strip()
            unidade = (unidade or "").strip()
            departamento = (departamento or "").strip()
            role = _determine_role_from_cargo(cargo)
            username = email.lower()

            cursor.execute(
                "SELECT id FROM users_new WHERE LOWER(email) = LOWER(%s)", (email,)
            )
            existing = cursor.fetchone()

            if existing:
                # Retry em caso de deadlock
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        cursor.execute(
                            """
                            UPDATE users_new
                            SET nome_completo = %s,
                                cargo_original = %s,
                                departamento = %s,
                                unidade = %s,
                                role = %s,
                                updated_at = NOW()
                            WHERE id = %s
                            """,
                            (
                                nome,
                                cargo or None,
                                departamento or None,
                                unidade or None,
                                role,
                                existing["id"],
                            ),
                        )
                        updated += 1
                        break
                    except psycopg2.errors.DeadlockDetected:
                        if attempt < max_retries - 1:
                            import time
                            import random
                            time.sleep(0.1 * (attempt + 1) + random.uniform(0, 0.1))  # Backoff exponencial com jitter
                            conn.rollback()
                            continue
                        else:
                            app.logger.warning(
                                f"[IMPORTACAO] Deadlock após múltiplas tentativas para {email}, pulando usuário",
                                exc_info=True,
                            )
                            skipped += 1
                            break  # Pula este usuário e continua
            else:
                temp_password = secrets.token_urlsafe(16)
                password_hash = bcrypt.hashpw(
                    temp_password.encode("utf-8"), bcrypt.gensalt()
                )
                # Retry em caso de deadlock ou unique violation
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        cursor.execute(
                            """
                            INSERT INTO users_new (
                                username,
                                password,
                                nome_completo,
                                cargo_original,
                                departamento,
                                role,
                                email,
                                unidade,
                                first_login,
                                nome_usuario
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, TRUE, NULL)
                            ON CONFLICT (username) DO UPDATE SET
                                nome_completo = EXCLUDED.nome_completo,
                                cargo_original = EXCLUDED.cargo_original,
                                departamento = EXCLUDED.departamento,
                                unidade = EXCLUDED.unidade,
                                role = EXCLUDED.role,
                                email = EXCLUDED.email,
                                updated_at = NOW()
                            """,
                            (
                                username,
                                psycopg2.Binary(password_hash),
                                nome,
                                cargo or None,
                                departamento or None,
                                role,
                                email,
                                unidade or None,
                            ),
                        )
                        # ON CONFLICT sempre retorna rowcount > 0, precisa verificar se foi insert ou update
                        if cursor.rowcount > 0:
                            # Verifica se foi realmente insert (created_at == updated_at) ou update
                            cursor.execute("SELECT created_at, updated_at FROM users_new WHERE username = %s", (username,))
                            user_check = cursor.fetchone()
                            if user_check and user_check.get("created_at") == user_check.get("updated_at"):
                                inserted += 1
                            else:
                                updated += 1
                        break
                    except (psycopg2.errors.DeadlockDetected, psycopg2.errors.UniqueViolation) as e:
                        if attempt < max_retries - 1:
                            import time
                            import random
                            time.sleep(0.1 * (attempt + 1) + random.uniform(0, 0.1))  # Backoff exponencial com jitter
                            conn.rollback()
                            # Tenta novamente como UPDATE se for unique violation
                            if isinstance(e, psycopg2.errors.UniqueViolation):
                                cursor.execute("SELECT id FROM users_new WHERE username = %s", (username,))
                                existing = cursor.fetchone()
                                if existing:
                                    # Tenta fazer UPDATE em vez de INSERT
                                    try:
                                        cursor.execute(
                                            """
                                            UPDATE users_new
                                            SET nome_completo = %s,
                                                cargo_original = %s,
                                                departamento = %s,
                                                unidade = %s,
                                                role = %s,
                                                email = %s,
                                                updated_at = NOW()
                                            WHERE id = %s
                                            """,
                                            (
                                                nome,
                                                cargo or None,
                                                departamento or None,
                                                unidade or None,
                                                role,
                                                email,
                                                existing["id"],
                                            ),
                                        )
                                        updated += 1
                                        break
                                    except psycopg2.errors.DeadlockDetected:
                                        continue
                            continue
                        else:
                            app.logger.warning(
                                f"[IMPORTACAO] Erro após múltiplas tentativas para {email}, pulando usuário",
                                exc_info=True,
                            )
                            skipped += 1
                            break  # Pula este usuário e continua

        conn.commit()
        app.logger.info(
            "[IMPORTACAO] Usuários sincronizados. Inseridos=%s | Atualizados=%s | Ignorados=%s",
            inserted,
            updated,
            skipped,
        )
    except Exception as exc:
        if conn:
            conn.rollback()
        app.logger.exception(
            "[IMPORTACAO] Erro ao sincronizar usuários da planilha: %s", exc
        )
    finally:
        if conn:
            conn.close()
        workbook.close()


def _run_db_init_with_retry(max_attempts: int = 5, delay_seconds: float = 2.0) -> None:
    """Executa inicialização do banco com retry (útil quando o PostgreSQL sobe após o app no Railway)."""
    last_error = None
    for attempt in range(1, max_attempts + 1):
        try:
            ensure_schema()
            ensure_agent_tables()
            seed_dashboards()
            normalize_roles()
            app.logger.info("Inicialização do banco concluída com sucesso.")
            return
        except Exception as e:
            last_error = e
            app.logger.warning(
                "Tentativa %d/%d de inicialização do banco falhou: %s",
                attempt, max_attempts, e,
            )
            if attempt < max_attempts:
                time.sleep(delay_seconds)
    app.logger.error("Erro na inicialização do banco após %d tentativas: %s", max_attempts, last_error)


with app.app_context():
    _run_db_init_with_retry()


# --------------------------------------------------------------------------- #
# Funções auxiliares de dados
# --------------------------------------------------------------------------- #
ASSET_TYPES = {"PBI", "interno", "grafico", "rpa", "outro"}
ASSET_STATUSES = {"ativo", "legado", "experimental"}


def normalize_asset_type(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    if value.upper() == "PBI":
        return "PBI"
    return value.lower()


def normalize_asset_status(value: str) -> str:
    return (value or "").strip().lower()


def extract_embed_url(embed_input: str) -> str | None:
    if not embed_input:
        return None
    cleaned = embed_input.strip()
    if cleaned.startswith("http://") or cleaned.startswith("https://"):
        return cleaned
    url_match = re.search(r'src="([^"]+)"', cleaned)
    if not url_match:
        return None
    return url_match.group(1)


def is_safe_embed_url(embed_url: str) -> bool:
    if not embed_url:
        return False
    lowered = embed_url.lower()
    blocked_markers = ("access_token=", "token=", "sig=")
    return not any(marker in lowered for marker in blocked_markers)

def fetch_dashboards(active_only: bool = True) -> List[Dict]:
    conn = get_db()
    cursor = conn.cursor()
    query = "SELECT * FROM dashboards"
    if active_only:
        query += " WHERE is_active = true"
    query += " ORDER BY display_order, title"
    cursor.execute(query)
    dashboards = cursor.fetchall()
    conn.close()
    return dashboards


def fetch_assets(statuses: List[str] | None = None) -> List[Dict]:
    conn = get_db()
    cursor = conn.cursor()
    query = "SELECT a.*, a.config AS asset_config FROM assets a"
    params: Tuple = ()
    if statuses:
        query += " WHERE status = ANY(%s)"
        params = (statuses,)
    query += " ORDER BY ordem_padrao, nome"
    cursor.execute(query, params)
    assets = cursor.fetchall()
    conn.close()
    return assets


def fetch_asset_by_id(asset_id: int) -> Dict | None:
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM assets WHERE id = %s", (asset_id,))
    asset = cursor.fetchone()
    conn.close()
    return dict(asset) if asset else None


def log_asset_action(
    action_type: str,
    asset_id: int | None,
    actor_id: int | None,
    target_type: str | None = None,
    target_value: str | None = None,
    details: Dict | None = None,
) -> None:
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO asset_logs (action_type, asset_id, actor_id, target_type, target_value, details)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (action_type, asset_id, actor_id, target_type, target_value, json.dumps(details) if details else None),
        )
        conn.commit()
        conn.close()
    except Exception as exc:
        app.logger.error(f"[ASSET] Falha ao registrar log: {exc}", exc_info=True)


def fetch_asset_assignments(
    target_type: str,
    target_value: str,
    include_inactive: bool = False,
) -> List[Dict]:
    conn = get_db()
    cursor = conn.cursor()
    status_clause = "" if include_inactive else "AND a.status IN ('ativo', 'experimental')"
    if target_type == "user":
        cursor.execute(
            f"""
            SELECT aa.*,
                   aa.config AS assignment_config,
                   a.*,
                   a.config AS asset_config
            FROM asset_assignments aa
            JOIN assets a ON a.id = aa.asset_id
            WHERE aa.user_id = %s {status_clause}
            ORDER BY aa.ordem, a.ordem_padrao, a.nome
            """,
            (int(target_value),),
        )
    else:
        cursor.execute(
            f"""
            SELECT aa.*,
                   aa.config AS assignment_config,
                   a.*,
                   a.config AS asset_config
            FROM asset_assignments aa
            JOIN assets a ON a.id = aa.asset_id
            WHERE aa.group_name = %s {status_clause}
            ORDER BY aa.ordem, a.ordem_padrao, a.nome
            """,
            (target_value,),
        )
    rows = cursor.fetchall()
    conn.close()
    return rows


def fetch_assets_for_user(
    user_id: int | None,
    department: str | None,
    include_inactive: bool = False,
) -> List[Dict]:
    conn = get_db()
    cursor = conn.cursor()

    status_clause = "" if include_inactive else "AND a.status IN ('ativo', 'experimental')"

    user_rows: List[Dict] = []
    if user_id is not None:
        cursor.execute(
            f"""
            SELECT aa.*,
                   aa.config AS assignment_config,
                   a.*,
                   a.config AS asset_config
            FROM asset_assignments aa
            JOIN assets a ON a.id = aa.asset_id
            WHERE aa.user_id = %s AND aa.visivel = true {status_clause}
            """,
            (user_id,),
        )
        user_rows = cursor.fetchall()

    group_rows: List[Dict] = []
    if department:
        dept_trimmed = (department or "").strip()
        cursor.execute(
            f"""
            SELECT aa.*,
                   aa.config AS assignment_config,
                   a.*,
                   a.config AS asset_config
            FROM asset_assignments aa
            JOIN assets a ON a.id = aa.asset_id
            WHERE LOWER(TRIM(aa.group_name)) = LOWER(TRIM(%s)) AND aa.visivel = true {status_clause}
            """,
            (dept_trimmed,),
        )
        group_rows = cursor.fetchall()

    conn.close()

    merged: Dict[int, Dict] = {}
    for row in group_rows:
        merged[row["asset_id"]] = dict(row)
        merged[row["asset_id"]]["assignment_source"] = "grupo"
    for row in user_rows:
        merged[row["asset_id"]] = dict(row)
        merged[row["asset_id"]]["assignment_source"] = "usuario"

    assets = list(merged.values())
    assets.sort(
        key=lambda r: (
            r.get("ordem", 0),
            r.get("ordem_padrao", 0),
            r.get("nome", ""),
        )
    )
    return assets


def _fetch_default_relatorio_asset() -> Dict | None:
    """Retorna o asset Relatório de Resultados com todos os componentes, para fallback quando o usuário não tem atribuições."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT a.*, a.config AS asset_config
        FROM assets a
        WHERE a.tipo = 'interno'
          AND a.config->>'legacy_view' = 'relatorio_resultados'
          AND a.status IN ('ativo', 'experimental')
        LIMIT 1
        """
    )
    row = cursor.fetchone()
    conn.close()
    if not row:
        return None
    asset = dict(row)
    available = (asset.get("asset_config") or {}).get("available_components") or []
    component_keys = [c["key"] for c in available if isinstance(c, dict) and c.get("key")]
    if not component_keys:
        component_keys = [c if isinstance(c, str) else c.get("key", "") for c in available]
    asset["assignment_config"] = {"components": component_keys}
    asset["assignment_source"] = "padrao"
    return asset


def save_asset_assignments(
    target_type: str,
    target_value: str,
    assignments: List[Dict],
    actor_id: int,
) -> None:
    conn = get_db()
    cursor = conn.cursor()
    if target_type == "user":
        cursor.execute("DELETE FROM asset_assignments WHERE user_id = %s", (int(target_value),))
    else:
        cursor.execute("DELETE FROM asset_assignments WHERE group_name = %s", (target_value,))

    if assignments:
        rows = []
        for item in assignments:
            rows.append(
                (
                    item["asset_id"],
                    int(target_value) if target_type == "user" else None,
                    target_value if target_type == "group" else None,
                    item.get("ordem", 0),
                    item.get("visivel", True),
                    item.get("config"),
                    actor_id,
                )
            )
        cursor.executemany(
            """
            INSERT INTO asset_assignments (
                asset_id, user_id, group_name, ordem, visivel, config, created_by, created_at, updated_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
            """,
            rows,
        )

    conn.commit()
    conn.close()

def fetch_users() -> List[Dict]:
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT id, nome_completo, username, email, role
        FROM users_new
        WHERE is_active = true
        ORDER BY nome_completo
        """
    )
    users = cursor.fetchall()
    conn.close()
    return users


def get_user_dashboard_map() -> Dict[int, Dict[str, List]]:
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT ud.user_id, d.id as dashboard_id, d.title, d.category
        FROM user_dashboards ud
        JOIN dashboards d ON d.id = ud.dashboard_id
        WHERE d.is_active = true
        ORDER BY d.display_order, d.title
        """
    )
    data: Dict[int, Dict[str, List]] = {}
    for row in cursor.fetchall():
        entry = data.setdefault(row["user_id"], {"ids": set(), "items": []})
        entry["ids"].add(row["dashboard_id"])
        entry["items"].append({"title": row["title"], "category": row["category"]})

    conn.close()

    for entry in data.values():
        entry["items"].sort(key=lambda i: i["title"])
        entry["ids"] = list(entry["ids"])

    return data


def save_user_dashboards(user_id: int, dashboard_ids: List[int], actor_id: int) -> None:
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM user_dashboards WHERE user_id = %s", (user_id,))
    if dashboard_ids:
        cursor.executemany(
            """
            INSERT INTO user_dashboards (user_id, dashboard_id, created_by)
            VALUES (%s, %s, %s)
            """,
            [(user_id, dash_id, actor_id) for dash_id in dashboard_ids],
        )
    conn.commit()
    conn.close()


def log_planner_sync(
    user_id: int,
    user_name: str,
    dashboard_count: int,
    status: str,
    message: str,
    task_id: str | None = None,
) -> None:
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO planner_sync_logs (user_id, user_name, dashboard_count, status, message, task_id)
        VALUES (%s, %s, %s, %s, %s, %s)
        """,
        (user_id, user_name, dashboard_count, status, message, task_id),
    )
    conn.commit()
    conn.close()


def get_recent_planner_logs(limit: int = 8) -> List[Dict]:
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT user_name, dashboard_count, status, message, task_id, created_at
        FROM planner_sync_logs
        ORDER BY created_at DESC
        LIMIT %s
        """,
        (limit,),
    )
    logs = cursor.fetchall()
    conn.close()
    return logs


def sync_dashboards_to_planner() -> Tuple[int, List[str]]:
    if not planner_client.is_configured:
        raise PlannerIntegrationError(
            "Configure MS_TENANT_ID, MS_CLIENT_ID, MS_CLIENT_SECRET, "
            "MS_PLANNER_PLAN_ID e MS_PLANNER_BUCKET_ID para usar esta função."
        )

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT u.id as user_id, u.nome_completo, u.email,
               d.title, d.category, d.embed_url
        FROM user_dashboards ud
        JOIN users_new u ON u.id = ud.user_id
        JOIN dashboards d ON d.id = ud.dashboard_id
        WHERE u.is_active = true AND d.is_active = true
        ORDER BY u.nome_completo, d.display_order, d.title
        """
    )

    assignments: Dict[int, Dict[str, List[Dict[str, str]]]] = {}
    for row in cursor.fetchall():
        user_entry = assignments.setdefault(
            row["user_id"],
            {"name": row["nome_completo"], "email": row["email"], "dashboards": []},
        )
        user_entry["dashboards"].append(
            {
                "title": row["title"],
                "category": row["category"],
                "url": row["embed_url"],
            }
        )

    conn.close()

    if not assignments:
        raise PlannerIntegrationError("Nenhum usuário possui dashboards atribuídos.")

    successes = 0
    errors: List[str] = []
    today = date.today().strftime("%d/%m/%Y")
    start_time = datetime.utcnow().replace(hour=11, minute=0, second=0, microsecond=0)
    due_time = start_time + timedelta(hours=6)

    for user_id, payload in assignments.items():
        dashboards = payload["dashboards"]
        if not dashboards:
            continue

        title = f"Agenda de dashboards - {payload['name']} ({today})"
        description_lines = [
            f"Agenda automática gerada em {datetime.now().strftime('%d/%m/%Y %H:%M')}.",
            "",
            "Dashboards liberados para hoje:",
        ]
        for idx, dash in enumerate(dashboards, start=1):
            description_lines.append(
                f"{idx}. {dash['title']} ({dash['category']}) - {dash['url']}"
            )

        description = "\n".join(description_lines)

        try:
            task = planner_client.create_dashboard_task(
                title=title,
                description=description,
                start_time=start_time,
                due_time=due_time,
            )
            log_planner_sync(
                user_id=user_id,
                user_name=payload["name"],
                dashboard_count=len(dashboards),
                status="success",
                message="Tarefa criada e agenda enviada.",
                task_id=task.get("id"),
            )
            successes += 1
        except PlannerIntegrationError as exc:
            errors.append(f"{payload['name']}: {exc}")
            log_planner_sync(
                user_id=user_id,
                user_name=payload["name"],
                dashboard_count=len(dashboards),
                status="error",
                message=str(exc),
            )

    return successes, errors


# --------------------------------------------------------------------------- #
# Funções de autenticação
# --------------------------------------------------------------------------- #
def authenticate_user(identifier: str, password: str):
    try:
        conn = get_db()
        cursor = conn.cursor()
        # Aceita email OU nome_usuario OU username
        cursor.execute(
            """
            SELECT id, username, password, nome_completo, cargo_original,
                   departamento, role, email, nome_usuario, first_login
            FROM users_new
            WHERE (LOWER(username) = LOWER(%s) 
                   OR LOWER(email) = LOWER(%s)
                   OR (nome_usuario IS NOT NULL AND LOWER(nome_usuario) = LOWER(%s)))
              AND is_active = true
            """,
            (identifier, identifier, identifier),
        )
        user = cursor.fetchone()
        conn.close()
        
        if not user:
            app.logger.debug(f"[AUTH] Usuário não encontrado: {identifier}")
            return None
        
        # Verifica senha
        password_valid = bcrypt.checkpw(password.encode("utf-8"), _as_bytes(user["password"]))
        if not password_valid:
            app.logger.debug(f"[AUTH] Senha incorreta para: {identifier}")
            return None
        
        app.logger.info(f"[AUTH] Login bem-sucedido: {identifier} (ID: {user['id']}, first_login: {user['first_login']})")
        role = "admin" if user["role"] == "admin" else "usuario"
        return {
            "id": user["id"],
            "username": user["username"],
            "nome_completo": user["nome_completo"],
            "cargo_original": user["cargo_original"],
            "departamento": user["departamento"],
            "role": role,
            "email": user["email"],
            "nome_usuario": user.get("nome_usuario"),
            "first_login": user["first_login"],
        }
    except Exception as exc:
        app.logger.error(f"[AUTH] Erro na autenticação: {exc}", exc_info=True)
        return None


def update_user_password(user_id: int, new_password: str, new_email: str = None) -> bool:
    try:
        conn = get_db()
        cursor = conn.cursor()
        password_hash = bcrypt.hashpw(new_password.encode("utf-8"), bcrypt.gensalt())
        
        # Se novo email fornecido e termina com @portoex.com.br, atualiza
        if new_email and new_email.lower().endswith("@portoex.com.br"):
            cursor.execute(
                """
                UPDATE users_new
                SET password = %s, email = %s, first_login = FALSE, 
                    updated_at = CURRENT_TIMESTAMP, last_login = CURRENT_TIMESTAMP
                WHERE id = %s
                """,
                (psycopg2.Binary(password_hash), new_email.lower(), user_id),
            )
        else:
            cursor.execute(
                """
                UPDATE users_new
                SET password = %s, first_login = FALSE, updated_at = CURRENT_TIMESTAMP,
                    last_login = CURRENT_TIMESTAMP
                WHERE id = %s
                """,
                (psycopg2.Binary(password_hash), user_id),
            )
        conn.commit()
        conn.close()
        return True
    except Exception as exc:
        print(f"Erro ao atualizar senha: {exc}")
        return False
            
            
def get_user_by_id(user_id: int):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT id, username, nome_completo, cargo_original,
                   departamento, role, email, is_active, nome_usuario,
                   avatar_url
            FROM users_new
            WHERE id = %s
            """,
            (user_id,),
        )
        user = cursor.fetchone()
        conn.close()
        if user:
            return dict(user)
        return None
    except Exception as exc:
        print(f"Erro ao buscar usuário: {exc}")
        return None


# --------------------------------------------------------------------------- #
# Rotas principais
# --------------------------------------------------------------------------- #
@app.route("/")
def index():
    if "user_id" in session:
        return (
            redirect(url_for("admin_dashboard"))
            if is_admin_session()
            else redirect(url_for("team_dashboard"))
        )
    return redirect(url_for("login"))


def _build_frontend_redirect(path: str = "") -> str:
    suffix = f"/{path.lstrip('/')}" if path else ""
    return f"{FRONTEND_APP_URL}{suffix}" if FRONTEND_APP_URL else ""


@app.route("/assets/<path:asset_path>")
def frontend_assets_root(asset_path: str):
    """Serve assets do build Vite quando o SPA roda pelo backend."""
    assets_dir = FRONTEND_DIST_DIR / "assets"
    if assets_dir.exists():
        return send_from_directory(str(assets_dir), asset_path)
    target = _build_frontend_redirect(f"assets/{asset_path}")
    if target:
        return redirect(target)
    return jsonify({"error": "Frontend build não encontrado em frontend/dist"}), 404


@app.route("/app/assets/<path:asset_path>")
def frontend_assets_app(asset_path: str):
    """Alias para servir assets sob /app/assets."""
    return frontend_assets_root(asset_path)


@app.route("/app")
@app.route("/app/<path:subpath>")
@login_required
def frontend_spa(subpath: str = ""):
    """
    Modo híbrido:
    - "/" mantém fluxo legado (login/dashboard)
    - "/app" entrega o SPA Vite (local) ou redireciona para FRONTEND_APP_URL.
    """
    index_file = FRONTEND_DIST_DIR / "index.html"
    if index_file.exists():
        return send_from_directory(str(FRONTEND_DIST_DIR), "index.html")

    target = _build_frontend_redirect(subpath)
    if target:
        return redirect(target)
    return jsonify({"error": "Frontend SPA não configurado. Defina FRONTEND_APP_URL ou publique frontend/dist."}), 404


@app.route("/signin")
def signin():
    """Alias legado para a página de login."""
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard_redirect():
    """Alias legado para encaminhar ao dashboard correto por perfil."""
    return redirect(url_for("admin_dashboard") if is_admin_session() else url_for("team_dashboard"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if "new_password" in request.form:
            user_id = session.get("temp_user_id")
            new_password = request.form.get("new_password", "").strip()
            confirm_password = request.form.get("confirm_password", "").strip()

            if not user_id:
                flash("Sessão expirada. Faça login novamente.", "error")
                return redirect(url_for("login"))

            if len(new_password) < 6:
                flash("A nova senha deve ter pelo menos 6 caracteres.", "error")
                user = get_user_by_id(user_id)
                return render_template(get_template("first_login.html"), user=user)

            if new_password != confirm_password:
                flash("As senhas não coincidem.", "error")
                user = get_user_by_id(user_id)
                return render_template(get_template("first_login.html"), user=user)

            # Permite atualizar email no primeiro acesso se fornecido
            new_email = request.form.get("new_email", "").strip()
            if new_email and not new_email.lower().endswith("@portoex.com.br"):
                flash("O email deve terminar com @portoex.com.br", "error")
                user = get_user_by_id(user_id)
                return render_template(get_template("first_login.html"), user=user)
            
            if update_user_password(user_id, new_password, new_email if new_email else None):
                user = get_user_by_id(user_id)
                if user:
                    session.update(
                        {
                            "user_id": user["id"],
                            "username": user["username"],
                            "role": user["role"],
                            "email": user.get("email", ""),
                            "nome_completo": user["nome_completo"],
                            "departamento": user["departamento"],
                        }
                    )
                    session.pop("temp_user_id", None)
                    flash(
                        f"Senha atualizada! Bem-vindo, {user['nome_completo']}!",
                        "success",
                    )
                    next_url = _safe_redirect_url(request.form.get("next") or request.args.get("next"))
                    if next_url:
                        return redirect(next_url)
                    return redirect(url_for("index"))
                flash("Não foi possível carregar o usuário.", "error")
            else:
                flash("Erro ao atualizar senha. Tente novamente.", "error")
                user = get_user_by_id(user_id)
                return render_template(get_template("first_login.html"), user=user)

        identifier = (
            request.form.get("username") or request.form.get("email", "")
        ).strip()
        password = request.form.get("password", "").strip()

        if identifier and password:
            # Primeiro tenta autenticar
            user = authenticate_user(identifier, password)
            
            if user:
                # Se é primeiro acesso, permite qualquer email
                # Se não é primeiro acesso, valida email @portoex.com.br
                if not user["first_login"] and "@" in identifier:
                    # Se não é primeiro acesso e o email usado não termina com @portoex.com.br
                    if not identifier.lower().endswith("@portoex.com.br"):
                        # Verifica se o email do usuário no banco termina com @portoex.com.br
                        user_email = user.get("email", "").lower()
                        if user_email and not user_email.endswith("@portoex.com.br"):
                            flash(
                                "Use um email @portoex.com.br para acessar o sistema. "
                                "Se este é seu primeiro acesso, use seu email pessoal para definir a senha.",
                                "error",
                            )
                            return render_template(get_template("login.html"))
                
                # Primeiro acesso: redireciona para definir senha
                if user["first_login"]:
                    session["temp_user_id"] = user["id"]
                    flash(
                        f"Bem-vindo, {user['nome_completo']}! Defina uma nova senha.",
                        "info",
                    )
                    return render_template(get_template("first_login.html"), user=user)

                session.update(
                    {
                        "user_id": user["id"],
                        "username": user["username"],
                        "role": user["role"],
                        "email": user.get("email", ""),
                        "nome_completo": user["nome_completo"],
                        "departamento": user["departamento"],
                    }
                )
                flash(f"Bem-vindo de volta, {user['nome_completo']}!", "success")
                next_url = _safe_redirect_url(request.form.get("next") or request.args.get("next"))
                if next_url:
                    return redirect(next_url)
                return redirect(url_for("index"))
            else:
                flash("Usuário ou senha incorretos!", "error")
        else:
            flash("Por favor, informe usuário/email e senha.", "error")

    return render_template(get_template("login.html"))


@app.route("/logout")
def logout():
    session.clear()
    flash("Logout realizado com sucesso!", "success")
    next_url = request.args.get("next")
    if next_url:
        return redirect(url_for("login", next=next_url))
    return redirect(url_for("login"))


# --------------------------------------------------------------------------- #
# API para o frontend SPA: sessão e dashboards (mesmo modelo do team_dashboard)
# --------------------------------------------------------------------------- #

def _safe_redirect_url(next_url: str | None) -> str | None:
    """Retorna next_url apenas se for segura (mesmo host do backend ou FRONTEND_APP_URL)."""
    if not next_url or not next_url.strip():
        return None
    next_url = next_url.strip()
    try:
        parsed = urlparse(next_url)
        if not parsed.scheme or not parsed.netloc:
            return None
        allowed_hosts = [request.host]
        if FRONTEND_APP_URL:
            front_parsed = urlparse(FRONTEND_APP_URL)
            if front_parsed.netloc:
                allowed_hosts.append(front_parsed.netloc)
        if parsed.netloc.lower() in [h.lower() for h in allowed_hosts]:
            return next_url
    except Exception:
        pass
    return None


@app.route("/api/session", methods=["GET"])
def api_session():
    """Retorna dados da sessão atual. 401 se não autenticado (SPA usa para redirecionar ao login)."""
    if not session.get("user_id"):
        return jsonify({"error": "Não autenticado"}), 401
    return jsonify({
        "user_id": session.get("user_id"),
        "username": session.get("username"),
        "role": session.get("role"),
        "nome_completo": session.get("nome_completo"),
        "departamento": session.get("departamento"),
        "email": session.get("email"),
    }), 200


def _serialize_asset(a: Dict) -> Dict:
    """Converte um asset (RealDict) para dict JSON-serializável."""
    out = {}
    for k, v in a.items():
        if v is None:
            out[k] = None
        elif isinstance(v, (datetime, date)):
            out[k] = v.isoformat() if hasattr(v, "isoformat") else str(v)
        elif isinstance(v, (bytes, memoryview)):
            continue
        elif isinstance(v, dict):
            out[k] = _serialize_asset(v) if v else v
        elif isinstance(v, list):
            out[k] = [_serialize_asset(x) if isinstance(x, dict) else x for x in v]
        else:
            out[k] = v
    return out


@app.route("/api/team-dashboard", methods=["GET"])
@login_required
def api_team_dashboard():
    """Retorna os mesmos dados que team_dashboard passa ao template: regular_assets, internal_assets, is_admin."""
    show_all = is_admin_session() and request.args.get("all") == "1"
    preview_user_id = request.args.get("preview_user_id", type=int)
    preview_group = request.args.get("preview_group", "").strip() or None

    effective_user_id = session.get("user_id")
    effective_department = session.get("departamento")

    if is_admin_session():
        if preview_user_id:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute(
                "SELECT id, nome_completo, departamento FROM users_new WHERE id = %s",
                (preview_user_id,),
            )
            preview_user = cursor.fetchone()
            conn.close()
            if preview_user:
                effective_user_id = preview_user["id"]
                effective_department = preview_user.get("departamento")
        elif preview_group:
            effective_user_id = None
            effective_department = preview_group

    assets: List[Dict] = []
    if show_all and is_admin_session():
        assets = fetch_assets(["ativo", "experimental"])
    else:
        assets = fetch_assets_for_user(effective_user_id, effective_department)

    if not assets and effective_user_id is not None:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT d.title, d.description, d.category, d.embed_url
            FROM dashboards d
            JOIN user_dashboards ud ON ud.dashboard_id = d.id
            WHERE ud.user_id = %s AND d.is_active = true
            ORDER BY d.display_order, d.title
            """,
            (effective_user_id,),
        )
        dashboards = cursor.fetchall()
        conn.close()
        assets = [
            {
                "id": idx + 1,
                "nome": row["title"],
                "tipo": "PBI",
                "categoria": row.get("category"),
                "descricao": row.get("description"),
                "embed_url": row.get("embed_url"),
                "resource_url": None,
                "status": "ativo",
                "ordem_padrao": idx,
                "config": None,
                "assignment_source": "legado",
            }
            for idx, row in enumerate(dashboards or [])
        ]
        if not assets:
            default_asset = _fetch_default_relatorio_asset()
            if default_asset:
                assets = [default_asset]

    internal_assets = []
    regular_assets = []
    for asset in assets:
        if (
            asset.get("tipo") == "interno"
            and (asset.get("asset_config") or {}).get("embed_mode")
            and (asset.get("assignment_config") or {}).get("components")
        ):
            internal_assets.append(asset)
        else:
            regular_assets.append(asset)

    return jsonify({
        "regular_assets": [_serialize_asset(a) for a in regular_assets],
        "internal_assets": [_serialize_asset(a) for a in internal_assets],
        "is_admin": is_admin_session(),
    }), 200


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    user = get_user_by_id(session["user_id"])
    if not user:
        flash("Não foi possível carregar seu perfil.", "error")
        return redirect(url_for("index"))

    if request.method == "POST":
        user_id = session["user_id"]
        conn = get_db()
        cursor = conn.cursor()

        errors: List[str] = []
        updates: List[str] = []
        params: List = []
        avatar_payload = None
        avatar_meta = None

        # Campos principais
        nome_completo = request.form.get("nome_completo", "").strip()
        if not nome_completo:
            errors.append("Informe seu nome completo.")
        elif nome_completo != user["nome_completo"]:
            updates.append("nome_completo = %s")
            params.append(nome_completo)

        new_username = request.form.get("username", "").strip()
        if not new_username:
            errors.append("Informe um nome de usuário.")
        elif new_username.lower() != user["username"].lower():
            cursor.execute(
                "SELECT id FROM users_new WHERE LOWER(username) = LOWER(%s) AND id <> %s",
                (new_username, user_id),
            )
            if cursor.fetchone():
                errors.append("Este nome de usuário já está em uso.")
            else:
                updates.append("username = %s")
                params.append(new_username)

        new_nome_usuario = _sanitize_optional(request.form.get("nome_usuario"))
        current_nome_usuario = _sanitize_optional(user.get("nome_usuario"))
        if new_nome_usuario != current_nome_usuario:
            if new_nome_usuario:
                cursor.execute(
                    """
                    SELECT id FROM users_new
                    WHERE LOWER(nome_usuario) = LOWER(%s) AND id <> %s
                    """,
                    (new_nome_usuario, user_id),
                )
                if cursor.fetchone():
                    errors.append("Este usuário público já está em uso.")
                else:
                    updates.append("nome_usuario = %s")
                    params.append(new_nome_usuario)
            else:
                updates.append("nome_usuario = %s")
                params.append(None)

        new_email = _sanitize_optional(request.form.get("email"))
        current_email = _sanitize_optional(user.get("email"))
        if new_email != current_email:
            if new_email:
                if not re.match(r"[^@]+@[^@]+\.[^@]+", new_email):
                    errors.append("Informe um email válido.")
                else:
                    cursor.execute(
                        "SELECT id FROM users_new WHERE LOWER(email) = LOWER(%s) AND id <> %s",
                        (new_email, user_id),
                    )
                    if cursor.fetchone():
                        errors.append("Este email já está em uso.")
                    else:
                        updates.append("email = %s")
                        params.append(new_email.lower())
            else:
                updates.append("email = %s")
                params.append(None)

        new_departamento = _sanitize_optional(request.form.get("departamento"))
        if new_departamento != _sanitize_optional(user.get("departamento")):
            updates.append("departamento = %s")
            params.append(new_departamento)

        new_cargo = _sanitize_optional(request.form.get("cargo_original"))
        if new_cargo != _sanitize_optional(user.get("cargo_original")):
            updates.append("cargo_original = %s")
            params.append(new_cargo)

        # Upload de avatar
        avatar_file = request.files.get("avatar")
        if avatar_file and avatar_file.filename:
            if not is_allowed_avatar_file(avatar_file.filename):
                errors.append("Formato de imagem não suportado. Use PNG, JPG, JPEG, WEBP ou GIF.")
            else:
                avatar_bytes = avatar_file.read()
                if not avatar_bytes:
                    errors.append("Não foi possível ler o arquivo da foto.")
                elif len(avatar_bytes) > MAX_AVATAR_SIZE_BYTES:
                    errors.append("A foto deve ter no máximo 5MB.")
                else:
                    ext = avatar_file.filename.rsplit(".", 1)[-1].lower()
                    base_name = secure_filename(os.path.splitext(avatar_file.filename)[0]) or "avatar"
                    unique_name = f"user_{user_id}_{int(time.time())}_{base_name}.{ext}"
                    avatar_meta = (
                        avatar_bytes,
                        unique_name,
                        avatar_file.mimetype
                        or mimetypes.guess_type(avatar_file.filename)[0]
                        or "application/octet-stream",
                    )

        # Alteração de senha
        new_password = request.form.get("new_password", "").strip()
        if new_password:
            current_password = request.form.get("current_password", "").strip()
            confirm_password = request.form.get("confirm_password", "").strip()
            if not current_password:
                errors.append("Informe sua senha atual para definir uma nova.")
            elif len(new_password) < 6:
                errors.append("A nova senha deve ter pelo menos 6 caracteres.")
            elif new_password != confirm_password:
                errors.append("A confirmação da nova senha não confere.")
            else:
                cursor.execute("SELECT password FROM users_new WHERE id = %s", (user_id,))
                stored = cursor.fetchone()
                if not stored or not bcrypt.checkpw(
                    current_password.encode("utf-8"), _as_bytes(stored["password"])
                ):
                    errors.append("Senha atual incorreta.")
                else:
                    password_hash = bcrypt.hashpw(new_password.encode("utf-8"), bcrypt.gensalt())
                    updates.append("password = %s")
                    params.append(psycopg2.Binary(password_hash))
                    updates.append("first_login = FALSE")

        new_avatar_url = None
        if not errors and avatar_meta:
            try:
                avatar_buffer = BytesIO(avatar_meta[0])
                new_avatar_url, _ = upload_to_supabase(
                    avatar_buffer, avatar_meta[1], avatar_meta[2], folder="avatars"
                )
                updates.append("avatar_url = %s")
                params.append(new_avatar_url)
            except Exception as avatar_exc:
                app.logger.error(f"Erro ao enviar avatar: {avatar_exc}")
                errors.append("Não foi possível salvar a nova foto de perfil. Tente novamente.")

        if errors:
            for error in errors:
                flash(error, "error")
        else:
            if updates:
                updates.append("updated_at = NOW()")
                query = f"UPDATE users_new SET {', '.join(updates)} WHERE id = %s"
                params.append(user_id)
                try:
                    cursor.execute(query, params)
                    conn.commit()
                    user = refresh_session_user_cache() or get_user_by_id(user_id)
                    flash("Perfil atualizado com sucesso!", "success")
                except Exception as exc:
                    conn.rollback()
                    app.logger.error(f"Erro ao atualizar perfil do usuário {user_id}: {exc}")
                    flash("Erro ao atualizar perfil. Tente novamente.", "error")
            else:
                flash("Nenhuma alteração detectada.", "info")

        cursor.close()

    return render_template(get_template("profile.html"), user=user)


@app.route("/profile/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        current_password = request.form.get("current_password", "").strip()
        new_password = request.form.get("new_password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()

        if not current_password or not new_password or not confirm_password:
            flash("Preencha todos os campos.", "error")
            return render_template("change_password.html")

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT password FROM users_new WHERE id = %s", (session["user_id"],)
        )
        result = cursor.fetchone()
        conn.close()

        if not result or not bcrypt.checkpw(
            current_password.encode("utf-8"), _as_bytes(result["password"])
        ):
            flash("Senha atual incorreta.", "error")
            return render_template("change_password.html")

        if len(new_password) < 6:
            flash("A nova senha deve conter pelo menos 6 caracteres.", "error")
            return render_template("change_password.html")

        if new_password != confirm_password:
            flash("As senhas não coincidem.", "error")
            return render_template("change_password.html")

        if update_user_password(session["user_id"], new_password):
            flash("Senha alterada com sucesso!", "success")
            return redirect(url_for("profile"))

        flash("Erro ao alterar senha. Tente novamente.", "error")

    return render_template("change_password.html")


@app.route("/admin/dashboard")
@login_required
@admin_required
def admin_dashboard():
    users = fetch_users()
    dashboards = fetch_dashboards()
    dashboard_map = get_user_dashboard_map()
    for user in users:
        dashboard_map.setdefault(user["id"], {"ids": [], "items": []})

    selected_user_id = request.args.get("user_id", type=int)
    if selected_user_id is None and users:
        selected_user_id = users[0]["id"]

    selected_tab = (request.args.get("tab") or "overview").strip().lower()
    relatorio_meta = fetch_relatorio_meta_settings()

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) as count FROM users_new WHERE is_active = true")
    total_users = cursor.fetchone()["count"]
    cursor.execute(
        "SELECT COUNT(*) as count FROM users_new WHERE is_active = true AND role = 'admin'"
    )
    total_admins = cursor.fetchone()["count"]
    cursor.execute("SELECT COUNT(*) as count FROM dashboards WHERE is_active = true")
    active_dashboards = cursor.fetchone()["count"]
    cursor.execute("SELECT COUNT(*) as count FROM user_dashboards")
    total_assignments = cursor.fetchone()["count"]
    cursor.execute(
        "SELECT status, created_at FROM planner_sync_logs ORDER BY created_at DESC LIMIT 1"
    )
    last_sync = cursor.fetchone()
    conn.close()

    stats = {
        "total_users": total_users,
        "total_admins": total_admins,
        "active_dashboards": active_dashboards,
        "total_assignments": total_assignments,
        "last_sync": last_sync["created_at"] if last_sync else None,
        "last_sync_status": last_sync["status"] if last_sync else None,
    }

    return render_template(
        get_template("admin_dashboard.html"),
        stats=stats,
        users=users,
        dashboards=dashboards,
        selected_user_id=selected_user_id,
        user_dashboards=dashboard_map,
        planner_enabled=planner_client.is_configured,
        planner_logs=get_recent_planner_logs(),
        selected_tab=selected_tab,
        relatorio_meta=relatorio_meta,
    )


@app.route("/admin/dashboard/permissions", methods=["POST"])
@login_required
@admin_required
def update_dashboard_permissions():
    user_id = request.form.get("user_id", type=int)
    if not user_id:
        flash("Selecione um usuário.", "error")
        return redirect(url_for("admin_dashboard"))

    dashboard_ids = request.form.getlist("dashboards")
    dashboard_ids_int = [int(d_id) for d_id in dashboard_ids]
    save_user_dashboards(user_id, dashboard_ids_int, session["user_id"])
    flash("Visibilidade atualizada com sucesso!", "success")
    return redirect(url_for("admin_dashboard", user_id=user_id))


@app.route("/admin/dashboard/meta", methods=["POST"])
@login_required
@admin_required
def update_relatorio_meta_settings():
    meta_valor = request.form.get("meta_valor", type=float)
    meta_percentual = request.form.get("meta_percentual", type=float)
    if meta_valor is None or meta_percentual is None:
        flash("Informe os valores de meta e percentual.", "error")
        return redirect(url_for("admin_dashboard", tab="metas"))

    save_relatorio_meta_settings(meta_valor, meta_percentual, session.get("user_id"))
    flash("Metas do Relatório de Resultados atualizadas!", "success")
    return redirect(url_for("admin_dashboard", tab="metas"))


@app.route("/api/relatorio/layout", methods=["GET", "POST"])
@login_required
def relatorio_layout_api():
    user_id = session.get("user_id")
    if request.method == "GET":
        layout = fetch_relatorio_layout(user_id)
        return jsonify({"layout": layout})

    payload = request.get_json() or {}
    layout = payload.get("layout") or {}
    if not isinstance(layout, dict):
        return jsonify({"error": "layout inválido"}), 400
    save_relatorio_layout(user_id, layout)
    return jsonify({"ok": True})


@app.route("/admin/users")
@login_required
@admin_required
def admin_users():
    """Página de gerenciamento de usuários"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute(
            """
            SELECT id, username, nome_completo, email, nome_usuario, role, 
                   departamento, is_active, first_login, created_at
            FROM users_new
            ORDER BY nome_completo
            """
        )
        users = [dict(row) for row in cursor.fetchall()]
        
        cursor.execute(
            """
            SELECT DISTINCT departamento 
            FROM users_new 
            WHERE departamento IS NOT NULL
            ORDER BY departamento
            """
        )
        departments = [row["departamento"] for row in cursor.fetchall()]
        
        return render_template(
            get_template("admin_users.html"),
            users=users,
            departments=departments
        )
    finally:
        conn.close()


@app.route("/api/agent/knowledge/ingest-documents", methods=["POST"])
@login_required
def ingest_documents_route():
    if session.get("role") != "admin":
        return jsonify({"error": "Apenas administradores podem executar a ingestão"}), 403

    try:
        from scripts.doc_ingest import ingest_documents
        summary = ingest_documents()
        return jsonify({"success": True, "summary": summary})
    except Exception as e:
        app.logger.error(f"[AGENT-KB] Erro na ingestão de documentos: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/admin/users/add", methods=["POST"])
@login_required
@admin_required
def add_user():
    """Adiciona um novo usuário"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        nome_completo = request.form.get("nome_completo", "").strip()
        email = request.form.get("email", "").strip()
        nome_usuario = request.form.get("nome_usuario", "").strip()
        role = request.form.get("role", "usuario").strip()
        password = request.form.get("password", "").strip()
        departamento = request.form.get("departamento", "").strip()
        cargo_original = request.form.get("cargo_original", "").strip()
        
        if not nome_completo or not email:
            flash("Nome completo e email são obrigatórios.", "error")
            return redirect(url_for("admin_users"))
        
        # Valida email
        if not email.lower().endswith("@portoex.com.br"):
            flash("O email deve terminar com @portoex.com.br", "error")
            return redirect(url_for("admin_users"))
        
        # Gera senha padrão se não fornecida
        if not password:
            password = "portoex123"  # Senha padrão
            first_login = True
        else:
            first_login = False
        
        # Verifica se email já existe
        cursor.execute("SELECT id FROM users_new WHERE LOWER(email) = LOWER(%s)", (email,))
        if cursor.fetchone():
            flash("Email já cadastrado.", "error")
            return redirect(url_for("admin_users"))
        
        # Verifica se nome_usuario já existe (se fornecido)
        if nome_usuario:
            cursor.execute("SELECT id FROM users_new WHERE LOWER(nome_usuario) = LOWER(%s)", (nome_usuario,))
            if cursor.fetchone():
                flash("Nome de usuário já cadastrado.", "error")
                return redirect(url_for("admin_users"))
        
        # Hash da senha
        password_hash = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt())
        username = email.lower()
        
        # Insere usuário
        cursor.execute(
            """
            INSERT INTO users_new (
                username, password, nome_completo, email, nome_usuario,
                role, departamento, cargo_original, is_active, first_login
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                username,
                psycopg2.Binary(password_hash),
                nome_completo,
                email.lower(),
                nome_usuario.lower() if nome_usuario else None,
                role if role in ["admin", "usuario"] else "usuario",
                departamento or None,
                cargo_original or None,
                True,
                first_login,
            ),
        )
        conn.commit()
        flash("Usuário adicionado com sucesso!", "success")
        return redirect(url_for("admin_users"))
    except Exception as exc:
        conn.rollback()
        app.logger.error(f"Erro ao adicionar usuário: {exc}", exc_info=True)
        flash(f"Erro ao adicionar usuário: {exc}", "error")
        return redirect(url_for("admin_users"))
    finally:
        conn.close()


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@login_required
@admin_required
def delete_user(user_id):
    """Exclui um usuário (soft delete - desativa)"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verifica se o usuário existe
        cursor.execute("SELECT id, nome_completo FROM users_new WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        if not user:
            flash("Usuário não encontrado.", "error")
            return redirect(url_for("admin_users"))
        
        # Não permite excluir a si mesmo
        if user_id == session.get("user_id"):
            flash("Você não pode excluir seu próprio usuário.", "error")
            return redirect(url_for("admin_users"))
        
        # Soft delete - desativa o usuário
        cursor.execute(
            "UPDATE users_new SET is_active = FALSE, updated_at = CURRENT_TIMESTAMP WHERE id = %s",
            (user_id,),
        )
        conn.commit()
        flash(f"Usuário {user['nome_completo']} foi desativado com sucesso!", "success")
        return redirect(url_for("admin_users"))
    except Exception as exc:
        conn.rollback()
        app.logger.error(f"Erro ao excluir usuário: {exc}", exc_info=True)
        flash(f"Erro ao excluir usuário: {exc}", "error")
        return redirect(url_for("admin_users"))
    finally:
        conn.close()


@app.route("/admin/users/<int:user_id>/update", methods=["POST"])
@login_required
@admin_required
def update_user(user_id):
    """Atualiza dados do usuário"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verifica se o usuário existe
        cursor.execute("SELECT id FROM users_new WHERE id = %s", (user_id,))
        if not cursor.fetchone():
            flash("Usuário não encontrado.", "error")
            return redirect(url_for("admin_users"))
        
        # Coleta dados do formulário
        new_email = request.form.get("email", "").strip()
        new_nome_usuario = request.form.get("nome_usuario", "").strip()
        new_role = request.form.get("role", "usuario").strip()
        new_password = request.form.get("password", "").strip()
        reset_first_login = request.form.get("reset_first_login") == "on"
        
        # Novos campos
        new_nome_completo = request.form.get("nome_completo", "").strip()
        new_cargo = request.form.get("cargo_original", "").strip()
        new_departamento = request.form.get("departamento", "").strip()
        new_is_active = request.form.get("is_active") == "true"
        
        # Valida email
        if new_email and not new_email.lower().endswith("@portoex.com.br"):
            flash("O email deve terminar com @portoex.com.br", "error")
            return redirect(url_for("admin_users"))
        
        # Atualiza campos
        updates = []
        params = []
        
        if new_nome_completo:
            updates.append("nome_completo = %s")
            params.append(new_nome_completo)
            
        # Cargo e departamento podem ser vazios
        updates.append("cargo_original = %s")
        params.append(new_cargo)
        
        updates.append("departamento = %s")
        params.append(new_departamento)
        
        # Status
        updates.append("is_active = %s")
        params.append(new_is_active)
        
        if new_email:
            updates.append("email = %s")
            params.append(new_email.lower())
        
        if new_nome_usuario:
            updates.append("nome_usuario = %s")
            params.append(new_nome_usuario.lower())
        
        if new_role in ["admin", "usuario"]:
            updates.append("role = %s")
            params.append(new_role)
        
        if new_password:
            password_hash = bcrypt.hashpw(
                new_password.encode("utf-8"), bcrypt.gensalt()
            )
            updates.append("password = %s")
            params.append(psycopg2.Binary(password_hash))
            updates.append("first_login = FALSE")
        
        if reset_first_login:
            updates.append("first_login = TRUE")
        
        if updates:
            updates.append("updated_at = CURRENT_TIMESTAMP")
            params.append(user_id)
            
            query = f"UPDATE users_new SET {', '.join(updates)} WHERE id = %s"
            cursor.execute(query, params)
            conn.commit()
            flash("Usuário atualizado com sucesso!", "success")
        else:
            flash("Nenhuma alteração foi feita.", "info")
        
        return redirect(url_for("admin_users"))
    except Exception as exc:
        conn.rollback()
        app.logger.error(f"Erro ao atualizar usuário: {exc}", exc_info=True)
        flash(f"Erro ao atualizar usuário: {exc}", "error")
        return redirect(url_for("admin_users"))
    finally:
        conn.close()


@app.route("/admin/planner/sync", methods=["POST"])
@login_required
@admin_required
def admin_planner_sync():
    try:
        success_count, errors = sync_dashboards_to_planner()
        if success_count:
            flash(
                f"Agenda enviada para {success_count} usuário(s) no Planner.",
                "success",
            )
        if errors:
            flash("Falha para: " + "; ".join(errors), "error")
    except PlannerIntegrationError as exc:
        flash(str(exc), "error")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/environments")
@login_required
@admin_required
def admin_environments():
    """Página para gerenciar ambientes do CD"""
    return render_template(get_template("admin_environments.html"))


@app.route("/admin/dashboards/add", methods=["GET", "POST"])
@login_required
@admin_required
def admin_add_dashboard():
    """Adiciona ou edita um dashboard do BI"""
    if request.method == "POST":
        conn = get_db()
        cursor = conn.cursor()
        
        try:
            dashboard_id = request.form.get("dashboard_id", type=int)
            title = request.form.get("title", "").strip()
            description = request.form.get("description", "").strip()
            category = request.form.get("category", "").strip()
            embed_code = request.form.get("embed_code", "").strip()
            display_order = request.form.get("display_order", type=int) or 0
            
            if not title or not embed_code:
                flash("Nome e código de incorporação são obrigatórios.", "error")
                return redirect(url_for("admin_add_dashboard", dashboard_id=dashboard_id) if dashboard_id else url_for("admin_add_dashboard"))
            
            # Extrai URL do iframe
            url_match = re.search(r'src="([^"]+)"', embed_code)
            if not url_match:
                flash("Código de incorporação inválido. Deve conter um iframe com src.", "error")
                return redirect(url_for("admin_add_dashboard", dashboard_id=dashboard_id) if dashboard_id else url_for("admin_add_dashboard"))
            
            embed_url = url_match.group(1)
            
            # Gera slug do título
            slug = title.lower().replace(" ", "-").replace("ç", "c").replace("ã", "a").replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
            slug = re.sub(r'[^a-z0-9-]', '', slug)
            
            if dashboard_id:
                # Atualiza dashboard existente
                cursor.execute(
                    """
                    UPDATE dashboards
                    SET title = %s, description = %s, category = %s, 
                        embed_url = %s, display_order = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                    """,
                    (title, description, category, embed_url, display_order, dashboard_id),
                )
                flash("Dashboard atualizado com sucesso!", "success")
            else:
                # Cria novo dashboard
                cursor.execute(
                    """
                    INSERT INTO dashboards (slug, title, description, category, embed_url, display_order)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (slug) DO UPDATE SET
                        title = EXCLUDED.title,
                        description = EXCLUDED.description,
                        category = EXCLUDED.category,
                        embed_url = EXCLUDED.embed_url,
                        display_order = EXCLUDED.display_order,
                        updated_at = CURRENT_TIMESTAMP
                    """,
                    (slug, title, description, category, embed_url, display_order),
                )
                flash("Dashboard adicionado com sucesso!", "success")
            
            conn.commit()
            return redirect(url_for("admin_dashboard"))
        except Exception as exc:
            conn.rollback()
            app.logger.error(f"Erro ao salvar dashboard: {exc}", exc_info=True)
            flash(f"Erro ao salvar dashboard: {exc}", "error")
            return redirect(url_for("admin_add_dashboard"))
        finally:
            conn.close()
    
    # GET: mostra formulário
    dashboard_id = request.args.get("dashboard_id", type=int)
    dashboard = None
    
    if dashboard_id:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM dashboards WHERE id = %s", (dashboard_id,))
        dashboard = cursor.fetchone()
        conn.close()
        
        if not dashboard:
            flash("Dashboard não encontrado.", "error")
            return redirect(url_for("admin_dashboard"))
    
    return render_template(get_template("admin_add_dashboard.html"), dashboard=dashboard)


@app.route("/admin/assets")
@login_required
@admin_required
def admin_assets():
    assets = fetch_assets()
    return render_template(get_template("admin_assets.html"), assets=assets)


@app.route("/admin/assets/add", methods=["GET", "POST"])
@login_required
@admin_required
def admin_add_asset():
    if request.method == "POST":
        conn = get_db()
        cursor = conn.cursor()
        try:
            asset_id = request.form.get("asset_id", type=int)
            nome = request.form.get("nome", "").strip()
            tipo = normalize_asset_type(request.form.get("tipo", ""))
            categoria = request.form.get("categoria", "").strip() or None
            descricao = request.form.get("descricao", "").strip() or None
            status = normalize_asset_status(request.form.get("status", ""))
            ordem_padrao = request.form.get("ordem_padrao", type=int) or 0
            embed_input = request.form.get("embed_code", "").strip()
            resource_url = request.form.get("resource_url", "").strip() or None
            config_text = request.form.get("config_json", "").strip()

            if not nome or not tipo:
                flash("Nome e tipo do ativo são obrigatórios.", "error")
                return redirect(
                    url_for("admin_add_asset", asset_id=asset_id)
                    if asset_id
                    else url_for("admin_add_asset")
                )

            if tipo not in ASSET_TYPES:
                flash("Tipo inválido. Use PBI, interno, gráfico, rpa ou outro.", "error")
                return redirect(url_for("admin_add_asset", asset_id=asset_id) if asset_id else url_for("admin_add_asset"))

            if status not in ASSET_STATUSES:
                flash("Status inválido. Use ativo, legado ou experimental.", "error")
                return redirect(url_for("admin_add_asset", asset_id=asset_id) if asset_id else url_for("admin_add_asset"))

            embed_url = extract_embed_url(embed_input) if embed_input else None
            if tipo == "PBI":
                if not embed_url:
                    flash("O código de incorporação do Power BI é obrigatório.", "error")
                    return redirect(url_for("admin_add_asset", asset_id=asset_id) if asset_id else url_for("admin_add_asset"))
                if not is_safe_embed_url(embed_url):
                    flash("O embed do Power BI não pode conter tokens sensíveis.", "error")
                    return redirect(url_for("admin_add_asset", asset_id=asset_id) if asset_id else url_for("admin_add_asset"))

            config_data = None
            if config_text:
                try:
                    config_data = json.loads(config_text)
                except json.JSONDecodeError:
                    flash("Configuração JSON inválida. Verifique a formatação.", "error")
                    return redirect(url_for("admin_add_asset", asset_id=asset_id) if asset_id else url_for("admin_add_asset"))

            if asset_id:
                cursor.execute(
                    """
                    UPDATE assets
                    SET nome = %s,
                        tipo = %s,
                        categoria = %s,
                        descricao = %s,
                        status = %s,
                        ordem_padrao = %s,
                        embed_url = %s,
                        resource_url = %s,
                        config = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                    """,
                    (
                        nome,
                        tipo,
                        categoria,
                        descricao,
                        status,
                        ordem_padrao,
                        embed_url,
                        resource_url,
                        psycopg2.extras.Json(config_data) if config_data is not None else None,
                        asset_id,
                    ),
                )
                log_asset_action(
                    action_type="update",
                    asset_id=asset_id,
                    actor_id=session.get("user_id"),
                    details={"nome": nome, "tipo": tipo, "status": status},
                )
                flash("Ativo atualizado com sucesso!", "success")
            else:
                cursor.execute(
                    """
                    INSERT INTO assets (
                        nome, tipo, categoria, descricao, status, ordem_padrao,
                        embed_url, resource_url, config
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (
                        nome,
                        tipo,
                        categoria,
                        descricao,
                        status,
                        ordem_padrao,
                        embed_url,
                        resource_url,
                        psycopg2.extras.Json(config_data) if config_data is not None else None,
                    ),
                )
                new_id = cursor.fetchone()["id"]
                log_asset_action(
                    action_type="create",
                    asset_id=new_id,
                    actor_id=session.get("user_id"),
                    details={"nome": nome, "tipo": tipo, "status": status},
                )
                flash("Ativo cadastrado com sucesso!", "success")

            conn.commit()
            return redirect(url_for("admin_assets"))
        except Exception as exc:
            conn.rollback()
            app.logger.error(f"Erro ao salvar ativo: {exc}", exc_info=True)
            flash(f"Erro ao salvar ativo: {exc}", "error")
            return redirect(url_for("admin_add_asset"))
        finally:
            conn.close()

    asset_id = request.args.get("asset_id", type=int)
    asset = fetch_asset_by_id(asset_id) if asset_id else None
    if asset_id and not asset:
        flash("Ativo não encontrado.", "error")
        return redirect(url_for("admin_assets"))
    return render_template(get_template("admin_asset_form.html"), asset=asset)


@app.route("/admin/assets/<int:asset_id>/delete", methods=["POST"])
@login_required
@admin_required
def admin_delete_asset(asset_id: int):
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, nome FROM assets WHERE id = %s", (asset_id,))
        asset = cursor.fetchone()
        if not asset:
            flash("Ativo não encontrado.", "error")
            return redirect(url_for("admin_assets"))

        cursor.execute("DELETE FROM assets WHERE id = %s", (asset_id,))
        conn.commit()
        log_asset_action(
            action_type="delete",
            asset_id=asset_id,
            actor_id=session.get("user_id"),
            details={"nome": asset["nome"]},
        )
        flash("Ativo removido com sucesso.", "success")
        return redirect(url_for("admin_assets"))
    except Exception as exc:
        conn.rollback()
        app.logger.error(f"Erro ao remover ativo: {exc}", exc_info=True)
        flash(f"Erro ao remover ativo: {exc}", "error")
        return redirect(url_for("admin_assets"))
    finally:
        conn.close()


@app.route("/admin/assets/assignments", methods=["GET", "POST"])
@login_required
@admin_required
def admin_asset_assignments():
    if request.method == "POST":
        target_type = request.form.get("target_type", "user")
        if target_type == "user":
            target_value = request.form.get("user_id", type=int)
            if not target_value:
                flash("Selecione um usuário para atribuição.", "error")
                return redirect(url_for("admin_asset_assignments"))
            target_value_str = str(target_value)
        else:
            target_value = request.form.get("group_name", "").strip()
            if not target_value:
                flash("Selecione um grupo para atribuição.", "error")
                return redirect(url_for("admin_asset_assignments", target_type="group"))
            target_value_str = target_value

        asset_ids = request.form.getlist("asset_ids")
        asset_meta: Dict[int, Dict] = {}
        if asset_ids:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute(
                "SELECT id, tipo, config FROM assets WHERE id = ANY(%s)",
                (list(map(int, asset_ids)),),
            )
            for row in cursor.fetchall():
                asset_meta[row["id"]] = {"tipo": row["tipo"], "config": row.get("config") or {}}
            conn.close()

        assignments: List[Dict] = []
        for asset_id in asset_ids:
            enabled = request.form.get(f"asset_{asset_id}_enabled") == "on"
            if not enabled:
                continue
            ordem = request.form.get(f"asset_{asset_id}_order", type=int) or 0
            visivel = request.form.get(f"asset_{asset_id}_visible") == "on"
            config = None
            meta = asset_meta.get(int(asset_id), {})
            if meta.get("tipo") == "interno":
                available_components = (meta.get("config") or {}).get("available_components")
                if available_components:
                    components = request.form.getlist(f"asset_{asset_id}_components")
                    if components:
                        config = {"components": components}
                else:
                    chart_types = request.form.getlist(f"asset_{asset_id}_chart_types")
                    if chart_types:
                        config = {"chart_types": chart_types}
            assignments.append(
                {
                    "asset_id": int(asset_id),
                    "ordem": ordem,
                    "visivel": visivel,
                    "config": psycopg2.extras.Json(config) if config else None,
                }
            )

        save_asset_assignments(target_type, target_value_str, assignments, session["user_id"])
        log_asset_action(
            action_type="assignments_update",
            asset_id=None,
            actor_id=session.get("user_id"),
            target_type=target_type,
            target_value=target_value_str,
            details={"total": len(assignments)},
        )
        flash("Atribuições salvas com sucesso!", "success")

        return redirect(
            url_for(
                "admin_asset_assignments",
                target_type=target_type,
                user_id=target_value_str if target_type == "user" else None,
                group_name=target_value_str if target_type == "group" else None,
            ),
            code=303,
        )

    target_type = request.args.get("target_type", "user")
    selected_user_id = request.args.get("user_id", type=int)
    selected_group = request.args.get("group_name", "").strip() or None

    users = fetch_users()
    if target_type == "user" and selected_user_id is None and users:
        selected_user_id = users[0]["id"]

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT DISTINCT departamento
        FROM users_new
        WHERE departamento IS NOT NULL
        ORDER BY departamento
        """
    )
    groups = [row["departamento"] for row in cursor.fetchall()]
    conn.close()
    if target_type == "group" and not selected_group and groups:
        selected_group = groups[0]

    assets = fetch_assets()
    assignments = []
    preview_label = None
    if target_type == "user" and selected_user_id:
        assignments = fetch_asset_assignments("user", str(selected_user_id), include_inactive=True)
        user = next((u for u in users if u["id"] == selected_user_id), None)
        preview_label = user["nome_completo"] if user else None
    elif target_type == "group" and selected_group:
        assignments = fetch_asset_assignments("group", selected_group, include_inactive=True)
        preview_label = selected_group

    assignment_map = {item["asset_id"]: item for item in assignments}
    preview_url = None
    if target_type == "user" and selected_user_id:
        preview_url = url_for("team_dashboard", preview_user_id=selected_user_id)
    elif target_type == "group" and selected_group:
        preview_url = url_for("team_dashboard", preview_group=selected_group)

    return render_template(
        get_template("admin_asset_assignments.html"),
        users=users,
        groups=groups,
        assets=assets,
        target_type=target_type,
        selected_user_id=selected_user_id,
        selected_group=selected_group,
        assignment_map=assignment_map,
        preview_url=preview_url,
        preview_label=preview_label,
    )


@app.route("/favicon.ico")
def favicon():
    """Serve o favicon"""
    return send_from_directory(
        app.static_folder,
        "test_favicon.ico",
        mimetype="image/x-icon"
    )


@app.route("/team/dashboard")
@app.route("/dashboards")
@login_required
def team_dashboard():
    show_all = is_admin_session() and request.args.get("all") == "1"
    preview_user_id = request.args.get("preview_user_id", type=int)
    preview_group = request.args.get("preview_group", "").strip() or None

    effective_user_id = session.get("user_id")
    effective_department = session.get("departamento")
    preview_label = None

    if is_admin_session():
        if preview_user_id:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute(
                "SELECT id, nome_completo, departamento FROM users_new WHERE id = %s",
                (preview_user_id,),
            )
            preview_user = cursor.fetchone()
            conn.close()
            if preview_user:
                effective_user_id = preview_user["id"]
                effective_department = preview_user.get("departamento")
                preview_label = preview_user.get("nome_completo")
            else:
                flash("Usuário de pré-visualização não encontrado.", "error")
        elif preview_group:
            effective_user_id = None
            effective_department = preview_group
            preview_label = preview_group

    assets: List[Dict] = []
    if show_all and is_admin_session():
        assets = fetch_assets(["ativo", "experimental"])
    else:
        assets = fetch_assets_for_user(effective_user_id, effective_department)

    if not assets and effective_user_id is not None:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT d.title, d.description, d.category, d.embed_url
            FROM dashboards d
            JOIN user_dashboards ud ON ud.dashboard_id = d.id
            WHERE ud.user_id = %s AND d.is_active = true
            ORDER BY d.display_order, d.title
            """,
            (effective_user_id,),
        )
        dashboards = cursor.fetchall()
        conn.close()
        assets = [
            {
                "id": idx + 1,
                "nome": row["title"],
                "tipo": "PBI",
                "categoria": row.get("category"),
                "descricao": row.get("description"),
                "embed_url": row.get("embed_url"),
                "resource_url": None,
                "status": "ativo",
                "ordem_padrao": idx,
                "config": None,
                "assignment_source": "legado",
            }
            for idx, row in enumerate(dashboards or [])
        ]
        # Fallback: se nem atribuicoes nem legado, exibir Relatorio de Resultados com todos os componentes
        if not assets:
            default_asset = _fetch_default_relatorio_asset()
            if default_asset:
                assets = [default_asset]

    today_label = datetime.now().strftime("%d/%m/%Y")
    chart_only_types = {"table", "card"}
    has_chart_assets = any(
        (a.get("tipo") == "grafico" and (a.get("asset_config") or {}).get("chart"))
        or (
            a.get("tipo") == "interno"
            and (a.get("asset_config") or {}).get("chart")
            and any(
                t not in chart_only_types
                for t in ((a.get("assignment_config") or {}).get("chart_types") or [])
            )
        )
        for a in assets
    )
    internal_assets = []
    regular_assets = []
    for asset in assets:
        if (
            asset.get("tipo") == "interno"
            and (asset.get("asset_config") or {}).get("embed_mode")
            and (asset.get("assignment_config") or {}).get("components")
        ):
            internal_assets.append(asset)
        else:
            regular_assets.append(asset)

    relatorio_meta = fetch_relatorio_meta_settings()

    return render_template(
        get_template("team_dashboard.html"),
        user=session.get("nome_completo", "Usuário"),
        assets=assets,
        internal_assets=internal_assets,
        regular_assets=regular_assets,
        today=today_label,
        show_all=show_all,
        is_admin=is_admin_session(),
        preview_label=preview_label,
        has_chart_assets=has_chart_assets,
        relatorio_meta=relatorio_meta,
    )


@app.route("/cd/facilities")
@login_required
def cd_facilities():
    """Página para visualizar planta 3D do CD"""
    conn = get_db()
    cursor = conn.cursor()
    try:
        # Buscar ambientes ordenados por display_order
        cursor.execute("""
            SELECT id, code, name, description, icon, floor
            FROM environments 
            WHERE is_active = true 
            ORDER BY display_order ASC
        """)
        environments = [dict(row) for row in cursor.fetchall()]
        
        # Buscar recursos (modelos 3D, fotos, plantas) para cada ambiente
        # Isso é importante para que o JS saiba o que carregar
        cursor.execute("""
            SELECT environment_id, resource_type, file_url, is_primary
            FROM environment_resources
            ORDER BY is_primary DESC
        """)
        resources = [dict(row) for row in cursor.fetchall()]
        
        # Agrupar recursos por ambiente
        env_resources = {}
        for res in resources:
            env_id = res['environment_id']
            if env_id not in env_resources:
                env_resources[env_id] = []
            env_resources[env_id].append(res)
            
        # Adicionar recursos aos ambientes
        for env in environments:
            env['resources'] = env_resources.get(env['id'], [])
            
    except Exception as e:
        app.logger.error(f"Erro ao carregar ambientes: {str(e)}")
        environments = []
    finally:
        conn.close()
        
    return render_template(get_template("cd_facilities.html"), environments=environments)


@app.route("/cd/booking")
@login_required
def cd_booking():
    """Página para agendar salas de reunião"""
    return render_template(get_template("cd_booking.html"))


@app.route("/api/room-bookings", methods=["GET", "POST"])
@login_required
def room_bookings_api():
    """API para listar e criar agendamentos de salas"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        if request.method == "GET":
            cursor.execute("""
                SELECT 
                    rb.id, rb.room, rb.title, rb.date, rb.start_time, rb.end_time,
                    rb.participants, rb.subject, rb.created_at,
                    u.nome_completo as user_name
                FROM room_bookings rb
                JOIN users_new u ON rb.user_id = u.id
                WHERE rb.is_active = true
                ORDER BY rb.date DESC, rb.start_time DESC
            """)
            bookings = cursor.fetchall()
            
            # Serializar objetos de data/hora
            result = []
            for row in bookings:
                booking = dict(row)
                booking['date'] = str(booking['date'])
                booking['start_time'] = str(booking['start_time'])
                booking['end_time'] = str(booking['end_time'])
                booking['created_at'] = str(booking['created_at'])
                result.append(booking)
                
            return jsonify(result)
        
        elif request.method == "POST":
            data = request.get_json()
            
            required_fields = ['room', 'title', 'date', 'start_time', 'end_time', 'participants', 'subject']
            for field in required_fields:
                if field not in data:
                    return jsonify({"error": f"Campo obrigatório ausente: {field}"}), 400
            
            user_id = session.get('user_id')
            if not user_id:
                app.logger.error("Tentativa de agendamento sem user_id na sessão")
                return jsonify({"error": "Sessão inválida. Faça login novamente."}), 401
                
            app.logger.info(f"Tentando criar agendamento: User={user_id}, Room={data['room']}, Date={data['date']}")
            
            cursor.execute("""
                SELECT id FROM room_bookings
                WHERE room = %s AND date = %s AND is_active = true
                AND (
                    (start_time <= %s AND end_time > %s) OR
                    (start_time < %s AND end_time >= %s) OR
                    (start_time >= %s AND end_time <= %s)
                )
            """, (
                data['room'], data['date'],
                data['start_time'], data['start_time'],
                data['end_time'], data['end_time'],
                data['start_time'], data['end_time']
            ))
            
            conflict = cursor.fetchone()
            if conflict:
                return jsonify({"error": "Já existe um agendamento neste horário para esta sala"}), 409
            
            cursor.execute("""
                INSERT INTO room_bookings 
                (user_id, room, title, date, start_time, end_time, participants, subject)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                user_id,
                data['room'],
                data['title'],
                data['date'],
                data['start_time'],
                data['end_time'],
                data['participants'],
                data['subject']
            ))
            
            booking_id = cursor.fetchone()['id']
            conn.commit()
            
            app.logger.info(f"Agendamento criado com sucesso: ID={booking_id}")
            return jsonify({"success": True, "id": booking_id}), 201
    
    except Exception as e:
        conn.rollback()
        import traceback
        error_details = traceback.format_exc()
        app.logger.error(f"Erro ao criar agendamento: {str(e)}\n{error_details}")
        # Retorna o erro detalhado apenas em debug, ou uma mensagem genérica
        return jsonify({"error": f"Erro interno: {str(e)}"}), 500
    finally:
        conn.close()


@app.route("/api/room-bookings/<int:booking_id>", methods=["GET", "PUT", "DELETE"])
@login_required
def room_booking_detail_api(booking_id):
    """API para obter, atualizar ou deletar um agendamento específico"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        if request.method == "GET":
            cursor.execute("""
                SELECT 
                    rb.id, rb.room, rb.title, rb.date, rb.start_time, rb.end_time,
                    rb.participants, rb.subject, rb.created_at, rb.user_id,
                    u.nome_completo as user_name
                FROM room_bookings rb
                JOIN users_new u ON rb.user_id = u.id
                WHERE rb.id = %s AND rb.is_active = true
            """, (booking_id,))
            
            booking = cursor.fetchone()
            if not booking:
                return jsonify({"error": "Agendamento não encontrado"}), 404
            
            # Serializar
            result = dict(booking)
            result['date'] = str(result['date'])
            result['start_time'] = str(result['start_time'])
            result['end_time'] = str(result['end_time'])
            result['created_at'] = str(result['created_at'])
            
            return jsonify(result)
        
        elif request.method == "PUT":
            cursor.execute(
                "SELECT user_id FROM room_bookings WHERE id = %s AND is_active = true",
                (booking_id,)
            )
            booking = cursor.fetchone()
            
            if not booking:
                return jsonify({"error": "Agendamento não encontrado"}), 404
            
            if booking['user_id'] != session['user_id'] and session.get('role') != 'admin':
                return jsonify({"error": "Sem permissão para editar este agendamento"}), 403
            
            data = request.get_json()
            
            cursor.execute("""
                UPDATE room_bookings
                SET room = %s, title = %s, date = %s, start_time = %s, 
                    end_time = %s, participants = %s, subject = %s
                WHERE id = %s
            """, (
                data.get('room'),
                data.get('title'),
                data.get('date'),
                data.get('start_time'),
                data.get('end_time'),
                data.get('participants'),
                data.get('subject'),
                booking_id
            ))
            
            conn.commit()
            return jsonify({"success": True})
        
        elif request.method == "DELETE":
            cursor.execute(
                "SELECT user_id FROM room_bookings WHERE id = %s AND is_active = true",
                (booking_id,)
            )
            booking = cursor.fetchone()
            
            if not booking:
                return jsonify({"error": "Agendamento não encontrado"}), 404
            
            if booking['user_id'] != session['user_id'] and session.get('role') != 'admin':
                return jsonify({"error": "Sem permissão para excluir este agendamento"}), 403
            
            cursor.execute(
                "UPDATE room_bookings SET is_active = false WHERE id = %s",
                (booking_id,)
            )
            
            conn.commit()
            return jsonify({"success": True})
    
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/3d-model/<model_type>", methods=['GET', 'OPTIONS'])
def proxy_3d_model(model_type):
    """Proxy otimizado para modelos 3D com streaming e CORS"""
    
    # Responder a preflight OPTIONS
    if request.method == 'OPTIONS':
        response = app.make_default_options_response()
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response
    
    app.logger.info(f"[3D PROXY] Requisitando modelo: {model_type}")
    
    urls = {
        'glb': 'https://github.com/anaissiabraao/GeRot/releases/download/v1.0-3d-models/Cd_front_12_50_53.glb',
        'fbx': 'https://github.com/anaissiabraao/GeRot/releases/download/v1.0-3d-models/Cd_front_12_10_17.fbx'
    }
    
    if model_type not in urls:
        app.logger.error(f"[3D PROXY] Modelo inválido: {model_type}")
        return jsonify({"error": "Modelo inválido"}), 404
    
    try:
        # Streaming otimizado: chunks grandes + timeout longo
        app.logger.info(f"[3D PROXY] Baixando: {urls[model_type]}")
        
        # Fazer requisição (sem context manager para não fechar antes do generator terminar)
        # Timeout de 300s (5min) para arquivos grandes
        r = requests.get(urls[model_type], stream=True, timeout=(30, 300))
        r.raise_for_status()
        
        content_length = r.headers.get('content-length', '0')
        content_type = r.headers.get('content-type', 'application/octet-stream')
        
        app.logger.info(f"[3D PROXY] Tamanho do arquivo: {content_length} bytes ({int(content_length)/(1024*1024):.2f} MB)")
        
        def generate():
            """Generator para streaming eficiente com keep-alive"""
            try:
                bytes_sent = 0
                # Chunks de 256KB para velocidade máxima
                for chunk in r.iter_content(chunk_size=262144):
                    if chunk:
                        bytes_sent += len(chunk)
                        yield chunk
                        # Log a cada 10MB para monitorar progresso
                        if bytes_sent % (10 * 1024 * 1024) < 262144:
                            app.logger.info(f"[3D PROXY] {bytes_sent/(1024*1024):.1f}MB enviados")
            finally:
                # Fechar conexão após streaming completo
                r.close()
                app.logger.info(f"[3D PROXY] Streaming concluído: {bytes_sent/(1024*1024):.1f}MB para {model_type}")
        
        # Retornar resposta com CORS, Content-Length e streaming
        from flask import Response
        response = Response(generate(), mimetype=content_type, direct_passthrough=True)
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        response.headers['Content-Length'] = content_length
        response.headers['Cache-Control'] = 'public, max-age=86400'
        response.headers['Connection'] = 'keep-alive'
        response.headers['Keep-Alive'] = 'timeout=300'
        
        app.logger.info(f"[3D PROXY] Streaming iniciado para {model_type}")
        return response
        
    except requests.exceptions.Timeout:
        app.logger.error(f"[3D PROXY] Timeout ao baixar {model_type}")
        return jsonify({"error": "Timeout ao baixar modelo"}), 504
    except requests.exceptions.RequestException as e:
        app.logger.error(f"[3D PROXY] Erro de rede: {str(e)}")
        return jsonify({"error": f"Erro ao baixar modelo: {str(e)}"}), 502
    except Exception as e:
        app.logger.error(f"[3D PROXY] Erro inesperado: {str(e)}")
        return jsonify({"error": str(e)}), 500


# --------------------------------------------------------------------------- #
# API pública básica
# --------------------------------------------------------------------------- #
class UsersAPI(Resource):
    def get(self, user_id=None):
        # Validação de admin para acessar lista de usuários
        if session.get("role") != "admin":
            return jsonify({"error": "Acesso restrito aos administradores"}), 403
        
        conn = get_db()
        cursor = conn.cursor()
        try:
            if user_id:
                cursor.execute(
                    "SELECT id, username, nome_completo, role, departamento FROM users_new WHERE id = %s AND is_active = true",
                    (user_id,),
                )
                user = cursor.fetchone()
                if not user:
                    return jsonify({"error": "User not found"}), 404
                return jsonify({"user": dict(user)})

            cursor.execute(
                "SELECT id, username, nome_completo, role, departamento FROM users_new WHERE is_active = true"
            )
            users = [dict(row) for row in cursor.fetchall()]
            return jsonify({"users": users})
        finally:
            conn.close()


api.add_resource(UsersAPI, "/api/users", "/api/users/<int:user_id>")


# --------------------------------------------------------------------------- #
# API para gerenciar ambientes do CD
# --------------------------------------------------------------------------- #
@app.route("/api/environments", methods=["GET", "POST"])
@login_required
def environments_api():
    """API para listar e criar ambientes"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        if request.method == "GET":
            cursor.execute(""" 
                SELECT 
                    e.id, e.code, e.name, e.description, e.icon, 
                    e.capacity, e.area_m2, e.floor, e.is_active,
                    e.display_order, e.created_at,
                    COUNT(DISTINCT er.id) as resource_count,
                    COUNT(DISTINCT CASE WHEN er.resource_type = 'model_3d' THEN er.id END) as models_3d,
                    COUNT(DISTINCT CASE WHEN er.resource_type = 'plant_2d' THEN er.id END) as plants_2d,
                    COUNT(DISTINCT CASE WHEN er.resource_type = 'photo' THEN er.id END) as photos
                FROM environments e
                LEFT JOIN environment_resources er ON e.id = er.environment_id
                WHERE e.is_active = true
                GROUP BY e.id, e.code, e.name, e.description, e.icon, 
                         e.capacity, e.area_m2, e.floor, e.is_active,
                         e.display_order, e.created_at
                ORDER BY e.display_order, e.name
            """)
            environments = cursor.fetchall()
            return jsonify([dict(row) for row in environments])
        
        elif request.method == "POST":
            # Apenas admins podem criar ambientes
            if session.get("role") != "manager":
                return jsonify({"error": "Apenas administradores podem criar ambientes"}), 403
            
            data = request.get_json()
            
            required_fields = ['code', 'name']
            for field in required_fields:
                if field not in data:
                    return jsonify({"error": f"Campo obrigatório ausente: {field}"}), 400
            
            cursor.execute(""" 
                INSERT INTO environments 
                (code, name, description, icon, capacity, area_m2, floor, display_order)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                data['code'],
                data['name'],
                data.get('description', ''),
                data.get('icon', 'fas fa-building'),
                data.get('capacity'),
                data.get('area_m2'),
                data.get('floor', 1),
                data.get('display_order', 0)
            ))
            
            environment_id = cursor.fetchone()['id']
            conn.commit()
            
            return jsonify({"success": True, "id": environment_id}), 201
    
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Erro ao processar ambientes: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/environments/<int:environment_id>", methods=["GET", "PUT", "DELETE"])
@login_required
def environment_detail_api(environment_id):
    """API para obter, atualizar ou deletar um ambiente específico"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        if request.method == "GET":
            cursor.execute(""" 
                SELECT 
                    e.*, 
                    s.camera_position_x, s.camera_position_y, s.camera_position_z,
                    s.camera_target_x, s.camera_target_y, s.camera_target_z,
                    s.model_scale, s.rotation_speed, s.enable_shadows,
                    s.background_color, s.grid_size
                FROM environments e
                LEFT JOIN environment_3d_settings s ON e.id = s.environment_id
                WHERE e.id = %s AND e.is_active = true
            """, (environment_id,))
            
            environment = cursor.fetchone()
            if not environment:
                return jsonify({"error": "Ambiente não encontrado"}), 404
            
            # Buscar recursos associados
            cursor.execute(""" 
                SELECT * FROM environment_resources 
                WHERE environment_id = %s 
                ORDER BY resource_type, display_order
            """, (environment_id,))
            resources = cursor.fetchall()
            
            result = dict(environment)
            result['resources'] = [dict(r) for r in resources]
            
            return jsonify(result)
        
        elif request.method == "PUT":
            # Apenas admins podem atualizar ambientes
            if session.get("role") != "manager":
                return jsonify({"error": "Apenas administradores podem editar ambientes"}), 403
            
            data = request.get_json()
            
            # Atualizar ambiente
            cursor.execute(""" 
                UPDATE environments 
                SET name = %s, description = %s, icon = %s, 
                    capacity = %s, area_m2 = %s, floor = %s, 
                    display_order = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s AND is_active = true
            """, (
                data.get('name'),
                data.get('description'),
                data.get('icon'),
                data.get('capacity'),
                data.get('area_m2'),
                data.get('floor'),
                data.get('display_order'),
                environment_id
            ))
            
            # Atualizar configurações 3D se fornecidas
            if '3d_settings' in data:
                settings = data['3d_settings']
                
                # Verificar se já existe configuração
                cursor.execute(
                    "SELECT id FROM environment_3d_settings WHERE environment_id = %s",
                    (environment_id,)
                )
                exists = cursor.fetchone()
                
                if exists:
                    cursor.execute(""" 
                        UPDATE environment_3d_settings
                        SET camera_position_x = %s, camera_position_y = %s, camera_position_z = %s,
                            camera_target_x = %s, camera_target_y = %s, camera_target_z = %s,
                            model_scale = %s, rotation_speed = %s, enable_shadows = %s,
                            background_color = %s, grid_size = %s,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE environment_id = %s
                    """, (
                        settings.get('camera_position_x', 5),
                        settings.get('camera_position_y', 5),
                        settings.get('camera_position_z', 5),
                        settings.get('camera_target_x', 0),
                        settings.get('camera_target_y', 0),
                        settings.get('camera_target_z', 0),
                        settings.get('model_scale', 1.0),
                        settings.get('rotation_speed', 0.01),
                        settings.get('enable_shadows', True),
                        settings.get('background_color', '#1a1a2e'),
                        settings.get('grid_size', 20),
                        environment_id
                    ))
                else:
                    cursor.execute(""" 
                        INSERT INTO environment_3d_settings
                        (environment_id, camera_position_x, camera_position_y, camera_position_z,
                         camera_target_x, camera_target_y, camera_target_z,
                         model_scale, rotation_speed, enable_shadows, background_color, grid_size)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        environment_id,
                        settings.get('camera_position_x', 5),
                        settings.get('camera_position_y', 5),
                        settings.get('camera_position_z', 5),
                        settings.get('camera_target_x', 0),
                        settings.get('camera_target_y', 0),
                        settings.get('camera_target_z', 0),
                        settings.get('model_scale', 1.0),
                        settings.get('rotation_speed', 0.01),
                        settings.get('enable_shadows', True),
                        settings.get('background_color', '#1a1a2e'),
                        settings.get('grid_size', 20)
                    ))
            
            conn.commit()
            return jsonify({"success": True})
        
        elif request.method == "DELETE":
            # Apenas admins podem deletar ambientes
            if session.get("role") != "manager":
                return jsonify({"error": "Apenas administradores podem excluir ambientes"}), 403
            
            # Soft delete - apenas marca como inativo
            cursor.execute(""" 
                UPDATE environments 
                SET is_active = false, updated_at = CURRENT_TIMESTAMP 
                WHERE id = %s
            """, (environment_id,))
            
            conn.commit()
            return jsonify({"success": True})
    
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Erro ao processar ambiente {environment_id}: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/environments/<int:environment_id>/resources", methods=["GET", "POST"])
@login_required
def environment_resources_api(environment_id):
    """API para gerenciar recursos (imagens, modelos 3D) de um ambiente"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        if request.method == "GET":
            resource_type = request.args.get('type')
            
            query = """
                SELECT * FROM environment_resources 
                WHERE environment_id = %s
            """
            params = [environment_id]
            
            if resource_type:
                query += " AND resource_type = %s"
                params.append(resource_type)
            
            query += " ORDER BY display_order, created_at DESC"
            
            cursor.execute(query, params)
            resources = cursor.fetchall()
            
            return jsonify([dict(r) for r in resources])
        
        elif request.method == "POST":
            # Apenas admins podem adicionar recursos
            if session.get("role") != "manager":
                return jsonify({"error": "Apenas administradores podem adicionar recursos"}), 403
            
            data = request.get_json()
            
            required_fields = ['resource_type', 'file_name', 'file_url']
            for field in required_fields:
                if field not in data:
                    return jsonify({"error": f"Campo obrigatório ausente: {field}"}), 400
            
            # Se marcar como primário, desmarcar outros do mesmo tipo
            if data.get('is_primary'):
                cursor.execute(""" 
                    UPDATE environment_resources 
                    SET is_primary = false 
                    WHERE environment_id = %s AND resource_type = %s
                """, (environment_id, data['resource_type']))
            
            cursor.execute(""" 
                INSERT INTO environment_resources
                (environment_id, resource_type, file_name, file_url, file_size, 
                 mime_type, description, is_primary, display_order, uploaded_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                environment_id,
                data['resource_type'],
                data['file_name'],
                data['file_url'],
                data.get('file_size'),
                data.get('mime_type'),
                data.get('description'),
                data.get('is_primary', False),
                data.get('display_order', 0),
                session.get('user_id')
            ))
            
            resource_id = cursor.fetchone()['id']
            conn.commit()
            
            return jsonify({"success": True, "id": resource_id}), 201
    
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Erro ao processar recursos do ambiente {environment_id}: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


def get_supabase_config():
    """Recupera configurações do Supabase das variáveis de ambiente"""
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_KEY")
    return url, key


def upload_to_supabase(file_obj, filename, content_type, folder="environments"):
    """Realiza upload de arquivo para o Supabase Storage"""
    url, key = get_supabase_config()
    
    if not url or not key:
        # Fallback para desenvolvimento local ou erro
        app.logger.error("Supabase credentials not found")
        raise Exception("Serviço de armazenamento não configurado")
    
    # Limpar URL base
    url = url.rstrip('/')
    bucket = "environment-assets"
    
    folder = folder.strip("/") if folder else "environments"
    storage_path = f"{folder}/{filename}"
    
    # Endpoint da API de Storage
    # POST /storage/v1/object/{bucket}/{path}
    api_url = f"{url}/storage/v1/object/{bucket}/{storage_path}"
    
    headers = {
        "Authorization": f"Bearer {key}",
        "Content-Type": content_type,
        "x-upsert": "true"
    }
    
    # Ler conteúdo do arquivo
    file_content = file_obj.read()
    
    try:
        response = requests.post(api_url, data=file_content, headers=headers)
        
        if response.status_code not in [200, 201]:
            # Se falhar, tentar criar o bucket e tentar novamente?
            # Por simplicidade, assumimos que o bucket existe.
            # Se o erro for 404 (bucket not found), logar erro específico.
            error_msg = f"Supabase Upload Failed ({response.status_code}): {response.text}"
            app.logger.error(error_msg)
            raise Exception("Falha no upload para o storage remoto")
            
        # Retornar URL pública
        # {supabase_url}/storage/v1/object/public/{bucket}/{path}
        public_url = f"{url}/storage/v1/object/public/{bucket}/{storage_path}"
        return public_url, len(file_content)
        
    except requests.exceptions.RequestException as e:
        app.logger.error(f"Erro de conexão com Supabase: {e}")
        raise Exception("Erro de conexão com serviço de storage")


@app.route("/api/agent/knowledge/import-file", methods=["POST"])
@login_required
def import_knowledge_file():
    """
    Faz upload de PDF/DOCX para o Supabase Storage e solicita ingestão no RAG local.

    O processamento (Docling -> JSON + texto -> chunks/embeddings) acontece no agente local (rag_service worker).
    """
    if session.get("role") != "admin":
        return jsonify({"error": "Apenas administradores podem importar documentos"}), 403

    if not RAG_API_URL:
        return jsonify({"error": "RAG_API_URL não configurada"}), 503

    if "file" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400

    file = request.files["file"]
    if not file or not file.filename:
        return jsonify({"error": "Nome de arquivo inválido"}), 400

    filename = secure_filename(file.filename)
    ext = os.path.splitext(filename)[1].lower()
    if ext not in [".pdf", ".docx"]:
        return jsonify({"error": "Formato não suportado. Envie PDF ou DOCX."}), 400

    mime_type = file.mimetype or mimetypes.guess_type(filename)[0] or "application/octet-stream"
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    unique_filename = f"kb_{timestamp}_{filename}"

    try:
        public_url, file_size = upload_to_supabase(file, unique_filename, mime_type, folder="knowledge_uploads")

        ingest_payload = {
            "source_name": unique_filename,
            "file_url": public_url,
            "content_type": "file_url",
            "metadata": {
                "filename": filename,
                "stored_filename": unique_filename,
                "file_url": public_url,
                "file_size": file_size,
                "mime_type": mime_type,
                "uploaded_by": session.get("user_id"),
                "uploaded_role": session.get("role"),
                "uploaded_at": datetime.now().isoformat(),
            },
        }

        resp = requests.post(
            f"{RAG_API_URL.rstrip('/')}/v1/ingest-file-url",
            json=ingest_payload,
            headers={"x-api-key": RAG_API_KEY},
            timeout=30,
        )
        resp.raise_for_status()
        return jsonify({"success": True, "storage_url": public_url, "rag": resp.json()})
    except Exception as e:
        app.logger.error(f"[KB-IMPORT] Erro ao importar documento: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/environments/<int:environment_id>/upload", methods=["POST"])
@login_required
def environment_upload_api(environment_id):
    """API para upload de arquivos de ambiente"""
    # Validar permissão (admin/manager)
    if session.get("role") not in ["admin", "manager"]:
        return jsonify({"error": "Permissão negada"}), 403

    if "file" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400
        
    file = request.files["file"]
    
    if file.filename == "":
        return jsonify({"error": "Nome de arquivo inválido"}), 400
        
    if file:
        try:
            filename = secure_filename(file.filename)
            
            # Adicionar timestamp para evitar colisão de nomes
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            unique_filename = f"{environment_id}_{timestamp}_{filename}"
            
            # Detectar tipo de recurso baseado na extensão
            ext = os.path.splitext(filename)[1].lower()
            mime_type = file.mimetype or mimetypes.guess_type(filename)[0]
            
            resource_type = "document"
            if ext in ['.jpg', '.jpeg', '.png', '.webp']:
                resource_type = "photo"
            elif ext in ['.glb', '.gltf', '.fbx', '.obj']:
                resource_type = "model_3d"
            elif ext in ['.pdf']:
                # PDFs podem ser plantas ou docs
                resource_type = "plant_2d" # Assumindo planta por padrão para PDF neste contexto
            
            # Realizar upload
            public_url, file_size = upload_to_supabase(file, unique_filename, mime_type)
            
            # Salvar no banco
            conn = get_db()
            cursor = conn.cursor()
            
            # Verificar se já existe um recurso primário deste tipo
            cursor.execute("""
                SELECT id FROM environment_resources 
                WHERE environment_id = %s AND resource_type = %s AND is_primary = true
            """, (environment_id, resource_type))
            has_primary = cursor.fetchone() is not None
            
            # Inserir novo recurso
            cursor.execute("""
                INSERT INTO environment_resources
                (environment_id, resource_type, file_name, file_url, file_size, 
                 mime_type, is_primary, uploaded_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                environment_id,
                resource_type,
                filename,
                public_url,
                file_size,
                mime_type,
                not has_primary, # Se não tem primário, este será o primeiro
                session.get('user_id')
            ))
            
            resource_id = cursor.fetchone()['id']
            conn.commit()
            conn.close()
            
            return jsonify({
                "success": True, 
                "id": resource_id, 
                "url": public_url,
                "type": resource_type
            }), 201
            
        except Exception as e:
            app.logger.error(f"Erro no upload: {e}")
            return jsonify({"error": str(e)}), 500

    return jsonify({"error": "Erro desconhecido"}), 500


@app.route("/api/resources/<int:resource_id>", methods=["DELETE"])
@login_required
def delete_resource_api(resource_id):
    """API para excluir um recurso"""
    if session.get("role") not in ["admin", "manager"]:
        return jsonify({"error": "Permissão negada"}), 403
        
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verificar se existe
        cursor.execute("SELECT id FROM environment_resources WHERE id = %s", (resource_id,))
        if not cursor.fetchone():
            return jsonify({"error": "Recurso não encontrado"}), 404
            
        # Deletar do banco
        cursor.execute("DELETE FROM environment_resources WHERE id = %s", (resource_id,))
        conn.commit()
        
        return jsonify({"success": True}), 200
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# --------------------------------------------------------------------------- #
# Rotas do Agente IA
# --------------------------------------------------------------------------- #
@app.route("/agent")
@login_required
@admin_required
def agent_page():
    """Página principal: Biblioteca de Automações."""
    # Nova UI (fase 1): catálogo + execução com feedback via APIs unificadas.
    return render_template(get_template("agent_library.html"), automations=list_automations())


@app.route("/agent/legacy")
@login_required
def agent_page_legacy():
    """Página antiga do Agente IA (abas). Mantida temporariamente para comparação/debug."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Buscar tipos de RPA com contagem de RPAs
        cursor.execute("""
            SELECT t.id, t.name, t.description, t.icon, t.is_active,
                   COUNT(r.id) as rpa_count
            FROM agent_rpa_types t
            LEFT JOIN agent_rpas r ON r.rpa_type_id = t.id
            WHERE t.is_active = true 
            GROUP BY t.id, t.name, t.description, t.icon, t.is_active
            ORDER BY t.name
        """)
        rpa_types = [dict(row) for row in cursor.fetchall()]
        
        # Buscar fontes de dados
        cursor.execute("""
            SELECT id, name, description, source_type 
            FROM agent_data_sources 
            WHERE is_active = true 
            ORDER BY name
        """)
        data_sources = [dict(row) for row in cursor.fetchall()]
        
        # Buscar RPAs do usuário
        cursor.execute("""
            SELECT r.id, r.name, r.status, r.priority, r.created_at,
                   t.name as type_name
            FROM agent_rpas r
            LEFT JOIN agent_rpa_types t ON r.rpa_type_id = t.id
            WHERE r.created_by = %s
            ORDER BY r.created_at DESC
            LIMIT 20
        """, (session['user_id'],))
        rpas = [dict(row) for row in cursor.fetchall()]
        
        # Buscar dashboards gerados pelo usuário
        cursor.execute("""
            SELECT id, title, category, status, result_url, result_data, created_at
            FROM agent_dashboard_requests
            WHERE created_by = %s
            ORDER BY created_at DESC
            LIMIT 20
        """, (session['user_id'],))
        generated_dashboards = [dict(row) for row in cursor.fetchall()]
        
        # Estatísticas de RPA
        cursor.execute("""
            SELECT 
                COUNT(*) FILTER (WHERE status = 'pending') as pending,
                COUNT(*) FILTER (WHERE status = 'running') as running,
                COUNT(*) FILTER (WHERE status = 'completed') as completed,
                COUNT(*) FILTER (WHERE status = 'failed') as failed
            FROM agent_rpas
            WHERE created_by = %s
        """, (session['user_id'],))
        stats = dict(cursor.fetchone())
        
        # Estatísticas de Dashboard
        cursor.execute("""
            SELECT 
                COUNT(*) FILTER (WHERE status = 'pending') as pending,
                COUNT(*) FILTER (WHERE status = 'processing') as processing,
                COUNT(*) FILTER (WHERE status = 'completed') as completed
            FROM agent_dashboard_requests
            WHERE created_by = %s
        """, (session['user_id'],))
        dashboard_stats = dict(cursor.fetchone())
        
        # Buscar templates de dashboard do usuário
        dashboard_templates = []
        try:
            cursor.execute("""
                SELECT id, title, description, category, is_published, thumbnail_url, created_at
                FROM agent_dashboard_templates
                WHERE created_by = %s
                ORDER BY updated_at DESC
                LIMIT 20
            """, (session['user_id'],))
            dashboard_templates = [dict(row) for row in cursor.fetchall()]
        except Exception:
            pass  # Tabela pode não existir ainda
        
        relatorio_meta = fetch_relatorio_meta_settings()
        return render_template(
            get_template("agent.html"),
            rpa_types=rpa_types,
            data_sources=data_sources,
            rpas=rpas,
            generated_dashboards=generated_dashboards,
            stats=stats,
            dashboard_stats=dashboard_stats,
            dashboard_templates=dashboard_templates,
            relatorio_meta=relatorio_meta,
        )
        
    except Exception as e:
        app.logger.error(f"[AGENT] Erro ao carregar página: {e}")
        # Se as tabelas não existem, mostrar página com dados vazios
        relatorio_meta = fetch_relatorio_meta_settings()
        return render_template(
            get_template("agent.html"),
            rpa_types=[],
            data_sources=[],
            rpas=[],
            generated_dashboards=[],
            stats={'pending': 0, 'running': 0, 'completed': 0, 'failed': 0},
            dashboard_stats={'pending': 0, 'processing': 0, 'completed': 0},
            dashboard_templates=[],
            relatorio_meta=relatorio_meta,
        )
    finally:
        conn.close()


@app.route("/agent/automation/<string:automation_id>")
@login_required
@admin_required
def agent_automation_page(automation_id: str):
    """Página dedicada de um executável do catálogo."""
    a = get_automation((automation_id or "").strip())
    if not a:
        flash("Automação não encontrada.", "error")
        return redirect(url_for("agent_page"))

    # Para algumas automações, a UX correta é a do legado (tabs + formulários)
    if a.id == "auditoria_fiscal_manifestos":
        return redirect(url_for("agent_page_legacy", tab="auditoria"))
    if a.id == "relatorio_resultados_controladoria":
        return redirect(url_for("agent_page_legacy", tab="relatorio"))

    return render_template(get_template("agent_automation.html"), automation=a)


def _ensure_agent_library_models_table():
    """Cria tabela para modelos do laboratório caso não exista (sem migrations no projeto)."""
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS agent_library_models (
                id SERIAL PRIMARY KEY,
                slug TEXT UNIQUE,
                title TEXT NOT NULL,
                description TEXT,
                category TEXT NOT NULL,
                kind TEXT NOT NULL DEFAULT 'query', -- query|procedure
                database TEXT NOT NULL DEFAULT 'azportoex',
                query TEXT,
                procedure TEXT,
                procedure_params JSONB,
                is_published BOOLEAN NOT NULL DEFAULT FALSE,
                created_by INTEGER,
                created_at TIMESTAMP NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMP NOT NULL DEFAULT NOW()
            )
            """
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _list_library_models(published_only: bool):
    _ensure_agent_library_models_table()
    conn = get_db()
    cursor = conn.cursor()
    try:
        where = "WHERE is_published = true" if published_only else ""
        cursor.execute(
            f"""
            SELECT id, slug, title, description, category, kind, database, is_published, created_at, updated_at
            FROM agent_library_models
            {where}
            ORDER BY updated_at DESC
            """
        )
        return [dict(r) for r in (cursor.fetchall() or [])]
    finally:
        conn.close()


@app.route("/agent/lab")
@login_required
@admin_required
def agent_lab_page():
    """Laboratório: criação/publicação de modelos executáveis."""
    models = _list_library_models(published_only=False)
    return render_template(get_template("agent_lab.html"), models=models)


@app.route("/agent/lab/models/create", methods=["POST"])
@login_required
@admin_required
def agent_lab_create_model():
    """Cria um modelo executável (query/procedure) para aparecer na biblioteca quando publicado."""
    _ensure_agent_library_models_table()
    data = request.form or {}
    title = (data.get("title") or "").strip()
    description = (data.get("description") or "").strip() or None
    category = (data.get("category") or "").strip()
    kind = (data.get("kind") or "query").strip()
    database = (data.get("database") or "azportoex").strip()
    query = (data.get("query") or "").strip() or None
    procedure = (data.get("procedure") or "").strip() or None
    procedure_params_json = (data.get("procedure_params_json") or "").strip()

    if not title or not category:
        flash("Título e categoria são obrigatórios.", "error")
        return redirect(url_for("agent_lab_page"))

    if kind not in ("query", "procedure"):
        flash("Tipo inválido.", "error")
        return redirect(url_for("agent_lab_page"))

    if not _validate_database_choice(database):
        flash("Base inválida.", "error")
        return redirect(url_for("agent_lab_page"))

    procedure_params = None
    if procedure_params_json:
        try:
            procedure_params = json.loads(procedure_params_json)
        except Exception:
            flash("procedure_params_json inválido (JSON).", "error")
            return redirect(url_for("agent_lab_page"))

    if kind == "query":
        if not query:
            flash("Query é obrigatória para modelo do tipo query.", "error")
            return redirect(url_for("agent_lab_page"))
        if not query.strip().upper().startswith("SELECT"):
            flash("Por segurança, apenas queries SELECT são permitidas.", "error")
            return redirect(url_for("agent_lab_page"))
    else:
        if not procedure:
            flash("Procedure é obrigatória para modelo do tipo procedure.", "error")
            return redirect(url_for("agent_lab_page"))

    slug = re.sub(r"[^a-z0-9\\-]+", "-", title.lower()).strip("-")[:60] or f"model-{int(time.time())}"

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO agent_library_models
            (slug, title, description, category, kind, database, query, procedure, procedure_params, created_by, is_published)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, false)
            """,
            (
                slug,
                title,
                description,
                category,
                kind,
                database,
                query,
                procedure,
                psycopg2.extras.Json(procedure_params) if procedure_params else None,
                session["user_id"],
            ),
        )
        conn.commit()
        flash("Modelo criado. Publique para aparecer na Biblioteca.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao criar modelo: {e}", "error")
    finally:
        conn.close()
    return redirect(url_for("agent_lab_page"))


@app.route("/agent/lab/models/<int:model_id>/publish", methods=["POST"])
@login_required
@admin_required
def agent_lab_publish_model(model_id: int):
    _ensure_agent_library_models_table()
    publish = (request.form.get("publish") or "").strip().lower() == "true"
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            UPDATE agent_library_models
            SET is_published = %s, updated_at = NOW()
            WHERE id = %s
            """,
            (publish, model_id),
        )
        conn.commit()
        flash("Atualizado.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao atualizar: {e}", "error")
    finally:
        conn.close()
    return redirect(url_for("agent_lab_page"))


@app.route("/agent/model/<int:model_id>")
@login_required
@admin_required
def agent_model_page(model_id: int):
    _ensure_agent_library_models_table()
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            SELECT id, slug, title, description, category, kind, database, query, procedure, procedure_params, is_published
            FROM agent_library_models
            WHERE id = %s
            """,
            (model_id,),
        )
        model = cursor.fetchone()
        if not model:
            flash("Modelo não encontrado.", "error")
            return redirect(url_for("agent_lab_page"))
        model = dict(model)
    finally:
        conn.close()
    return render_template(get_template("agent_model.html"), model=model)


def _dashboard_request_by_id_for_user(request_id: int, user_id: int):
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            SELECT id, title, description, category, status, filters, result_data, error_message, updated_at
            FROM agent_dashboard_requests
            WHERE id = %s AND created_by = %s
            """,
            (request_id, user_id),
        )
        row = cursor.fetchone()
        if not row:
            return None
        # Evitar vazar credenciais sensíveis em qualquer resposta (mesmo para o próprio usuário)
        d = dict(row)
        try:
            f = d.get("filters") or {}
            if isinstance(f, dict):
                if "brudam_senha" in f and f.get("brudam_senha"):
                    f["brudam_senha"] = "***"
                d["filters"] = f
        except Exception:
            pass
        return d
    finally:
        conn.close()


def _insert_agent_dashboard_request(
    *,
    title: str,
    description: str,
    category: str,
    chart_types: list[str],
    filters: dict,
    created_by: int,
) -> int:
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO agent_dashboard_requests
            (title, description, category, chart_types, filters, created_by, status, created_at)
            VALUES (%s, %s, %s, %s::text[], %s, %s, 'pending', NOW())
            RETURNING id
            """,
            (
                title,
                description,
                category,
                chart_types,
                psycopg2.extras.Json(filters) if filters else None,
                created_by,
            ),
        )
        request_id = cursor.fetchone()["id"]
        conn.commit()
        return request_id
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _normalize_state(status: str | None) -> str:
    s = (status or "").lower().strip()
    if s in ("pending",):
        return "pending"
    if s in ("running",):
        return "running"
    if s in ("processing",):
        return "processing"
    if s in ("completed",):
        return "completed"
    if s in ("failed", "error"):
        return "failed"
    return "pending" if not s else s


def _validate_yyyy_mm_dd(value: str | None) -> bool:
    if not value:
        return False
    # YYYY-MM-DD
    return bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}", value.strip()))


def _validate_database_choice(value: str | None) -> bool:
    return (value or "").strip() in ("azportoex", "portoexsp", "ambas")


def _build_steps(automation_id: str, state: str) -> list[dict]:
    a = get_automation(automation_id)
    labels = a.steps if a else ["Enfileirar", "Executar", "Concluir"]
    steps = []
    if state in ("pending",):
        active_idx = 0
    elif state in ("running", "processing"):
        active_idx = 2 if len(labels) > 2 else 1
    elif state in ("completed",):
        active_idx = len(labels)  # all done
    else:
        active_idx = 1
    for idx, label in enumerate(labels):
        if state == "failed" and idx == active_idx:
            st = "error"
        elif idx < active_idx:
            st = "done"
        elif idx == active_idx and state in ("running", "processing", "pending"):
            st = "active"
        else:
            st = "pending"
        steps.append({"label": label, "status": st})
    return steps


@app.route("/api/agent/library/catalog", methods=["GET"])
@login_required
@admin_required
def agent_library_catalog():
    """Catálogo unificado de automações (fonte: automate/automation_catalog.py)."""
    items = []
    for a in list_automations():
        items.append(
            {
                "id": a.id,
                "name": a.name,
                "description": a.description,
                "category": a.category,
                "kind": a.kind,
                "steps": a.steps,
                "params": [
                    {
                        "key": p.key,
                        "label": p.label,
                        "type": p.type,
                        "required": p.required,
                        "default": p.default,
                        "placeholder": p.placeholder,
                        "options": p.options,
                        "help_text": p.help_text,
                    }
                    for p in (a.params or [])
                ],
            }
        )
    return jsonify({"success": True, "items": items}), 200


@app.route("/api/agent/library/run", methods=["POST"])
@login_required
@admin_required
def agent_library_run():
    """Dispara execução de uma automação da Biblioteca e retorna um run_id unificado."""
    payload = request.get_json() or {}
    automation_id = (payload.get("automation_id") or "").strip()
    params = payload.get("params") or {}

    a = get_automation(automation_id)
    if not a:
        return jsonify({"success": False, "error": "Automação não encontrada"}), 404

    # Validar params obrigatórios
    for p in a.params:
        if p.required and not (params.get(p.key) or "").strip():
            return jsonify({"success": False, "error": f"Parâmetro obrigatório: {p.key}"}), 400

    user_id = session["user_id"]

    try:
        if a.id == "auditoria_fiscal_manifestos":
            data_inicio = params.get("data_inicio")
            data_fim = params.get("data_fim")
            operador_id = params.get("operador_id")

            if not (_validate_yyyy_mm_dd(data_inicio) and _validate_yyyy_mm_dd(data_fim)):
                return jsonify({"success": False, "error": "Datas inválidas (use YYYY-MM-DD)"}), 400

            # Reaproveita a mesma query do endpoint legado, mas adiciona _automation_id
            query = f"""
                SELECT 
                    m.id_manifesto,
                    m.data_emissao,
                    m.tipo,
                    mt.tipo as tipo_descricao,
                    m.id_agente,
                    f.fantasia as agente_nome,
                    m.motorista,
                    func.nome as motorista_nome,
                    m.veiculo,
                    v.placa as veiculo_placa,
                    v.modelo as veiculo_modelo,
                    m.operador,
                    u.primeiro_nome as operador_nome,
                    m.picking,
                    m.km_inicial,
                    m.km_final,
                    m.km_rodado,
                    m.fatura,
                    m.total_nf_valor,
                    m.custo_motorista,
                    m.custo_motorista_extra,
                    m.adiantamento,
                    m.pedagio
                FROM azportoex.manifesto m
                LEFT JOIN azportoex.manifesto_tipo mt ON m.tipo = mt.id_tipo
                LEFT JOIN azportoex.fornecedores f ON m.id_agente = f.id_local
                LEFT JOIN azportoex.funcionario func ON m.motorista = func.id_funcionario
                LEFT JOIN azportoex.usuarios u ON m.operador = u.id_usuario
                LEFT JOIN azportoex.veiculos v ON m.veiculo = v.id_veiculo
                WHERE m.data_emissao BETWEEN '{data_inicio}' AND '{data_fim}'
            """
            if operador_id:
                query += f" AND m.operador = {int(operador_id)}"
            query += " ORDER BY m.data_emissao DESC, m.id_manifesto DESC"

            request_id = _insert_agent_dashboard_request(
                title=f"Auditoria {data_inicio} a {data_fim}",
                description="Auditoria Fiscal de Manifestos",
                category="auditoria",
                chart_types=["table"],
                filters={"query": query, "limit": 2000, "_automation_id": a.id},
                created_by=user_id,
            )
            return jsonify({"success": True, "run_id": f"dashreq:{request_id}"}), 200

        if a.id == "relatorio_resultados_controladoria":
            period = (params.get("period") or "mes_atual").strip()
            data_inicio = params.get("data_inicio")
            data_fim = params.get("data_fim")
            database = (params.get("database") or "azportoex").strip()
            fp_tokens_raw = params.get("forma_pagamento_tokens")
            if isinstance(fp_tokens_raw, list):
                forma_pagamento_tokens = [str(x).strip() for x in fp_tokens_raw if str(x).strip()]
            else:
                fp_single = (params.get("forma_pagamento") or "").strip()
                forma_pagamento_tokens = [fp_single] if fp_single else []
            if not forma_pagamento_tokens:
                forma_pagamento_tokens = ["Faturado", "A vista"]

            # Se custom, valida datas
            if (not period or period.lower() in ("custom", "personalizado")):
                if not (_validate_yyyy_mm_dd(data_inicio) and _validate_yyyy_mm_dd(data_fim)):
                    return jsonify({"success": False, "error": "Datas inválidas (use YYYY-MM-DD)"}), 400

            if not _validate_database_choice(database):
                return jsonify({"success": False, "error": "Base inválida"}), 400

            if database == "ambas":
                ids = []
                for db_name in ["azportoex", "portoexsp"]:
                    filters_payload = _build_relatorio_filters(
                        data_inicio,
                        data_fim,
                        db_name,
                        period=period,
                        forma_pagamento_tokens=forma_pagamento_tokens,
                    )
                    req_id = _insert_agent_dashboard_request(
                        title=f"Relatório Resultados {db_name} ({period})",
                        description=f"Relatório gerencial da controladoria - {db_name}",
                        category="relatorio_resultados",
                        chart_types=["table", "cards"],
                        filters={**filters_payload, "_automation_id": a.id},
                        created_by=user_id,
                    )
                    ids.append(req_id)
                return jsonify({"success": True, "run_id": f"dashreqmulti:{','.join([str(i) for i in ids])}"}), 200

            filters_payload = _build_relatorio_filters(
                data_inicio,
                data_fim,
                database,
                period=period,
                forma_pagamento_tokens=forma_pagamento_tokens,
            )
            if not filters_payload:
                return jsonify({"success": False, "error": "Base inválida"}), 400
            request_id = _insert_agent_dashboard_request(
                title=f"Relatório Resultados {database} ({period})",
                description=f"Relatório gerencial da controladoria - {database}",
                category="relatorio_resultados",
                chart_types=["table", "cards"],
                filters={**filters_payload, "_automation_id": a.id},
                created_by=user_id,
            )
            return jsonify({"success": True, "run_id": f"dashreq:{request_id}"}), 200

        if a.id == "gerar_dashboard":
            title = params.get("title")
            description = params.get("description")
            filters_json = (params.get("filters_json") or "").strip()
            filters = None
            if filters_json:
                try:
                    filters = json.loads(filters_json)
                except Exception:
                    return jsonify({"success": False, "error": "Filtros (JSON) inválidos"}), 400

            request_id = _insert_agent_dashboard_request(
                title=title,
                description=description,
                category="dashboard_gen",
                chart_types=["table", "cards"],
                filters={**(filters or {}), "_automation_id": a.id},
                created_by=user_id,
            )
            return jsonify({"success": True, "run_id": f"dashreq:{request_id}"}), 200

        if a.id == "logistica_km_por_motorista":
            data_inicio = params.get("data_inicio")
            data_fim = params.get("data_fim")
            database = (params.get("database") or "azportoex").strip()

            if not (_validate_yyyy_mm_dd(data_inicio) and _validate_yyyy_mm_dd(data_fim)):
                return jsonify({"success": False, "error": "Datas inválidas (use YYYY-MM-DD)"}), 400
            if not _validate_database_choice(database):
                return jsonify({"success": False, "error": "Base inválida"}), 400

            def _mk_query(db: str) -> str:
                # Consolidado por motorista; usa colunas já presentes na auditoria (manifesto).
                return f"""
                    SELECT
                        COALESCE(func.nome, CAST(m.motorista AS CHAR), 'N/A') AS motorista_nome,
                        COUNT(*) AS total_manifestos,
                        SUM(COALESCE(m.km_rodado, 0)) AS km_total_rodado,
                        SUM(COALESCE(m.total_nf_valor, 0)) AS valor_total_nf,
                        SUM(COALESCE(m.custo_motorista, 0) + COALESCE(m.custo_motorista_extra, 0) + COALESCE(m.pedagio, 0)) AS custo_total_estimado
                    FROM {db}.manifesto m
                    LEFT JOIN {db}.funcionario func ON m.motorista = func.id_funcionario
                    WHERE m.data_emissao BETWEEN '{data_inicio}' AND '{data_fim}'
                    GROUP BY COALESCE(func.nome, CAST(m.motorista AS CHAR), 'N/A')
                    ORDER BY km_total_rodado DESC
                """

            if database == "ambas":
                ids = []
                for db_name in ["azportoex", "portoexsp"]:
                    request_id = _insert_agent_dashboard_request(
                        title=f"Logística (KM por motorista) • {db_name} • {data_inicio} a {data_fim}",
                        description="Consolidação de KM, valores e custos por motorista",
                        category="logistica",
                        chart_types=["table", "cards"],
                        filters={"query": _mk_query(db_name), "limit": 5000, "database": db_name, "_automation_id": a.id},
                        created_by=user_id,
                    )
                    ids.append(request_id)
                return jsonify({"success": True, "run_id": f"dashreqmulti:{','.join([str(i) for i in ids])}"}), 200

            request_id = _insert_agent_dashboard_request(
                title=f"Logística (KM por motorista) • {database} • {data_inicio} a {data_fim}",
                description="Consolidação de KM, valores e custos por motorista",
                category="logistica",
                chart_types=["table", "cards"],
                filters={"query": _mk_query(database), "limit": 5000, "database": database, "_automation_id": a.id},
                created_by=user_id,
            )
            return jsonify({"success": True, "run_id": f"dashreq:{request_id}"}), 200

        if a.id == "brudam_relatorio_completo_selenium":
            data_inicio = params.get("data_inicio")
            data_fim = params.get("data_fim")
            modo = (params.get("modo") or "completo").strip()
            headless_raw = (params.get("headless") or "1").strip()
            brudam_usuario = (params.get("brudam_usuario") or "").strip()
            brudam_senha = (params.get("brudam_senha") or "").strip()
            brudam_url = (params.get("brudam_url") or "").strip()
            use_agent_default = (params.get("use_agent_default") or "").strip().lower() in ("1", "true", "yes", "on")

            if not (_validate_yyyy_mm_dd(data_inicio) and _validate_yyyy_mm_dd(data_fim)):
                return jsonify({"success": False, "error": "Datas inválidas (use YYYY-MM-DD)"}), 400
            if modo not in ("completo", "geral"):
                return jsonify({"success": False, "error": "Modo inválido"}), 400
            if headless_raw not in ("0", "1", "true", "false", "True", "False"):
                return jsonify({"success": False, "error": "Parâmetro headless inválido"}), 400

            headless = headless_raw in ("1", "true", "True")

            # Se não estiver usando o default do agente, exige usuário/senha (para evitar execuções que vão falhar)
            if not use_agent_default and (not brudam_usuario or not brudam_senha):
                return jsonify({"success": False, "error": "Informe usuário e senha do Brudam (ou selecione 'Usar credenciais do agente local')"}), 400

            request_id = _insert_agent_dashboard_request(
                title=f"Relatório Completo Brudam (Selenium) • {data_inicio} a {data_fim}",
                description="RPA local (Selenium) com geração de HTML completo baseado nos dados coletados.",
                category="rpa_selenium",
                chart_types=["html", "table"],
                filters={
                    "runner": "brudam_selenium_report",
                    "data_inicio": data_inicio,
                    "data_fim": data_fim,
                    "modo": modo,
                    "headless": headless,
                    # credenciais (somente quando fornecidas; o agente local pode usar seu .env)
                    "brudam_usuario": brudam_usuario if (not use_agent_default and brudam_usuario) else None,
                    "brudam_senha": brudam_senha if (not use_agent_default and brudam_senha) else None,
                    "brudam_url": brudam_url or None,
                    "_automation_id": a.id,
                },
                created_by=user_id,
            )
            return jsonify({"success": True, "run_id": f"dashreq:{request_id}"}), 200

        return jsonify({"success": False, "error": "Automação ainda não suportada"}), 400

    except Exception as e:
        app.logger.error("[LIBRARY] Erro ao iniciar run: %s", e)
        return jsonify({"success": False, "error": str(e)}), 500


def _parse_run_id(run_id: str) -> tuple[str, list[int]]:
    run_id = (run_id or "").strip()
    if run_id.startswith("dashreq:"):
        return ("dashreq", [int(run_id.split(":", 1)[1])])
    if run_id.startswith("dashreqmulti:"):
        raw = run_id.split(":", 1)[1]
        parts = [p for p in raw.split(",") if p.strip()]
        return ("dashreqmulti", [int(p) for p in parts])
    if run_id.startswith("rpa:"):
        return ("rpa", [int(run_id.split(":", 1)[1])])
    return ("unknown", [])


@app.route("/api/agent/library/run/<path:run_id>", methods=["GET"])
@login_required
@admin_required
def agent_library_run_status(run_id):
    kind, ids = _parse_run_id(run_id)
    if kind == "unknown" or not ids:
        return jsonify({"success": False, "error": "run_id inválido"}), 400

    user_id = session["user_id"]

    # (Fase 1) Suporte primário: agent_dashboard_requests
    if kind in ("dashreq", "dashreqmulti"):
        rows = []
        for rid in ids:
            row = _dashboard_request_by_id_for_user(rid, user_id)
            if not row:
                return jsonify({"success": False, "error": "Execução não encontrada"}), 404
            rows.append(row)

        # Consolidar estado: failed > processing > pending > completed (multi)
        states = [_normalize_state(r.get("status")) for r in rows]
        if any(s == "failed" for s in states):
            state = "failed"
        elif any(s == "processing" for s in states):
            state = "processing"
        elif any(s == "pending" for s in states):
            state = "pending"
        elif all(s == "completed" for s in states):
            state = "completed"
        else:
            state = states[0] if states else "pending"

        # automation_id: prefer filters._automation_id
        automation_id = None
        try:
            automation_id = (rows[0].get("filters") or {}).get("_automation_id")
        except Exception:
            automation_id = None

        steps = _build_steps(automation_id or "unknown", state)
        total_steps = max(len(steps), 1)
        active_idx = next((i for i, s in enumerate(steps) if s["status"] == "active"), total_steps)
        done_count = len([s for s in steps if s["status"] == "done"])
        percent = 0 if state == "pending" else int(((done_count + (1 if state in ("running", "processing") else 0)) / total_steps) * 100)
        percent = 100 if state == "completed" else percent

        # logs: combina logs parciais do agent_local (result_data._logs) + status básico
        logs: list[str] = []
        merged_progress = None
        for r in rows:
            logs.append(f"[{r.get('updated_at')}] status={r.get('status')} id={r.get('id')}")
            rd = r.get("result_data") or {}
            if isinstance(rd, dict):
                rd_logs = rd.get("_logs") or []
                if isinstance(rd_logs, list):
                    logs.extend([str(x) for x in rd_logs if x is not None])
                if not merged_progress and isinstance(rd.get("_progress"), dict):
                    merged_progress = rd.get("_progress")
            if r.get("error_message"):
                logs.append(f"[ERRO] {r.get('error_message')}")
        # limitar para não explodir o payload
        logs = logs[-1500:]

        # resultado: extrai dados
        row_count = 0
        combined_data: list[dict] = []
        combined_payloads: list[dict] = []
        error_msg = None
        for r in rows:
            if r.get("error_message"):
                error_msg = r.get("error_message")
            rd = r.get("result_data") or {}
            data = rd.get("data") or []
            if isinstance(data, list):
                combined_data.extend(data)
                row_count += len(data)
            elif isinstance(data, dict):
                # Para RPAs/relatórios, pode ser um payload estruturado (ex.: {"html": "...", ...})
                combined_payloads.append(data)

        artifacts = {}
        if state == "completed":
            artifacts = {
                "html_url": f"/api/agent/library/run/{run_id}/report.html",
                "pdf_url": f"/api/agent/library/run/{run_id}/report.pdf",
            }

        progress_text = f"{state.upper()} • {percent}%"
        if merged_progress:
            # Se o agent_local enviar progresso, preferimos isso
            try:
                if isinstance(merged_progress.get("percent"), (int, float)):
                    percent = int(merged_progress.get("percent"))
                if merged_progress.get("text"):
                    progress_text = str(merged_progress.get("text"))
            except Exception:
                pass

        return jsonify(
            {
                "success": True,
                "run_id": run_id,
                "state": state,
                "progress": {
                    "percent": percent,
                    "text": progress_text,
                    "current_step": active_idx,
                    "total_steps": total_steps,
                },
                "steps": steps,
                "logs": logs,
                "result": {
                    "row_count": row_count,
                    "error": error_msg,
                    "payloads": combined_payloads[:3],  # pequeno preview quando não é tabular
                },
                "artifacts": artifacts,
            }
        ), 200

    return jsonify({"success": False, "error": "run_id não suportado"}), 400


@app.route("/api/agent/library/run/<path:run_id>/report.html", methods=["GET"])
@login_required
@admin_required
def agent_library_run_report_html(run_id):
    kind, ids = _parse_run_id(run_id)
    if kind not in ("dashreq", "dashreqmulti") or not ids:
        return jsonify({"error": "run_id inválido"}), 400

    user_id = session["user_id"]
    rows = []
    for rid in ids:
        row = _dashboard_request_by_id_for_user(rid, user_id)
        if not row:
            return jsonify({"error": "Execução não encontrada"}), 404
        rows.append(row)

    # Consolidar dataset
    title = rows[0].get("title") or "Relatório"
    subtitle = rows[0].get("description") or None
    data = []
    for r in rows:
        rd = r.get("result_data") or {}
        d = rd.get("data") or []
        # Caso especial: payload estruturado vindo do RPA Selenium (contém HTML completo pronto)
        if isinstance(d, dict) and d.get("html"):
            html = str(d.get("html"))
            return html, 200, {"Content-Type": "text/html; charset=utf-8"}
        if isinstance(d, list):
            data.extend(d)

    report = build_tabular_report(title=title, records=data, subtitle=subtitle)
    html = render_report_html(report)
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}


@app.route("/api/agent/library/run/<path:run_id>/report.pdf", methods=["GET"])
@login_required
@admin_required
def agent_library_run_report_pdf(run_id):
    kind, ids = _parse_run_id(run_id)
    if kind not in ("dashreq", "dashreqmulti") or not ids:
        return jsonify({"error": "run_id inválido"}), 400

    user_id = session["user_id"]
    rows = []
    for rid in ids:
        row = _dashboard_request_by_id_for_user(rid, user_id)
        if not row:
            return jsonify({"error": "Execução não encontrada"}), 404
        rows.append(row)

    title = rows[0].get("title") or "Relatório"
    subtitle = rows[0].get("description") or None
    data = []
    for r in rows:
        rd = r.get("result_data") or {}
        d = rd.get("data") or []
        if isinstance(d, list):
            data.extend(d)

    report = build_tabular_report(title=title, records=data, subtitle=subtitle)
    pdf_bytes = generate_report_pdf(report)
    filename = f"relatorio_{run_id.replace(':', '_')}.pdf"
    return (
        pdf_bytes,
        200,
        {
            "Content-Type": "application/pdf",
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@app.route("/api/agent/rpa", methods=["POST"])
@login_required
def create_rpa():
    """API para criar uma nova automação RPA."""
    data = request.get_json()
    
    if not data.get('name') or not data.get('description'):
        return jsonify({"error": "Nome e descrição são obrigatórios"}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Validar parâmetros JSON se fornecido
        parameters = None
        if data.get('parameters'):
            try:
                import json
                parameters = json.loads(data['parameters']) if isinstance(data['parameters'], str) else data['parameters']
            except:
                return jsonify({"error": "Parâmetros JSON inválidos"}), 400
        
        cursor.execute("""
            INSERT INTO agent_rpas (name, description, rpa_type_id, priority, frequency, parameters, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            data['name'],
            data['description'],
            data.get('rpa_type') or None,
            data.get('priority', 'medium'),
            data.get('frequency', 'once'),
            psycopg2.extras.Json(parameters) if parameters else None,
            session['user_id']
        ))
        
        rpa_id = cursor.fetchone()['id']
        conn.commit()
        
        # Log da ação
        cursor.execute("""
            INSERT INTO agent_logs (action_type, entity_type, entity_id, user_id, details)
            VALUES ('create', 'rpa', %s, %s, %s)
        """, (rpa_id, session['user_id'], psycopg2.extras.Json({'name': data['name']})))
        conn.commit()
        
        return jsonify({"success": True, "id": rpa_id}), 201
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"[AGENT] Erro ao criar RPA: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/rpa/<int:rpa_id>", methods=["DELETE"])
@login_required
def delete_rpa(rpa_id):
    """API para excluir uma automação RPA."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verificar se pertence ao usuário ou se é admin
        cursor.execute("SELECT created_by FROM agent_rpas WHERE id = %s", (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            return jsonify({"error": "RPA não encontrada"}), 404
        
        if rpa['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        cursor.execute("DELETE FROM agent_rpas WHERE id = %s", (rpa_id,))
        conn.commit()
        
        return jsonify({"success": True}), 200
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-gen", methods=["GET"])
@login_required
def list_dashboard_gen():
    """Lista solicitações de geração de dashboard com filtros opcionais.
    Suporta filtros via query string: status=completed|pending|failed e has_data=true.
    Retorna itens com id, title, status, row_count e updated_at.
    """
    status = (request.args.get('status') or '').strip().lower()
    has_data = (request.args.get('has_data') or '').strip().lower() == 'true'

    conn = get_db()
    cursor = conn.cursor()

    try:
        where_clauses = []
        params = []

        # Restringir por usuário (a menos que admin)
        if session.get('role') != 'admin':
            where_clauses.append("created_by = %s")
            params.append(session['user_id'])

        # Filtro de status opcional
        if status in ('completed', 'pending', 'failed'):
            where_clauses.append("status = %s")
            params.append(status)

        where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ''

        cursor.execute(f"""
            SELECT id, title, status, result_data, updated_at
            FROM agent_dashboard_requests
            {where_sql}
            ORDER BY updated_at DESC
            LIMIT 100
        """, tuple(params))

        rows = cursor.fetchall() or []
        items = []
        for r in rows:
            rd = r.get('result_data') or {}
            data = rd.get('data') or []
            row_count = len(data) if isinstance(data, list) else 0
            if has_data and row_count == 0:
                continue
            items.append({
                'id': r['id'],
                'title': r['title'],
                'status': r['status'],
                'row_count': row_count,
                'updated_at': r.get('updated_at').isoformat() if r.get('updated_at') else None
            })

        return jsonify({'items': items}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-gen", methods=["POST"])
@login_required
def create_dashboard_gen():
    """API para solicitar geração de dashboard."""
    data = request.get_json()
    
    if not data.get('title') or not data.get('description'):
        return jsonify({"error": "Título e descrição são obrigatórios"}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Validar filtros JSON se fornecido
        filters = None
        if data.get('filters'):
            try:
                import json
                filters = json.loads(data['filters']) if isinstance(data['filters'], str) else data['filters']
            except:
                return jsonify({"error": "Filtros JSON inválidos"}), 400
        
        chart_types = data.get('chart_types', [])
        
        cursor.execute("""
            INSERT INTO agent_dashboard_requests 
            (title, description, category, data_source_id, chart_types, filters, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            data['title'],
            data['description'],
            data.get('category', 'Outros'),
            data.get('data_source') or None,
            chart_types,
            psycopg2.extras.Json(filters) if filters else None,
            session['user_id']
        ))
        
        request_id = cursor.fetchone()['id']
        conn.commit()
        
        # Log da ação
        cursor.execute("""
            INSERT INTO agent_logs (action_type, entity_type, entity_id, user_id, details)
            VALUES ('create', 'dashboard_request', %s, %s, %s)
        """, (request_id, session['user_id'], psycopg2.extras.Json({'title': data['title']})))
        conn.commit()
        
        return jsonify({"success": True, "id": request_id}), 201
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"[AGENT] Erro ao criar solicitação de dashboard: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-gen/<int:request_id>", methods=["DELETE"])
@login_required
def delete_dashboard_gen(request_id):
    """API para excluir uma solicitação de dashboard."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT created_by FROM agent_dashboard_requests WHERE id = %s", (request_id,))
        req = cursor.fetchone()
        
        if not req:
            return jsonify({"error": "Solicitação não encontrada"}), 404
        
        if req['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        cursor.execute("DELETE FROM agent_dashboard_requests WHERE id = %s", (request_id,))
        conn.commit()
        
        return jsonify({"success": True}), 200
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-gen/<int:request_id>", methods=["GET"])
@login_required
def get_dashboard_gen(request_id):
    """API para obter detalhes de uma solicitação de dashboard."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT d.*, u.nome_completo as created_by_name
            FROM agent_dashboard_requests d
            LEFT JOIN users_new u ON d.created_by = u.id
            WHERE d.id = %s
        """, (request_id,))
        dash = cursor.fetchone()
        
        if not dash:
            return jsonify({"error": "Solicitação não encontrada"}), 404
        
        if dash['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        return jsonify(dict(dash)), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/rpa/<int:rpa_id>", methods=["GET"])
@login_required
def get_rpa_details(rpa_id):
    """API para obter detalhes de uma RPA."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT r.*, t.name as type_name, u.nome_completo as created_by_name
            FROM agent_rpas r
            LEFT JOIN agent_rpa_types t ON r.rpa_type_id = t.id
            LEFT JOIN users_new u ON r.created_by = u.id
            WHERE r.id = %s
        """, (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            return jsonify({"error": "RPA não encontrada"}), 404
        
        if rpa['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        return jsonify(dict(rpa)), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/rpa/<int:rpa_id>/export", methods=["GET"])
@login_required
def export_rpa_to_excel(rpa_id):
    """Exporta os resultados de uma RPA para Excel."""
    from io import BytesIO
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT r.name, r.result, r.created_by
            FROM agent_rpas r
            WHERE r.id = %s
        """, (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            return jsonify({"error": "RPA não encontrada"}), 404
        
        if rpa['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        result = rpa.get('result') or {}
        data = result.get('data', [])
        
        if not data:
            return jsonify({"error": "Nenhum dado para exportar"}), 400
        
        # Criar Excel
        wb = load_workbook(filename=None) if False else None
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados"
        
        # Cabeçalhos
        if data and isinstance(data, list) and len(data) > 0:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Dados
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, '')
                    ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
        
        # Salvar em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nome do arquivo
        safe_name = re.sub(r'[^\w\s-]', '', rpa['name'])[:30]
        filename = f"rpa_{rpa_id}_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        app.logger.error(f"[EXPORT] Erro ao exportar RPA {rpa_id}: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/agent/rpa/<int:rpa_id>")
@login_required
def view_rpa_page(rpa_id):
    """Página para visualizar detalhes de uma RPA."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT r.*, t.name as type_name, u.nome_completo as created_by_name
            FROM agent_rpas r
            LEFT JOIN agent_rpa_types t ON r.rpa_type_id = t.id
            LEFT JOIN users_new u ON r.created_by = u.id
            WHERE r.id = %s
        """, (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            flash("RPA não encontrada", "error")
            return redirect(url_for('agent_page'))
        
        if rpa['created_by'] != session['user_id'] and session.get('role') != 'admin':
            flash("Permissão negada", "error")
            return redirect(url_for('agent_page'))
        
        # Buscar logs
        cursor.execute("""
            SELECT action_type, details, created_at
            FROM agent_logs
            WHERE entity_type = 'rpa' AND entity_id = %s
            ORDER BY created_at DESC
            LIMIT 20
        """, (rpa_id,))
        logs = cursor.fetchall()
        
        return render_template(
            "rpa_detail.html",
            rpa=rpa,
            logs=logs
        )
        
    except Exception as e:
        flash(f"Erro: {e}", "error")
        return redirect(url_for('agent_page'))
    finally:
        conn.close()


@app.route("/agent/dashboard-gen/<int:request_id>")
@login_required
def view_dashboard_gen_page(request_id):
    """Página para visualizar detalhes de uma solicitação de dashboard."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT d.*, u.nome_completo as created_by_name
            FROM agent_dashboard_requests d
            LEFT JOIN users_new u ON d.created_by = u.id
            WHERE d.id = %s
        """, (request_id,))
        dash = cursor.fetchone()
        
        if not dash:
            flash("Solicitação não encontrada", "error")
            return redirect(url_for('agent_page'))
        
        if dash['created_by'] != session['user_id'] and session.get('role') != 'admin':
            flash("Permissão negada", "error")
            return redirect(url_for('agent_page'))
        
        return render_template(
            "dashboard_gen_detail.html",
            dash=dash
        )
        
    except Exception as e:
        flash(f"Erro: {e}", "error")
        return redirect(url_for('agent_page'))
    finally:
        conn.close()


@app.route("/api/agent/dashboard-gen/<int:dash_id>/export", methods=["GET"])
@login_required
def export_dashboard_to_excel(dash_id):
    """Exporta os resultados de um dashboard para Excel."""
    from io import BytesIO
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT title, result_data, created_by
            FROM agent_dashboard_requests
            WHERE id = %s
        """, (dash_id,))
        dash = cursor.fetchone()
        
        if not dash:
            return jsonify({"error": "Dashboard não encontrado"}), 404
        
        if dash['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        result = dash.get('result_data') or {}
        data = result.get('data', [])
        
        if not data:
            return jsonify({"error": "Nenhum dado para exportar"}), 400
        
        # Criar Excel
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados"
        
        # Cabeçalhos
        if data and isinstance(data, list) and len(data) > 0:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Dados
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, '')
                    ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
        
        # Salvar em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nome do arquivo
        safe_name = re.sub(r'[^\w\s-]', '', dash['title'])[:30]
        filename = f"dashboard_{dash_id}_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        app.logger.error(f"[EXPORT] Erro ao exportar Dashboard {dash_id}: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-gen/<int:dash_id>/refresh", methods=["POST"])
@login_required
def refresh_dashboard(dash_id):
    """Recoloca um dashboard na fila para ser processado novamente."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT id, created_by FROM agent_dashboard_requests WHERE id = %s
        """, (dash_id,))
        dash = cursor.fetchone()
        
        if not dash:
            return jsonify({"error": "Dashboard não encontrado"}), 404
        
        if dash['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        # Recolocar na fila
        cursor.execute("""
            UPDATE agent_dashboard_requests 
            SET status = 'pending', 
                result_data = NULL,
                error_message = NULL,
                updated_at = NOW()
            WHERE id = %s
        """, (dash_id,))
        conn.commit()
        
        return jsonify({"success": True, "message": "Dashboard recolocado na fila"}), 200
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# --------------------------------------------------------------------------- #
# Editor de Dashboard (estilo Power BI)
# --------------------------------------------------------------------------- #
@app.route("/agent/dashboard-editor")
@app.route("/agent/dashboard-editor/new")
@login_required
def dashboard_editor_new():
    """Página do editor de dashboard - novo."""
    return render_template("dashboard_editor.html", template=None)


@app.route("/agent/dashboard-editor/<int:template_id>")
@login_required
def dashboard_editor(template_id):
    """Página do editor de dashboard - editar existente."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT * FROM agent_dashboard_templates WHERE id = %s AND created_by = %s
        """, (template_id, session['user_id']))
        template = cursor.fetchone()
        
        if not template:
            flash("Dashboard não encontrado", "error")
            return redirect(url_for('agent_page'))
        
        return render_template("dashboard_editor.html", template=dict(template))
        
    except Exception as e:
        flash(f"Erro: {e}", "error")
        return redirect(url_for('agent_page'))
    finally:
        conn.close()


@app.route("/agent/dashboard/<int:template_id>")
@login_required
def view_dashboard_template(template_id):
    """Visualizar dashboard publicado."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT t.*, u.nome_completo as created_by_name
            FROM agent_dashboard_templates t
            LEFT JOIN users_new u ON t.created_by = u.id
            WHERE t.id = %s
        """, (template_id,))
        template = cursor.fetchone()
        
        if not template:
            flash("Dashboard não encontrado", "error")
            return redirect(url_for('agent_page'))
        
        # Verificar permissão
        if not template['is_public'] and template['created_by'] != session['user_id'] and session.get('role') != 'admin':
            flash("Você não tem permissão para ver este dashboard", "error")
            return redirect(url_for('agent_page'))
        
        return render_template("dashboard_view.html", template=dict(template))
        
    except Exception as e:
        flash(f"Erro: {e}", "error")
        return redirect(url_for('agent_page'))
    finally:
        conn.close()


@app.route("/api/agent/dashboard-template", methods=["POST"])
@login_required
def create_dashboard_template():
    """Criar novo template de dashboard."""
    data = request.get_json()
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            INSERT INTO agent_dashboard_templates 
            (title, description, category, query_config, charts_config, layout_config, is_published, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            data.get('title', 'Novo Dashboard'),
            data.get('description', ''),
            data.get('category', 'Outros'),
            psycopg2.extras.Json(data.get('query_config', {})),
            psycopg2.extras.Json(data.get('charts_config', [])),
            psycopg2.extras.Json(data.get('layout_config', {})),
            data.get('is_published', False),
            session['user_id']
        ))
        
        new_id = cursor.fetchone()['id']
        conn.commit()
        
        return jsonify({"success": True, "id": new_id}), 201
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-template", methods=["PUT"])
@login_required
def update_dashboard_template():
    """Atualizar template de dashboard existente."""
    data = request.get_json()
    template_id = data.get('id')
    
    if not template_id:
        return jsonify({"error": "ID do template não fornecido"}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verificar propriedade
        cursor.execute("SELECT created_by FROM agent_dashboard_templates WHERE id = %s", (template_id,))
        template = cursor.fetchone()
        
        if not template or (template['created_by'] != session['user_id'] and session.get('role') != 'admin'):
            return jsonify({"error": "Permissão negada"}), 403
        
        cursor.execute("""
            UPDATE agent_dashboard_templates 
            SET title = %s, 
                description = %s, 
                category = %s, 
                query_config = %s, 
                charts_config = %s, 
                layout_config = %s, 
                is_published = %s,
                updated_at = NOW()
            WHERE id = %s
        """, (
            data.get('title'),
            data.get('description', ''),
            data.get('category', 'Outros'),
            psycopg2.extras.Json(data.get('query_config', {})),
            psycopg2.extras.Json(data.get('charts_config', [])),
            psycopg2.extras.Json(data.get('layout_config', {})),
            data.get('is_published', False),
            template_id
        ))
        
        conn.commit()
        return jsonify({"success": True, "id": template_id}), 200
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-template/<int:template_id>", methods=["DELETE"])
@login_required
def delete_dashboard_template(template_id):
    """Excluir template de dashboard."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT created_by FROM agent_dashboard_templates WHERE id = %s", (template_id,))
        template = cursor.fetchone()
        
        if not template or (template['created_by'] != session['user_id'] and session.get('role') != 'admin'):
            return jsonify({"error": "Permissão negada"}), 403
        
        cursor.execute("DELETE FROM agent_dashboard_templates WHERE id = %s", (template_id,))
        conn.commit()
        
        return jsonify({"success": True}), 200
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard-editor/execute-query", methods=["POST"])
@login_required
def execute_dashboard_query():
    """Executar query SQL para preview no editor de dashboard."""
    data = request.get_json()
    query = data.get('query', '').strip()
    
    if not query:
        return jsonify({"error": "Query não fornecida"}), 400
    
    # Validar query (apenas SELECT)
    if not query.upper().startswith('SELECT'):
        return jsonify({"error": "Apenas queries SELECT são permitidas"}), 400
    
    # Limitar resultados
    if 'LIMIT' not in query.upper():
        query = query.rstrip(';') + ' LIMIT 500'
    
    try:
        import pymysql
        
        host = os.getenv("MYSQL_AZ_HOST", "")
        port = int(os.getenv("MYSQL_AZ_PORT", "3307"))
        user = os.getenv("MYSQL_AZ_USER", "")
        password = os.getenv("MYSQL_AZ_PASSWORD", "")
        database = os.getenv("MYSQL_AZ_DB", "")
        
        if not all([host, user, password, database]):
            return jsonify({"error": "Credenciais MySQL não configuradas"}), 500
        
        mysql_conn = pymysql.connect(
            host=host, port=port, user=user, password=password, database=database,
            charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor,
            connect_timeout=30, read_timeout=60
        )
        
        with mysql_conn.cursor() as cursor:
            cursor.execute(query)
            rows = cursor.fetchall()
            
            # Converter tipos não serializáveis
            from decimal import Decimal
            from datetime import datetime, date
            
            def convert_value(v):
                if isinstance(v, Decimal):
                    return float(v)
                if isinstance(v, (datetime, date)):
                    return v.isoformat()
                if isinstance(v, bytes):
                    return v.decode('utf-8', errors='ignore')
                return v
            
            data = [{k: convert_value(v) for k, v in row.items()} for row in rows]
            fields = list(rows[0].keys()) if rows else []
        
        mysql_conn.close()
        
        return jsonify({
            "success": True,
            "data": data,
            "fields": fields,
            "row_count": len(data)
        }), 200
        
    except Exception as e:
        app.logger.error(f"[DASHBOARD-EDITOR] Erro ao executar query: {e}")
        return jsonify({"error": str(e)}), 500


# --------------------------------------------------------------------------- #
# Executor de RPAs - Conexão MySQL Brudam
# --------------------------------------------------------------------------- #
def get_brudam_db(database_name: str = None):
    """Conecta ao banco MySQL Brudam (azportoex ou portoexsp). Credenciais via .env"""
    import pymysql
    
    # Host atualizado conforme Workbench
    host = os.getenv("MYSQL_AZ_HOST", "portoex.db.brudam.com.br")
    
    # Tenta pegar porta do env, se falhar usa 3306 como padrão
    env_port = os.getenv("MYSQL_AZ_PORT", "3306")
    try:
        port = int(env_port)
    except ValueError:
        port = 3306
        
    user = os.getenv("MYSQL_AZ_USER", "")
    password = os.getenv("MYSQL_AZ_PASSWORD", "")
    database = database_name or os.getenv("MYSQL_AZ_DB", "")
    
    if not all([user, password, database]):
        # Se faltar credenciais, tenta usar valores hardcoded de emergência (baseado nas memórias)
        if not user: user = "consulta_portoex"
        if not database: database = database_name or "azportoex"
        # A senha não deve ficar no código, mas se necessário para teste local:
        # if not password: password = "..." 
        
    if not all([host, user, password, database]):
        raise ValueError("Credenciais MySQL Brudam não configuradas no .env")
    
    # Função auxiliar para tentar conexão
    def try_connect(target_host, target_port):
        return pymysql.connect(
            host=target_host,
            port=target_port,
            user=user,
            password=password,
            database=database,
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor,
            connect_timeout=10,
            read_timeout=60
        )

    # Lógica de tentativa:
    # 1. Tenta host/porta configurados
    # 2. Se falhar e for IP, tenta o DNS
    # 3. Se falhar e for DNS, tenta o IP antigo (fallback)
    
    try:
        app.logger.info(f"[MYSQL] Tentando conexão em {host}:{port}...")
        return try_connect(host, port)
    except pymysql.err.OperationalError as e:
        app.logger.warning(f"[MYSQL] Falha na conexão principal: {e}")
        
        # Se falhou e estamos usando o DNS, tenta o IP antigo como fallback
        if "portoex.db.brudam.com.br" in host:
            fallback_ip = "10.147.17.88"
            app.logger.info(f"[MYSQL] Tentando fallback para IP {fallback_ip}...")
            try:
                return try_connect(fallback_ip, port)
            except Exception:
                pass # Ignora erro do fallback para lançar o original
        
        # Se falhou e estamos usando IP, tenta o DNS
        elif "10." in host:
            fallback_host = "portoex.db.brudam.com.br"
            app.logger.info(f"[MYSQL] Tentando fallback para DNS {fallback_host}...")
            try:
                return try_connect(fallback_host, port)
            except Exception:
                pass
                
        raise e


def execute_rpa(rpa_id: int) -> dict:
    """Executa uma automação RPA e retorna o resultado."""
    import json
    from datetime import datetime
    
    conn = get_db()
    cursor = conn.cursor()
    logs = []
    result = {"success": False, "data": None, "error": None}
    
    try:
        # Buscar RPA
        cursor.execute("""
            SELECT r.*, t.name as type_name 
            FROM agent_rpas r
            LEFT JOIN agent_rpa_types t ON r.rpa_type_id = t.id
            WHERE r.id = %s
        """, (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            return {"success": False, "error": "RPA não encontrada"}
        
        # Atualizar status para 'running'
        cursor.execute("""
            UPDATE agent_rpas 
            SET status = 'running', executed_at = NOW() 
            WHERE id = %s
        """, (rpa_id,))
        conn.commit()
        
        logs.append(f"[{datetime.now().isoformat()}] Iniciando execução: {rpa['name']}")
        logs.append(f"[{datetime.now().isoformat()}] Tipo: {rpa['type_name']}")
        
        # Executar baseado no tipo
        type_name = rpa['type_name'] or ''
        parameters = rpa['parameters'] or {}
        
        if 'Extração de Dados' in type_name or 'brudam' in str(parameters).lower():
            # Conectar ao Brudam e executar query
            logs.append(f"[{datetime.now().isoformat()}] Conectando ao MySQL Brudam...")
            
            try:
                brudam_conn = get_brudam_db()
                brudam_cursor = brudam_conn.cursor()
                logs.append(f"[{datetime.now().isoformat()}] Conexão estabelecida com sucesso!")
                
                # Query padrão ou customizada
                query = parameters.get('query', 'SELECT 1 as test')
                limit = parameters.get('limit', 100)
                
                # Adicionar LIMIT se não existir
                if 'LIMIT' not in query.upper():
                    query = f"{query} LIMIT {limit}"
                
                logs.append(f"[{datetime.now().isoformat()}] Executando query...")
                brudam_cursor.execute(query)
                data = brudam_cursor.fetchall()
                
                logs.append(f"[{datetime.now().isoformat()}] Query executada! {len(data)} registros retornados.")
                
                result["success"] = True
                result["data"] = data
                result["row_count"] = len(data)
                
                brudam_cursor.close()
                brudam_conn.close()
                logs.append(f"[{datetime.now().isoformat()}] Conexão fechada.")
                
            except Exception as e:
                logs.append(f"[{datetime.now().isoformat()}] ERRO ao conectar/executar: {str(e)}")
                result["error"] = str(e)
        
        else:
            # Tipo genérico - apenas simula execução
            logs.append(f"[{datetime.now().isoformat()}] Executando automação genérica...")
            import time as time_module
            time_module.sleep(1)  # Simula processamento
            result["success"] = True
            result["data"] = {"message": "Automação executada com sucesso (simulação)"}
            logs.append(f"[{datetime.now().isoformat()}] Automação concluída.")
        
        # Atualizar RPA com resultado
        final_status = 'completed' if result["success"] else 'failed'
        cursor.execute("""
            UPDATE agent_rpas 
            SET status = %s, 
                completed_at = NOW(),
                result = %s,
                error_message = %s,
                updated_at = NOW()
            WHERE id = %s
        """, (
            final_status,
            psycopg2.extras.Json({"data": result.get("data"), "row_count": result.get("row_count", 0)}),
            result.get("error"),
            rpa_id
        ))
        conn.commit()
        
        # Salvar logs
        cursor.execute("""
            INSERT INTO agent_logs (action_type, entity_type, entity_id, user_id, details)
            VALUES ('execute', 'rpa', %s, %s, %s)
        """, (rpa_id, rpa['created_by'], psycopg2.extras.Json({"logs": logs, "success": result["success"]})))
        conn.commit()
        
        result["logs"] = logs
        return result
        
    except Exception as e:
        app.logger.error(f"[RPA] Erro ao executar RPA {rpa_id}: {e}")
        logs.append(f"[{datetime.now().isoformat()}] ERRO FATAL: {str(e)}")
        
        # Atualizar status para failed
        try:
            cursor.execute("""
                UPDATE agent_rpas 
                SET status = 'failed', error_message = %s, updated_at = NOW()
                WHERE id = %s
            """, (str(e), rpa_id))
            conn.commit()
        except:
            pass
        
        result["error"] = str(e)
        result["logs"] = logs
        return result
    finally:
        conn.close()


@app.route("/api/agent/rpa/<int:rpa_id>/execute", methods=["POST"])
@login_required
def execute_rpa_api(rpa_id):
    """API para executar uma automação RPA manualmente."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verificar permissão
        cursor.execute("SELECT created_by, status FROM agent_rpas WHERE id = %s", (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            return jsonify({"error": "RPA não encontrada"}), 404
        
        if rpa['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        if rpa['status'] == 'running':
            return jsonify({"error": "RPA já está em execução"}), 400
        
        conn.close()
        
        # Executar RPA
        result = execute_rpa(rpa_id)
        
        return jsonify(result), 200 if result["success"] else 500
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


@app.route("/api/agent/rpa/<int:rpa_id>/logs", methods=["GET"])
@login_required
def get_rpa_logs(rpa_id):
    """API para buscar logs de execução de uma RPA."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verificar permissão
        cursor.execute("SELECT created_by FROM agent_rpas WHERE id = %s", (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            return jsonify({"error": "RPA não encontrada"}), 404
        
        if rpa['created_by'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Permissão negada"}), 403
        
        # Buscar logs
        cursor.execute("""
            SELECT action_type, details, created_at
            FROM agent_logs
            WHERE entity_type = 'rpa' AND entity_id = %s
            ORDER BY created_at DESC
            LIMIT 50
        """, (rpa_id,))
        logs = [dict(row) for row in cursor.fetchall()]
        
        # Buscar resultado atual da RPA
        cursor.execute("""
            SELECT status, result, error_message, executed_at, completed_at
            FROM agent_rpas WHERE id = %s
        """, (rpa_id,))
        rpa_status = dict(cursor.fetchone())
        
        return jsonify({
            "logs": logs,
            "status": rpa_status
        }), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/brudam/test", methods=["POST"])
@login_required
@admin_required
def test_brudam_connection():
    """API para testar conexão com o banco Brudam."""
    logs = []
    
    try:
        from datetime import datetime
        logs.append(f"[{datetime.now().isoformat()}] Iniciando teste de conexão...")
        
        brudam_conn = get_brudam_db()
        logs.append(f"[{datetime.now().isoformat()}] Conexão estabelecida!")
        
        brudam_cursor = brudam_conn.cursor()
        
        # Listar tabelas
        brudam_cursor.execute("SHOW TABLES")
        tables = [list(row.values())[0] for row in brudam_cursor.fetchall()]
        logs.append(f"[{datetime.now().isoformat()}] {len(tables)} tabelas encontradas")
        
        brudam_cursor.close()
        brudam_conn.close()
        logs.append(f"[{datetime.now().isoformat()}] Conexão fechada com sucesso!")
        
        return jsonify({
            "success": True,
            "tables": tables[:20],  # Primeiras 20 tabelas
            "total_tables": len(tables),
            "logs": logs
        }), 200
        
    except Exception as e:
        logs.append(f"[{datetime.now().isoformat()}] ERRO: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "logs": logs
        }), 500


@app.route("/api/agent/auditoria-fiscal/request", methods=["POST"])
@login_required
def request_auditoria_fiscal():
    """Cria uma solicitação de auditoria fiscal para o agente local executar."""
    data = request.get_json()
    data_inicio = data.get('data_inicio')
    data_fim = data.get('data_fim')
    operador_id = data.get('operador_id')
    
    if not data_inicio or not data_fim:
        return jsonify({"error": "Datas de início e fim são obrigatórias"}), 400
    
    # Construir a query SQL para o agente executar
    query = f"""
        SELECT 
            m.id_manifesto,
            m.data_emissao,
            m.tipo,
            mt.tipo as tipo_descricao,
            m.id_agente,
            f.fantasia as agente_nome,
            m.motorista,
            func.nome as motorista_nome,
            m.veiculo,
            v.placa as veiculo_placa,
            v.modelo as veiculo_modelo,
            m.operador,
            u.primeiro_nome as operador_nome,
            m.picking,
            m.km_inicial,
            m.km_final,
            m.km_rodado,
            m.fatura,
            m.total_nf_valor,
            m.custo_motorista,
            m.custo_motorista_extra,
            m.adiantamento,
            m.pedagio
        FROM azportoex.manifesto m
        LEFT JOIN azportoex.manifesto_tipo mt ON m.tipo = mt.id_tipo
        LEFT JOIN azportoex.fornecedores f ON m.id_agente = f.id_local
        LEFT JOIN azportoex.funcionario func ON m.motorista = func.id_funcionario
        LEFT JOIN azportoex.usuarios u ON m.operador = u.id_usuario
        LEFT JOIN azportoex.veiculos v ON m.veiculo = v.id_veiculo
        WHERE m.data_emissao BETWEEN '{data_inicio}' AND '{data_fim}'
    """
    
    if operador_id:
        query += f" AND m.operador = {operador_id}"
    
    query += " ORDER BY m.data_emissao DESC, m.id_manifesto DESC"
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Criar solicitação na tabela de dashboards (reusando a estrutura)
        # category='auditoria' para identificar
        cursor.execute("""
            INSERT INTO agent_dashboard_requests 
            (title, description, category, chart_types, filters, created_by, status, created_at)
            VALUES (%s, %s, %s, %s::text[], %s, %s, 'pending', NOW())
            RETURNING id
        """, (
            f"Auditoria {data_inicio} a {data_fim}",
            "Auditoria Fiscal de Manifestos",
            "auditoria",
            ["table"],
            psycopg2.extras.Json({"query": query, "limit": 2000}),
            session['user_id']
        ))
        
        request_id = cursor.fetchone()['id']
        conn.commit()
        
        return jsonify({
            "success": True,
            "request_id": request_id,
            "message": "Solicitação enviada para o Agente Local"
        }), 200
        
    except Exception as e:
        conn.rollback()
        import traceback
        error_details = traceback.format_exc()
        app.logger.error(f"Erro ao criar solicitação de auditoria: {error_details}")
        return jsonify({"success": False, "error": str(e), "details": error_details}), 500
    finally:
        conn.close()


@app.route("/api/agent/auditoria-fiscal/status/<int:request_id>", methods=["GET"])
@login_required
def check_auditoria_status(request_id):
    """Verifica o status da solicitação de auditoria."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT status, result_data, error_message, updated_at
            FROM agent_dashboard_requests
            WHERE id = %s AND created_by = %s
        """, (request_id, session['user_id']))
        
        row = cursor.fetchone()
        
        if not row:
            return jsonify({"error": "Solicitação não encontrada"}), 404
            
        data = dict(row)
        
        # Se concluído, processar estatísticas no backend para aliviar o frontend
        if data['status'] == 'completed' and data.get('result_data'):
            result = data['result_data']
            manifestos = result.get('data', [])
            
            # Calcular estatísticas aqui se necessário, ou mandar tudo pro front
            # Vamos mandar tudo pro front processar por enquanto
            return jsonify({
                "success": True,
                "status": "completed",
                "manifestos": manifestos,
                "total_registros": len(manifestos)
            }), 200
            
        elif data['status'] == 'failed':
            return jsonify({
                "success": False,
                "status": "failed",
                "error": data.get('error_message')
            }), 200
            
        else:
            return jsonify({
                "success": True,
                "status": data['status']
            }), 200
            
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


def _build_relatorio_filters(
    data_inicio: str | None,
    data_fim: str | None,
    database: str = "azportoex",
    *,
    period: str | None = None,
    forma_pagamento_tokens: list[str] | None = None,
    forma_pagamento_exclude_tokens: list[str] | None = None,
) -> dict | None:
    """
    Monta payload de filtros para a procedure de Relatório de Resultados.
    
    Args:
        data_inicio: Data de início (YYYY-MM-DD) (usado somente quando period='custom' ou None)
        data_fim: Data de fim (YYYY-MM-DD) (usado somente quando period='custom' ou None)
        database: Nome do banco ('azportoex', 'portoexsp', ou 'ambas')
    """
    
    if database in ("portoexsp", "azportoex"):
        # Mapear período -> procedure segmentada (DBA)
        period_norm = (period or "").strip().lower()
        proc_map = {
            "mes_atual": "relatorio147mesatual",
            "mesmenos1": "relatorio147mesmenos1",
            "mes_menos1": "relatorio147mesmenos1",
            "mesmenos2": "relatorio147mesmenos2",
            "mes_menos2": "relatorio147mesmenos2",
            "mesmenos3": "relatorio147mesmenos3",
            "mes_menos3": "relatorio147mesmenos3",
        }
        procedure = proc_map.get(period_norm) or "relatorio147"
    else:
        # Para "ambas", retornamos None para indicar que precisa criar múltiplas solicitações
        return None
    
    return {
        "procedure": procedure,
        # relatorio147 antigo: não aceita params; ainda usamos para filtro posterior no agente local
        "procedure_params": {"data_inicio": data_inicio, "data_fim": data_fim} if data_inicio and data_fim else {},
        # Regras de negócio (defaults): ambos são opções na coluna forma_pagamento
        # Enviamos tokens para filtro "contém" no agent_local (AND).
        "forma_pagamento_tokens": forma_pagamento_tokens or ["Faturado", "A vista"],
        # Excluir formas indesejadas (ex.: CORTESIA)
        "forma_pagamento_exclude_tokens": forma_pagamento_exclude_tokens or ["Cortesia"],
        "limit": 5000,
        "database": database  # Incluir informação do banco nos filtros
    }


@app.route("/api/agent/relatorio-resultados/request", methods=["POST"])
@login_required
def request_relatorio_resultados():
    """Cria solicitação para executar o Relatório de Resultados na controladoria."""
    data = request.get_json() or {}
    # Período (novas procedures) + compatibilidade com range custom
    period = (data.get("period") or "mes_atual").strip()
    data_inicio = data.get("data_inicio")
    data_fim = data.get("data_fim")
    database = data.get("database", "azportoex")  # unidade: matriz/filial
    # Ambos são filtros na MESMA coluna forma_pagamento (ex.: "Faturado", "À vista")
    fp_tokens_raw = data.get("forma_pagamento_tokens")
    if isinstance(fp_tokens_raw, list):
        forma_pagamento_tokens = [str(x).strip() for x in fp_tokens_raw if str(x).strip()]
    else:
        # Compatibilidade: payload antigo enviava forma_pagamento string
        fp_single = (data.get("forma_pagamento") or "").strip()
        forma_pagamento_tokens = [fp_single] if fp_single else []
    if not forma_pagamento_tokens:
        forma_pagamento_tokens = ["Faturado", "A vista"]

    fp_excl_raw = data.get("forma_pagamento_exclude_tokens")
    if isinstance(fp_excl_raw, list):
        forma_pagamento_exclude_tokens = [str(x).strip() for x in fp_excl_raw if str(x).strip()]
    else:
        forma_pagamento_exclude_tokens = ["Cortesia"]
    if not forma_pagamento_exclude_tokens:
        forma_pagamento_exclude_tokens = ["Cortesia"]

    def _parse_date(s: str):
        from datetime import datetime
        return datetime.strptime(s, "%Y-%m-%d")

    def _months_back_limit_ok(di: str, df: str) -> bool:
        from datetime import datetime
        today = datetime.now().date()
        # Janela: no máximo 3 meses atrás (aprox. por mês calendário)
        min_date = datetime(today.year, today.month, 1).date()
        # voltar 3 meses no calendário:
        y = min_date.year
        m = min_date.month - 3
        while m <= 0:
            m += 12
            y -= 1
        min_date = datetime(y, m, 1).date()
        d1 = _parse_date(di).date()
        d2 = _parse_date(df).date()
        return d1 >= min_date and d2 >= min_date and d1 <= today and d2 <= today

    # Se period for custom (ou vazio), exige datas e limita a -3 meses
    if (not period or period.strip().lower() in ("custom", "personalizado")):
        if not data_inicio or not data_fim:
            return jsonify({"success": False, "error": "Datas de início e fim são obrigatórias para período custom"}), 400
        if not _months_back_limit_ok(data_inicio, data_fim):
            return jsonify({"success": False, "error": "Período inválido: permitido apenas até 3 meses atrás."}), 400

    conn = get_db()
    cursor = conn.cursor()

    try:
        request_ids = []

        # Custom por data: quebrar por mês (-3) usando as procedures segmentadas e filtrar no agente
        if period.strip().lower() in ("custom", "personalizado"):
            from datetime import datetime

            di = _parse_date(data_inicio)
            df = _parse_date(data_fim)
            # Normalizar ordem
            if df < di:
                di, df = df, di

            # Iterar meses no range
            months = []
            cur = datetime(di.year, di.month, 1)
            end_m = datetime(df.year, df.month, 1)
            while cur <= end_m:
                months.append((cur.year, cur.month))
                # next month
                ny = cur.year + (1 if cur.month == 12 else 0)
                nm = 1 if cur.month == 12 else cur.month + 1
                cur = datetime(ny, nm, 1)

            # Mapear (ano,mes) -> offset (0..3) relativo ao mês atual
            now = datetime.now()
            cur_month = datetime(now.year, now.month, 1)

            def _month_offset(y: int, m: int) -> int:
                return (cur_month.year - y) * 12 + (cur_month.month - m)

            # Criar requests por mês (máximo 4)
            for (yy, mm) in months:
                off = _month_offset(yy, mm)
                if off < 0 or off > 3:
                    continue
                per = "mes_atual" if off == 0 else ("mes_menos1" if off == 1 else ("mes_menos2" if off == 2 else "mes_menos3"))

                if database == "ambas":
                    for db_name in ["azportoex", "portoexsp"]:
                        filters_payload = _build_relatorio_filters(
                            data_inicio,
                            data_fim,
                            db_name,
                            period=per,
                            forma_pagamento_tokens=forma_pagamento_tokens,
                            forma_pagamento_exclude_tokens=forma_pagamento_exclude_tokens,
                        )
                        cursor.execute("""
                            INSERT INTO agent_dashboard_requests
                            (title, description, category, chart_types, filters, created_by, status, created_at)
                            VALUES (%s, %s, %s, %s::text[], %s, %s, 'pending', NOW())
                            RETURNING id
                        """, (
                            f"Relatório Resultados {db_name} (custom {yy}-{str(mm).zfill(2)})",
                            f"Relatório gerencial da controladoria - {db_name}",
                            "relatorio_resultados",
                            ["table", "cards"],
                            psycopg2.extras.Json(filters_payload),
                            session['user_id']
                        ))
                        request_ids.append(cursor.fetchone()['id'])
                else:
                    filters_payload = _build_relatorio_filters(
                        data_inicio,
                        data_fim,
                        database,
                        period=per,
                        forma_pagamento_tokens=forma_pagamento_tokens,
                        forma_pagamento_exclude_tokens=forma_pagamento_exclude_tokens,
                    )
                    cursor.execute("""
                        INSERT INTO agent_dashboard_requests
                        (title, description, category, chart_types, filters, created_by, status, created_at)
                        VALUES (%s, %s, %s, %s::text[], %s, %s, 'pending', NOW())
                        RETURNING id
                    """, (
                        f"Relatório Resultados {database} (custom {yy}-{str(mm).zfill(2)})",
                        f"Relatório gerencial da controladoria - {database}",
                        "relatorio_resultados",
                        ["table", "cards"],
                        psycopg2.extras.Json(filters_payload),
                        session['user_id']
                    ))
                    request_ids.append(cursor.fetchone()['id'])

            conn.commit()
            return jsonify({
                "success": True,
                "request_ids": request_ids,
                "message": f"Solicitações enviadas para o Agente Local (custom por mês: {len(request_ids)})"
            }), 200
        
        if database == "ambas":
            # Criar duas solicitações separadas, uma para cada banco
            for db_name in ["azportoex", "portoexsp"]:
                filters_payload = _build_relatorio_filters(
                    data_inicio,
                    data_fim,
                    db_name,
                    period=period,
                    forma_pagamento_tokens=forma_pagamento_tokens,
                    forma_pagamento_exclude_tokens=forma_pagamento_exclude_tokens,
                )
                
                cursor.execute("""
                    INSERT INTO agent_dashboard_requests
                    (title, description, category, chart_types, filters, created_by, status, created_at)
                    VALUES (%s, %s, %s, %s::text[], %s, %s, 'pending', NOW())
                    RETURNING id
                """, (
                    f"Relatório Resultados {db_name} ({period})",
                    f"Relatório gerencial da controladoria - {db_name}",
                    "relatorio_resultados",
                    ["table", "cards"],
                    psycopg2.extras.Json(filters_payload),
                    session['user_id']
                ))
                
                request_id = cursor.fetchone()['id']
                request_ids.append(request_id)
            
            conn.commit()
            
            return jsonify({
                "success": True,
                "request_ids": request_ids,
                "message": f"Solicitações enviadas para o Agente Local (azportoex e portoexsp)"
            }), 200
        else:
            # Criar uma única solicitação
            filters_payload = _build_relatorio_filters(
                data_inicio,
                data_fim,
                database,
                period=period,
                forma_pagamento_tokens=forma_pagamento_tokens,
                forma_pagamento_exclude_tokens=forma_pagamento_exclude_tokens,
            )
            
            if not filters_payload:
                return jsonify({"success": False, "error": "Banco de dados inválido"}), 400

            cursor.execute("""
                INSERT INTO agent_dashboard_requests
                (title, description, category, chart_types, filters, created_by, status, created_at)
                VALUES (%s, %s, %s, %s::text[], %s, %s, 'pending', NOW())
                RETURNING id
            """, (
                f"Relatório Resultados {database} ({period})",
                f"Relatório gerencial da controladoria - {database}",
                "relatorio_resultados",
                ["table", "cards"],
                psycopg2.extras.Json(filters_payload),
                session['user_id']
            ))

            request_id = cursor.fetchone()['id']
            conn.commit()

            return jsonify({
                "success": True,
                "request_id": request_id,
                "message": f"Solicitação enviada para o Agente Local ({database})"
            }), 200

    except Exception as e:
        conn.rollback()
        app.logger.error(f"Erro ao criar solicitação Relatório de Resultados: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/relatorio-resultados/status/<int:request_id>", methods=["GET"])
@login_required
def check_relatorio_resultados_status(request_id):
    """Consulta o status da solicitação e retorna dados consolidados quando concluída."""
    conn = get_db()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT status, result_data, error_message, filters
            FROM agent_dashboard_requests
            WHERE id = %s AND created_by = %s
        """, (request_id, session['user_id']))

        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "Solicitação não encontrada"}), 404

        status = row['status']
        if status == 'completed' and row.get('result_data'):
            payload = row['result_data'] or {}
            registros = payload.get('data', [])
            # Usar row_count do payload (pode ser maior que len(registros) se houver truncamento)
            # Mas como agora não há truncamento, deve ser igual
            row_count = payload.get('row_count', len(registros))
            return jsonify({
                "success": True,
                "status": "completed",
                "registros": registros,
                "row_count": row_count,  # Usar contagem do payload para garantir precisão
                "requested_range": (row.get('filters') or {}).get('procedure_params', {})
            }), 200
        elif status == 'processing':
            # Ainda processando chunks
            payload = row.get('result_data') or {}
            chunks_info = {
                "chunks_received": payload.get('_chunks_received', []),
                "total_chunks": payload.get('_total_chunks', 0),
                "current_records": len(payload.get('data', []))
            }
            return jsonify({
                "success": True,
                "status": "processing",
                "message": f"Processando chunks: {len(chunks_info['chunks_received'])}/{chunks_info['total_chunks']} recebidos",
                "chunks_info": chunks_info
            }), 200

        if status == 'failed':
            return jsonify({
                "success": False,
                "status": "failed",
                "error": row.get('error_message')
            }), 200

        return jsonify({
            "success": True,
            "status": status
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/relatorio-entregas", methods=["GET"])
@login_required
def api_relatorio_entregas():
    """
    Relatório de Entregas (Script 376) - busca via Agente Local.
    Cria solicitação para o agente (brudam_agent.py) executar as queries no MySQL Brudam
    e retorna 202 com request_id para o frontend fazer polling no endpoint de status.
    """
    database = request.args.get("database", "azportoex")
    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")

    hoje = datetime.now().date()
    if not data_inicio:
        data_inicio = date(hoje.year, hoje.month, 1).strftime("%Y-%m-%d")
    if not data_fim:
        data_fim = hoje.strftime("%Y-%m-%d")

    filters = {
        "runner": "relatorio_entregas",
        "database": database,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
    }

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO agent_dashboard_requests
            (title, description, category, chart_types, filters, created_by, status, created_at)
            VALUES (%s, %s, %s, %s::text[], %s, %s, 'pending', NOW())
            RETURNING id
        """, (
            f"Relatório Entregas {database} ({data_inicio} a {data_fim})",
            "Relatório de Entregas por Status (Script 376)",
            "relatorio_entregas",
            ["doughnut", "table"],
            psycopg2.extras.Json(filters),
            session["user_id"],
        ))
        request_id = cursor.fetchone()["id"]
        conn.commit()
        return jsonify({
            "pending": True,
            "request_id": request_id,
            "message": "Solicitação enviada ao Agente Local. Aguarde o processamento.",
        }), 202
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Erro ao criar solicitação relatório entregas: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/relatorio-entregas/status/<int:request_id>", methods=["GET"])
@login_required
def api_relatorio_entregas_status(request_id):
    """
    Consulta o status da solicitação de Relatório de Entregas.
    Retorna os dados quando o agente concluir o processamento.
    """
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT id, status, result_data, error_message, category
            FROM agent_dashboard_requests
            WHERE id = %s AND created_by = %s AND category = 'relatorio_entregas'
        """, (request_id, session["user_id"]))
        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "Solicitação não encontrada"}), 404

        status = row["status"]
        if status == "pending" or status == "processing":
            return jsonify({
                "pending": True,
                "request_id": request_id,
                "status": status,
                "message": "Aguardando processamento pelo Agente Local...",
            }), 200

        if status == "failed":
            return jsonify({
                "success": False,
                "error": row.get("error_message") or "Falha no processamento",
                "status_counts": {},
                "total": 0,
            }), 200

        # completed
        rd = row.get("result_data") or {}
        if not isinstance(rd, dict):
            rd = {}
        data = rd.get("data")
        if isinstance(data, dict):
            return jsonify(data), 200
        return jsonify({
            "success": False,
            "error": "Formato de dados inválido",
            "status_counts": {},
            "total": 0,
        }), 200
    finally:
        conn.close()


@app.route("/api/agent/auditoria-fiscal/operadores", methods=["GET"])
@login_required
def get_operadores_auditoria():
    """
    API para listar operadores.
    NOTA: Como o servidor não tem acesso ao banco, esta API vai tentar 
    buscar de um cache local ou retornar lista vazia para o usuário digitar o ID.
    Futuramente pode ser implementado via request assíncrono.
    """
    return jsonify({
        "success": True,
        "operadores": [],
        "message": "Lista dinâmica indisponível offline. Digite o ID se souber."
    }), 200


@app.route("/api/agent/brudam/query", methods=["POST"])
@login_required
@admin_required
def execute_brudam_query():
    """API para executar query no banco Brudam (apenas admin)."""
    data = request.get_json()
    query = data.get('query', '')
    limit = data.get('limit', 100)
    
    if not query:
        return jsonify({"error": "Query é obrigatória"}), 400
    
    # Segurança: apenas SELECT permitido
    if not query.strip().upper().startswith('SELECT'):
        return jsonify({"error": "Apenas queries SELECT são permitidas"}), 403
    
    # Adicionar LIMIT se não existir
    if 'LIMIT' not in query.upper():
        query = f"{query} LIMIT {limit}"
    
    logs = []
    
    try:
        from datetime import datetime
        logs.append(f"[{datetime.now().isoformat()}] Executando query...")
        
        brudam_conn = get_brudam_db()
        brudam_cursor = brudam_conn.cursor()
        
        brudam_cursor.execute(query)
        data = brudam_cursor.fetchall()
        
        logs.append(f"[{datetime.now().isoformat()}] {len(data)} registros retornados")
        
        brudam_cursor.close()
        brudam_conn.close()
        
        return jsonify({
            "success": True,
            "data": data,
            "row_count": len(data),
            "logs": logs
        }), 200
        
    except Exception as e:
        logs.append(f"[{datetime.now().isoformat()}] ERRO: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "logs": logs
        }), 500


# --------------------------------------------------------------------------- #
# APIs para Agente Local (Brudam)
# --------------------------------------------------------------------------- #
AGENT_API_KEY = os.getenv("AGENT_API_KEY", "")


def verify_agent_api_key():
    """Verifica a chave de API do agente local."""
    # Se não houver chave configurada no servidor, permite acesso (modo desenvolvimento)
    if not AGENT_API_KEY:
        return True
        
    api_key = request.headers.get("X-API-Key", "")
    return api_key and api_key == AGENT_API_KEY


@app.route("/api/agent/sync/knowledge", methods=["POST"])
def sync_knowledge():
    """
    API para o agente local enviar dados do Brudam para a Base de Conhecimento.
    Usa X-API-Key para autenticação.
    """
    if not verify_agent_api_key():
        return jsonify({"error": "API Key inválida"}), 401
    
    data = request.get_json()
    items = data.get('items', []) # Lista de {question, answer, category}
    
    if not items:
        return jsonify({"error": "Nenhum item fornecido"}), 400
        
    conn = get_db()
    cursor = conn.cursor()
    
    count = 0
    try:
        for item in items:
            question = item.get('question')
            answer = item.get('answer')
            category = item.get('category', 'Brudam Sync')
            
            if question and answer:
                # Upsert (Insere ou Atualiza se a pergunta for idêntica)
                cursor.execute("""
                    SELECT id FROM agent_knowledge_base 
                    WHERE question = %s AND category = %s
                """, (question, category))
                existing = cursor.fetchone()
                
                if existing:
                    cursor.execute("""
                        UPDATE agent_knowledge_base 
                        SET answer = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (answer, existing['id']))
                else:
                    cursor.execute("""
                        INSERT INTO agent_knowledge_base (question, answer, category, created_by)
                        VALUES (%s, %s, %s, 0)
                    """, (question, answer, category))
                count += 1
        
        conn.commit()
        return jsonify({"success": True, "count": count}), 200
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"[AGENT-SYNC] Erro ao sincronizar conhecimento: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/rpas/pending", methods=["GET"])
def get_pending_rpas():
    """API para o agente local buscar RPAs pendentes."""
    if not verify_agent_api_key():
        return jsonify({"error": "API Key inválida"}), 401
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Buscar RPAs pendentes do tipo "Extração de Dados" ou com parâmetros brudam
        cursor.execute("""
            SELECT r.id, r.name, r.description, r.parameters, r.priority,
                   t.name as type_name
            FROM agent_rpas r
            LEFT JOIN agent_rpa_types t ON r.rpa_type_id = t.id
            WHERE r.status = 'pending'
              AND (t.name LIKE '%Extração%' OR r.parameters::text LIKE '%brudam%' OR r.parameters::text LIKE '%query%')
            ORDER BY 
                CASE r.priority 
                    WHEN 'critical' THEN 1 
                    WHEN 'high' THEN 2 
                    WHEN 'medium' THEN 3 
                    ELSE 4 
                END,
                r.created_at ASC
            LIMIT 10
        """)
        rpas = [dict(row) for row in cursor.fetchall()]
        
        # Marcar como "running" para evitar execução duplicada
        for rpa in rpas:
            cursor.execute("""
                UPDATE agent_rpas 
                SET status = 'running', executed_at = NOW() 
                WHERE id = %s AND status = 'pending'
            """, (rpa['id'],))
        conn.commit()
        
        return jsonify({"rpas": rpas}), 200
        
    except Exception as e:
        app.logger.error(f"[AGENT-API] Erro ao buscar RPAs pendentes: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/rpa/<int:rpa_id>/result", methods=["POST"])
def receive_rpa_result(rpa_id):
    """API para o agente local enviar resultado da execução."""
    if not verify_agent_api_key():
        return jsonify({"error": "API Key inválida"}), 401
    
    data = request.get_json()
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verificar se RPA existe
        cursor.execute("SELECT id, created_by FROM agent_rpas WHERE id = %s", (rpa_id,))
        rpa = cursor.fetchone()
        
        if not rpa:
            return jsonify({"error": "RPA não encontrada"}), 404
        
        # Atualizar RPA com resultado
        success = data.get("success", False)
        final_status = "completed" if success else "failed"
        
        # Limitar tamanho dos dados para evitar problemas de armazenamento
        result_data = data.get("data", [])
        if isinstance(result_data, list) and len(result_data) > 1000:
            result_data = result_data[:1000]  # Limitar a 1000 registros
        
        cursor.execute("""
            UPDATE agent_rpas 
            SET status = %s, 
                completed_at = NOW(),
                result = %s,
                error_message = %s,
                updated_at = NOW()
            WHERE id = %s
        """, (
            final_status,
            psycopg2.extras.Json({
                "data": result_data,
                "row_count": data.get("row_count", 0),
                "source": "agent_local"
            }),
            data.get("error"),
            rpa_id
        ))
        
        # Salvar logs
        logs = data.get("logs", [])
        cursor.execute("""
            INSERT INTO agent_logs (action_type, entity_type, entity_id, user_id, details)
            VALUES ('execute_remote', 'rpa', %s, %s, %s)
        """, (rpa_id, rpa['created_by'], psycopg2.extras.Json({
            "logs": logs,
            "success": success,
            "source": "agent_local"
        })))
        
        conn.commit()
        
        app.logger.info(f"[AGENT-API] Resultado recebido para RPA #{rpa_id}: {final_status}")
        return jsonify({"success": True}), 200
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"[AGENT-API] Erro ao salvar resultado: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboards/pending", methods=["GET"])
def get_pending_dashboards():
    """API para o agente local buscar solicitações de dashboard pendentes."""
    if not verify_agent_api_key():
        return jsonify({"error": "API Key inválida"}), 401
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Garantir colunas de lease (migração incremental)
        cursor.execute("ALTER TABLE agent_dashboard_requests ADD COLUMN IF NOT EXISTS leased_by TEXT;")
        cursor.execute("ALTER TABLE agent_dashboard_requests ADD COLUMN IF NOT EXISTS leased_until TIMESTAMPTZ;")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_agent_dashboard_requests_lease ON agent_dashboard_requests(leased_until);")
        conn.commit()

        agent_id = (request.headers.get("X-Agent-Id") or request.headers.get("X-Agent-ID") or "").strip() or "agent_local"
        try:
            limit = int(request.args.get("limit", "5"))
        except Exception:
            limit = 5
        limit = max(1, min(limit, 25))

       
        cursor.execute(
            """
            WITH to_claim AS (
              SELECT id
              FROM agent_dashboard_requests
              WHERE status = 'pending'
                AND (leased_until IS NULL OR leased_until < NOW())
                AND filters IS NOT NULL
                AND (
                      filters::text ILIKE '%%"query"%%'
                   OR filters::text ILIKE '%%"procedure"%%'
                   OR filters::text ILIKE '%%"runner"%%'
                )
              ORDER BY created_at ASC
              FOR UPDATE SKIP LOCKED
              LIMIT %s
            )
            UPDATE agent_dashboard_requests d
            SET status = 'processing',
                leased_by = %s,
                leased_until = NOW() + INTERVAL '20 minutes',
                updated_at = NOW()
            FROM to_claim c
            WHERE d.id = c.id
            RETURNING d.id, d.title, d.description, d.category, d.chart_types, d.filters, d.created_by
            """,
            (limit, agent_id),
        )
        dashboards = [dict(row) for row in cursor.fetchall()]
        conn.commit()
        
        return jsonify({"dashboards": dashboards}), 200
        
    except Exception as e:
        app.logger.error(f"[AGENT-API] Erro ao buscar dashboards pendentes: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard/<int:dash_id>/progress", methods=["POST"])
def receive_dashboard_progress(dash_id: int):
    """API para o agente local enviar progresso/logs parciais do dashboard (para streaming/polling no modal)."""
    if not verify_agent_api_key():
        return jsonify({"error": "API Key inválida"}), 401

    payload = request.get_json() or {}
    logs_append = payload.get("logs_append") or []
    progress = payload.get("progress") or {}

    # Sanitização
    if not isinstance(logs_append, list):
        logs_append = [str(logs_append)]
    logs_append = [str(x) for x in logs_append if x is not None]
    if not isinstance(progress, dict):
        progress = {}

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            SELECT id, result_data
            FROM agent_dashboard_requests
            WHERE id = %s
            """,
            (dash_id,),
        )
        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "Dashboard não encontrado"}), 404

        rd = row.get("result_data") or {}
        if not isinstance(rd, dict):
            rd = {}

        existing_logs = rd.get("_logs") or []
        if not isinstance(existing_logs, list):
            existing_logs = []

        # Evitar crescimento infinito no Postgres
        combined_logs = (existing_logs + logs_append)[-3000:]
        rd["_logs"] = combined_logs

        # Progress opcional (ex.: {"percent": 35, "text": "..."} )
        if progress:
            rd["_progress"] = progress

        cursor.execute(
            """
            UPDATE agent_dashboard_requests
            SET status = 'processing',
                result_data = %s,
                updated_at = NOW()
            WHERE id = %s
            """,
            (psycopg2.extras.Json(rd), dash_id),
        )
        conn.commit()
        return jsonify({"success": True}), 200
    except Exception as e:
        conn.rollback()
        app.logger.error(f"[AGENT-API] Erro ao salvar progresso dashboard #{dash_id}: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/dashboard/<int:dash_id>/result", methods=["POST"])
def receive_dashboard_result(dash_id):
    """API para o agente local enviar resultado do dashboard (suporta chunks e compressão)."""
    if not verify_agent_api_key():
        return jsonify({"error": "API Key inválida"}), 401
    
    # Descomprimir dados se vierem comprimidos
    if request.headers.get("Content-Encoding") == "gzip":
        try:
            decompressed = gzip.decompress(request.data)
            data = json.loads(decompressed.decode('utf-8'))
            app.logger.info(f"[AGENT-API] Dados comprimidos recebidos e descomprimidos para Dashboard #{dash_id}")
        except Exception as e:
            app.logger.error(f"[AGENT-API] Erro ao descomprimir dados: {e}")
            return jsonify({"error": f"Erro ao descomprimir dados: {str(e)}"}), 400
    else:
        data = request.get_json()
        if not data:
            return jsonify({"error": "Dados inválidos"}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        success = data.get("success", True)
        # Verificar se dashboard existe e buscar dados existentes
        cursor.execute("""
            SELECT id, created_by, result_data 
            FROM agent_dashboard_requests 
            WHERE id = %s
        """, (dash_id,))
        dash = cursor.fetchone()
        
        if not dash:
            return jsonify({"error": "Dashboard não encontrado"}), 404
        
        # Verificar se é um chunk
        chunk_info = data.get("_chunk_info", {})
        is_chunk = bool(chunk_info)
        
        if is_chunk:
            # É um chunk - combinar com dados existentes
            chunk_index = chunk_info.get("chunk_index", 0)
            total_chunks = chunk_info.get("total_chunks", 1)
            total_records = chunk_info.get("total_records", 0)
            
            app.logger.info(f"[AGENT-API] Recebendo chunk {chunk_index + 1}/{total_chunks} para Dashboard #{dash_id} "
                          f"(registros {chunk_info.get('chunk_start', 0)}-{chunk_info.get('chunk_end', 0)})")
            
            # Buscar dados existentes ou inicializar
            existing_result = dash.get("result_data")
            if existing_result and isinstance(existing_result, dict):
                combined_data = list(existing_result.get("data", []))
                chunks_received = list(existing_result.get("_chunks_received", []))
            else:
                combined_data = []
                chunks_received = []
            
            # Verificar se este chunk já foi recebido (evitar duplicação)
            if chunk_index in chunks_received:
                app.logger.warning(f"[AGENT-API] Chunk {chunk_index} já foi recebido anteriormente, ignorando duplicata")
                return jsonify({"success": True, "message": "Chunk já recebido"}), 200
            
            # Adicionar dados do chunk
            chunk_data = data.get("data", [])
            combined_data.extend(chunk_data)
            chunks_received.append(chunk_index)
            chunks_received.sort()  # Ordenar para facilitar verificação
            
            # Preparar payload combinado
            all_chunks_received = len(chunks_received) >= total_chunks
            payload_data = {
                "data": combined_data,
                "row_count": total_records,  # Contagem total esperada
                "source": "agent_local",
                "_chunks_received": chunks_received,
                "_total_chunks": total_chunks,
                "_is_complete": all_chunks_received
            }
            
            # Se todos os chunks foram recebidos, marcar como completo
            if all_chunks_received:
                final_status = "completed" if success else "failed"
                app.logger.info(f"[AGENT-API] ✅ Todos os {total_chunks} chunks recebidos para Dashboard #{dash_id}: {len(combined_data)} registros totais")
            else:
                final_status = "processing"  # Ainda aguardando mais chunks
                app.logger.info(f"[AGENT-API] Chunk {chunk_index + 1}/{total_chunks} recebido. "
                              f"Chunks recebidos: {chunks_received}. Aguardando mais chunks...")
        else:
            # Não é chunk - dados completos de uma vez
            result_data = data.get("data", [])
            # Evitar 500 quando agente manda data=null em falhas
            if result_data is None:
                result_data = []
            final_status = "completed" if success else "failed"
            
            app.logger.info(f"[AGENT-API] Recebendo resultado completo para Dashboard #{dash_id}: {len(result_data)} registros")
            
            payload_data = {
                "data": result_data,
                "row_count": data.get("row_count", len(result_data)),
                "source": "agent_local"
            }
        
        # Atualizar dashboard (SEM LIMITAÇÃO - armazena todos os dados)
        cursor.execute("""
            UPDATE agent_dashboard_requests 
            SET status = %s, 
                result_data = %s,
                error_message = %s,
                completed_at = CASE WHEN %s = 'completed' THEN NOW() ELSE completed_at END,
                updated_at = NOW()
            WHERE id = %s
        """, (
            final_status,
            psycopg2.extras.Json(payload_data),
            data.get("error"),
            final_status,
            dash_id
        ))
        
        # Salvar logs
        logs = data.get("logs", [])
        cursor.execute("""
            INSERT INTO agent_logs (action_type, entity_type, entity_id, user_id, details)
            VALUES ('execute_remote', 'dashboard', %s, %s, %s)
        """, (dash_id, dash['created_by'], psycopg2.extras.Json({
            "logs": logs,
            "success": success,
            "source": "agent_local"
        })))
        
        conn.commit()
        
        app.logger.info(f"[AGENT-API] Resultado recebido para Dashboard #{dash_id}: {final_status}")
        return jsonify({"success": True}), 200
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"[AGENT-API] Erro ao salvar resultado dashboard: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/health", methods=["GET"])
def agent_health_check():
    """Health check para o agente local."""
    return jsonify({
        "status": "ok",
        "timestamp": datetime.now().isoformat(),
        "version": "1.0.0"
    }), 200


# --------------------------------------------------------------------------- #
# Chat IA & RAG
# --------------------------------------------------------------------------- #

@app.route("/api/agent/chat/history", methods=["GET"])
@login_required
def get_chat_history():
    """Retorna o histórico de conversas do usuário."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Buscar conversas ordenadas pela última atualização
        cursor.execute("""
            SELECT c.id, c.title, c.updated_at,
                   (SELECT content FROM agent_messages m WHERE m.conversation_id = c.id ORDER BY m.created_at DESC LIMIT 1) as last_message
            FROM agent_conversations c
            WHERE c.user_id = %s AND c.is_archived = false
            ORDER BY c.updated_at DESC
            LIMIT 50
        """, (session['user_id'],))
        
        conversations = [dict(row) for row in cursor.fetchall()]
        return jsonify({"conversations": conversations})
        
    except Exception as e:
        app.logger.error(f"Erro ao buscar histórico de chat: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/chat/<int:conversation_id>/messages", methods=["GET"])
@login_required
def get_chat_messages(conversation_id):
    """Retorna as mensagens de uma conversa."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verificar permissão
        cursor.execute("SELECT title, user_id FROM agent_conversations WHERE id = %s", (conversation_id,))
        conv = cursor.fetchone()
        
        if not conv:
            return jsonify({"error": "Conversa não encontrada"}), 404
            
        if conv['user_id'] != session['user_id'] and session.get('role') != 'admin':
            return jsonify({"error": "Acesso negado"}), 403
            
        # Buscar mensagens
        cursor.execute("""
            SELECT role, content, created_at, metadata
            FROM agent_messages
            WHERE conversation_id = %s
            ORDER BY created_at ASC
        """, (conversation_id,))
        
        messages = [dict(row) for row in cursor.fetchall()]
        return jsonify({
            "title": conv['title'],
            "messages": messages
        })
        
    except Exception as e:
        app.logger.error(f"Erro ao buscar mensagens: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/chat/message", methods=["POST"])
@login_required
def send_chat_message():
    """Envia uma mensagem e obtém resposta via RAG."""
    data = request.get_json()
    user_message = data.get('message')
    conversation_id = data.get('conversation_id')
    
    if not user_message:
        return jsonify({"error": "Mensagem vazia"}), 400
        
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # 1. Gerenciar Conversa (Criar ou Atualizar)
        if not conversation_id:
            title = ' '.join(user_message.split()[:5]) + '...'
            cursor.execute("""
                INSERT INTO agent_conversations (title, user_id)
                VALUES (%s, %s)
                RETURNING id
            """, (title, session['user_id']))
            conversation_id = cursor.fetchone()['id']
        else:
            # Se o frontend mandou um conversation_id antigo/inexistente (ex.: cache/localStorage),
            # garantimos que a conversa existe para evitar violação de FK em agent_messages.
            cursor.execute(
                "SELECT id FROM agent_conversations WHERE id = %s AND user_id = %s",
                (conversation_id, session["user_id"]),
            )
            existing = cursor.fetchone()
            if not existing:
                # Cria nova conversa e ignora o id antigo
                title = " ".join(user_message.split()[:5]) + "..."
                cursor.execute(
                    """
                    INSERT INTO agent_conversations (title, user_id)
                    VALUES (%s, %s)
                    RETURNING id
                    """,
                    (title, session["user_id"]),
                )
                conversation_id = cursor.fetchone()["id"]
            else:
                cursor.execute(
                    "UPDATE agent_conversations SET updated_at = NOW() WHERE id = %s AND user_id = %s",
                    (conversation_id, session["user_id"]),
                )
            
        # 2. Salvar mensagem do usuário
        cursor.execute("""
            INSERT INTO agent_messages (conversation_id, role, content)
            VALUES (%s, 'user', %s)
        """, (conversation_id, user_message))
        
        ai_response = ""

        normalized_message = user_message.lower()
        if normalized_message.startswith('/imagem ') or normalized_message.startswith('/img '):
            ai_response = (
                "No momento a geração de imagens está indisponível. "
                "Posso ajudar com uma descrição detalhada do que você precisa?"
            )
        else:
            # --- CHAT TEXTO (RAG) ---

            user_role = session.get('role', 'user')

            cursor.execute("""
                SELECT question, answer, category 
                FROM agent_knowledge_base 
                WHERE to_tsvector('portuguese', question || ' ' || answer) @@ plainto_tsquery('portuguese', %s)
                AND (allowed_roles IS NULL OR %s = ANY(allowed_roles))
                ORDER BY created_at DESC
                LIMIT 5
            """, (user_message, user_role))

            knowledge_items = cursor.fetchall()
            kb_blocks = []
            if knowledge_items:
                for idx, item in enumerate(knowledge_items, start=1):
                    kb_blocks.append(
                        f"{idx}. [{item['category'] or 'Geral'}] {item['question']}\n{item['answer']}"
                    )
                ai_response = (
                    "📚 Base de Conhecimento encontrada:\n"
                    + "\n\n".join(kb_blocks)
                )
            else:
                context_text = ""

                cursor.execute("""
                    SELECT role, content 
                    FROM agent_messages 
                    WHERE conversation_id = %s 
                    ORDER BY created_at DESC 
                    LIMIT 10
                """, (conversation_id,))
                history = [dict(row) for row in cursor.fetchall()][::-1]

                history_text = ""
                if history:
                    history_lines = []
                    for msg in history:
                        role_label = "Usuário" if msg["role"] == "user" else "Assistente"
                        history_lines.append(f"{role_label}: {msg['content']}")
                    history_text = "\n".join(history_lines)

                rag_response = None
                if RAG_API_URL:
                    try:
                        rag_response = call_rag_service(user_message)
                        ai_response = format_rag_response(rag_response)
                    except requests.HTTPError as rag_http_err:
                        status_code = rag_http_err.response.status_code if rag_http_err.response else "?"
                        error_text = ""
                        try:
                            if rag_http_err.response:
                                error_text = rag_http_err.response.text[:500]
                        except Exception:
                            pass
                        app.logger.error("[RAG] HTTP %s ao processar pergunta: %s - %s", status_code, rag_http_err, error_text)
                        # Se o erro mencionar Gemini/Google, é porque o RAG service ainda está desatualizado
                        if "gemini" in error_text.lower() or "generativelanguage" in error_text.lower() or "429" in str(status_code):
                            ai_response = (
                                "Erro: O serviço RAG local ainda está tentando usar Google Gemini (código desatualizado). "
                                "Por favor, atualize o código do RAG service no Render ou reinicie o serviço local."
                            )
                        else:
                            ai_response = (
                                "No momento não consegui consultar o RAG local (erro HTTP %s). "
                                "Verifique se o agente local está ligado e se o tunnel está ativo."
                            ) % status_code
                        # Evita qualquer fallback externo
                        rag_response = {"answer": ai_response, "sources": []}
                    except Exception as rag_error:
                        error_str = str(rag_error)
                        app.logger.error("[RAG] Falha ao processar pergunta: %s", rag_error)
                        # Se o erro mencionar Gemini/Google, é porque o RAG service ainda está desatualizado
                        if "gemini" in error_str.lower() or "generativelanguage" in error_str.lower():
                            ai_response = (
                                "Erro: O serviço RAG local ainda está tentando usar Google Gemini (código desatualizado). "
                                "Por favor, atualize o código do RAG service no Render ou reinicie o serviço local."
                            )
                        else:
                            ai_response = (
                                "No momento não consegui consultar o RAG local. "
                                "Verifique se o agente local está ligado e se o tunnel está ativo."
                            )
                        rag_response = {"answer": ai_response, "sources": []}

                # Se RAG está configurado e em modo estrito, não use nenhum fallback externo
                if RAG_API_URL and RAG_STRICT_MODE and not rag_response:
                    ai_response = (
                        "O RAG local está configurado, mas não respondeu. "
                        "Tente novamente em instantes ou reinicie o agente local."
                    )
                    rag_response = {"answer": ai_response, "sources": []}

                if not rag_response:
                    # Sem RAG e sem fallback: retorna erro claro.
                    return jsonify({
                        "error": "Serviço RAG indisponível. Configure RAG_API_URL e mantenha o agente local ligado."
                    }), 503

        # 6. Salvar resposta da IA
        cursor.execute("""
            INSERT INTO agent_messages (conversation_id, role, content)
            VALUES (%s, 'assistant', %s)
        """, (conversation_id, ai_response))
        
        conn.commit()
        
        return jsonify({
            "conversation_id": conversation_id,
            "response": ai_response
        })
        
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Erro no chat IA: {e}")
        return jsonify({"error": f"Erro ao processar mensagem: {str(e)}"}), 500
    finally:
        conn.close()


@app.route("/api/agent/knowledge", methods=["GET"])
@login_required
def list_knowledge():
    """Lista itens da base de conhecimento."""
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT id, question, answer, category, created_at 
            FROM agent_knowledge_base 
            ORDER BY created_at DESC
        """)
        items = [dict(row) for row in cursor.fetchall()]
        return jsonify({"items": items})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/knowledge", methods=["POST"])
@login_required
def add_knowledge():
    """Adiciona um novo item à base de conhecimento."""
    # Restrição de permissão
    if session.get("role") != "admin":
        return jsonify({"error": "Apenas administradores podem adicionar conhecimento"}), 403

    data = request.get_json()
    question = data.get('question')
    answer = data.get('answer')
    category = data.get('category', 'Geral')
    
    if not question or not answer:
        return jsonify({"error": "Pergunta e resposta são obrigatórias"}), 400
        
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO agent_knowledge_base (question, answer, category, created_by)
            VALUES (%s, %s, %s, %s)
            RETURNING id
        """, (question, answer, category, session['user_id']))
        new_id = cursor.fetchone()['id']
        conn.commit()
        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/knowledge/<int:item_id>", methods=["DELETE"])
@login_required
def delete_knowledge(item_id):
    """Remove um item da base de conhecimento."""
    # Restrição de permissão
    if session.get("role") != "admin":
        return jsonify({"error": "Apenas administradores podem remover conhecimento"}), 403

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM agent_knowledge_base WHERE id = %s", (item_id,))
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/agent/chat/<int:conversation_id>", methods=["DELETE"])
@login_required
def delete_conversation(conversation_id):
    """Apaga uma conversa."""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            DELETE FROM agent_conversations 
            WHERE id = %s AND user_id = %s
        """, (conversation_id, session['user_id']))
        
        conn.commit()
        return jsonify({"success": True})
        
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/indicadores-executivos", methods=["GET"])
@login_required
def api_indicadores_executivos():
    """
    Endpoint para retornar indicadores executivos por perfil de usuario
    Busca dados do Brudam e calcula indicadores financeiros e operacionais
    """
    try:
        # Parametros da query
        database = request.args.get("database", "azportoex")
        data_inicio = request.args.get("data_inicio")
        data_fim = request.args.get("data_fim")
        custos_dia = request.args.get("custos_dia", type=float)
        
        # Se nao informado periodo, usar mes atual
        hoje = datetime.now().date()
        if not data_inicio:
            data_inicio = date(hoje.year, hoje.month, 1).strftime("%Y-%m-%d")
        if not data_fim:
            data_fim = hoje.strftime("%Y-%m-%d")
        
        # Buscar dados do Brudam
        operacoes = []
        try:
            brudam_conn = get_brudam_db(database)
            cursor = brudam_conn.cursor()
            
            # Query para buscar coletas e entregas com dados financeiros
            query = """
                SELECT 
                    c.id_coleta as id_operacao,
                    c.id_coleta,
                    NULL as id_minuta,
                    'Coleta' as tipo_operacao,
                    c.data as data_operacao,
                    c.data as coleta_data,
                    c.total_nf_valor as valor_nf,
                    c.total_peso as peso,
                    c.total_volumes as volume,
                    c.total_cubo as cubagem,
                    c.status,
                    s.descricao as status_descricao,
                    c.placa,
                    c.motorista,
                    c.responsavel,
                    c.vendedor_nome,
                    c.prev_entrega_data,
                    c.prev_entrega,
                    NULL as entrega_data,
                    serv.nome as servico_nome
                FROM coletas c
                LEFT JOIN minuta_status s ON c.status = s.codigo
                LEFT JOIN servicos serv ON c.id_servico = serv.id_servico
                WHERE c.data BETWEEN %s AND %s
                
                UNION ALL
                
                SELECT 
                    m.id_minuta as id_operacao,
                    m.coleta_numero as id_coleta,
                    m.id_minuta,
                    'Entrega' as tipo_operacao,
                    m.data as data_operacao,
                    c.data as coleta_data,
                    m.total_nf_valor as valor_nf,
                    m.total_peso as peso,
                    m.total_volumes as volume,
                    m.total_cubo as cubagem,
                    m.status,
                    s.descricao as status_descricao,
                    c.placa,
                    c.motorista,
                    c.responsavel,
                    c.vendedor_nome,
                    m.prev_entrega_data,
                    m.prev_entrega,
                    m.data as entrega_data,
                    serv.nome as servico_nome
                FROM minutas m
                LEFT JOIN coletas c ON m.coleta_numero = c.id_coleta
                LEFT JOIN minuta_status s ON m.status = s.codigo
                LEFT JOIN servicos serv ON c.id_servico = serv.id_servico
                WHERE m.data BETWEEN %s AND %s
            """
            
            cursor.execute(query, (data_inicio, data_fim, data_inicio, data_fim))
            rows = cursor.fetchall()
            
            # Converter para lista de dicionarios
            from decimal import Decimal
            for row in rows:
                op = {}
                for key, value in row.items():
                    if isinstance(value, Decimal):
                        op[key] = float(value)
                    elif isinstance(value, (datetime, date)):
                        op[key] = value.isoformat() if isinstance(value, datetime) else str(value)
                    else:
                        op[key] = value
                operacoes.append(op)
            
            brudam_conn.close()
            
        except Exception as e:
            app.logger.error(f"[INDICADORES] Erro ao buscar dados do Brudam: {e}")
            # Retornar estrutura vazia mas valida
            operacoes = []
        
        # Calcular indicadores executivos
        indicadores_exec = montar_indicadores_executivos(operacoes, custos_dia)
        
        # Determinar painel baseado no usuario
        usuario = session.get("user_id")
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT username, email, departamento FROM users_new WHERE id = %s", (usuario,))
        user_data = cursor.fetchone()
        conn.close()
        
        username = (user_data.get("username") or user_data.get("email") or "").lower() if user_data else ""
        departamento = (user_data.get("departamento") or "").lower() if user_data else ""
        
        # Mapear usuario para painel
        panel_key = None
        if "marlon" in username:
            panel_key = "ceo"
        elif "michell" in username:
            panel_key = "diretoria"
        elif "murilo" in username:
            panel_key = "operacional"
        elif "mauro" in username:
            panel_key = "projetos"
        elif departamento in ["financeiro", "comercial", "controladoria"]:
            panel_key = "diretoria"
        elif departamento in ["operacional", "atendimento"]:
            panel_key = "operacional"
        
        # Retornar dados do painel especifico
        panel_map = {
            "ceo": "ceo_panel",
            "diretoria": "diretoria_panel",
            "operacional": "operacional_panel",
            "projetos": "projetos_panel"
        }
        
        panel_id = panel_map.get(panel_key) if panel_key else None
        panel_data = indicadores_exec["indicadores"].get(panel_id) if panel_id else None
        leitura = indicadores_exec["leituras_executivas"].get(panel_key) if panel_key else None

        # Gráficos prontos para o frontend (formato clean: type, title, labels, datasets)
        charts = []
        try:
            dir_panel = indicadores_exec["indicadores"].get("diretoria_panel") or {}
            resultado_por_servico = dir_panel.get("resultado_por_servico_d5")
            if isinstance(resultado_por_servico, dict) and resultado_por_servico:
                items = sorted(
                    resultado_por_servico.items(),
                    key=lambda x: float(x[1].get("faturamento") or 0),
                    reverse=True,
                )[:12]
                charts.append({
                    "id": "faturamento_por_servico",
                    "type": "bar",
                    "title": "Faturamento por serviço",
                    "labels": [k for k, _ in items],
                    "datasets": [{"label": "Faturamento (R$)", "data": [float(v.get("faturamento") or 0) for _, v in items]}],
                })
            pc = dir_panel.get("participacao_corporativo_percent")
            pv = dir_panel.get("participacao_vendedores_percent")
            if pc is not None and pv is not None:
                charts.append({
                    "id": "participacao_faturamento",
                    "type": "pie",
                    "title": "Participação no faturamento",
                    "labels": ["Corporativo", "Vendedores"],
                    "datasets": [{"label": "Participação (%)", "data": [float(pc), float(pv)]}],
                })
        except Exception as chart_err:
            app.logger.warning(f"[INDICADORES] Charts opcionais: {chart_err}")

        return jsonify({
            "success": True,
            "panel_key": panel_key,
            "panel_data": panel_data,
            "leitura_executiva": leitura,
            "charts": charts,
            "indicadores_completos": indicadores_exec,
            "periodo": {
                "inicio": data_inicio,
                "fim": data_fim
            },
            "database": database,
            "total_operacoes": len(operacoes)
        }), 200
        
    except Exception as e:
        app.logger.error(f"[INDICADORES] Erro ao calcular indicadores executivos: {e}", exc_info=True)
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


# --------------------------------------------------------------------------- #
# Tratamento de erros
# --------------------------------------------------------------------------- #
@app.errorhandler(404)
def not_found_error(error):
    return render_template("errors/404.html"), 404


@app.errorhandler(500)
def internal_error(error):
    return render_template("errors/500.html"), 500


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)

